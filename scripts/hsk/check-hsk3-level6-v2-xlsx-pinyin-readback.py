#!/usr/bin/env python3
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path.cwd()
RELEASE_ID = "hsk3_level_6_1800_v2"
DATE = "20260612"
EXPECTED_ROWS = 1800

XLSX_PATH = ROOT / "outputs/hsk/hsk3_level_6_1800_v2.xlsx"
SOURCE_PATH = ROOT / "outputs/hsk/source/hsk3_level_6_1800_v2.source.json"
EXAMPLES_PATH = ROOT / "config/hsk3-level6-v2-examples.json"
REPORT_PATH = ROOT / f"outputs/hsk/qa/{RELEASE_ID}_xlsx_pinyin_readback_{DATE}.json"
SHEET_NAME = "HSK3 Level 6"

HAN_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
TONE_NUMBER_RE = re.compile(r"[A-Za-züÜv:][1-5]\b")
TONE_MARK_RE = re.compile(r"[āáǎàēéěèīíǐìōóǒòūúǔùǖǘǚǜńňǹḿ]", re.IGNORECASE)
PINYIN_LATIN_RE = re.compile(r"[A-Za-züÜāáǎàēéěèīíǐìōóǒòūúǔùǖǘǚǜńňǹḿ]", re.IGNORECASE)


def issue(code, message, row=None, **extra):
    payload = {"code": code, "message": message}
    if row is not None:
        payload["row"] = row
    payload.update(extra)
    return payload


def main():
    source_rows = json.loads(SOURCE_PATH.read_text(encoding="utf-8"))
    examples = json.loads(EXAMPLES_PATH.read_text(encoding="utf-8"))

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Missing sheet: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    required_headers = ["ZH", "ZH transcription", "ZH example", "ZH example transcription"]

    blockers = []
    warnings = []
    for header in required_headers:
        if header not in header_index:
            blockers.append(issue("missing_header", f"Missing XLSX header {header}."))

    rows_checked = 0
    source_pinyin_exact_rows = 0
    example_pinyin_exact_rows = 0
    source_rows_with_tone_marks = 0
    example_rows_with_tone_marks = 0

    if not blockers:
        for excel_row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            source_index = excel_row_number - 2
            if source_index >= len(source_rows):
                blockers.append(issue("extra_xlsx_row", "XLSX has more rows than source.", excel_row_number))
                continue
            source = source_rows[source_index]
            key = source["hsk_key"]
            example = examples.get(key)
            rows_checked += 1

            zh = values[header_index["ZH"] - 1] or ""
            zh_transcription = values[header_index["ZH transcription"] - 1] or ""
            zh_example = values[header_index["ZH example"] - 1] or ""
            zh_example_transcription = values[header_index["ZH example transcription"] - 1] or ""

            if zh != source["simplified"]:
                blockers.append(issue("zh_mismatch", "XLSX ZH does not match official source simplified.", excel_row_number, expected=source["simplified"], actual=zh, hsk_key=key))
            if zh_transcription != source["pinyin"]:
                blockers.append(issue("zh_transcription_mismatch", "XLSX ZH transcription does not match official source pinyin.", excel_row_number, expected=source["pinyin"], actual=zh_transcription, hsk_key=key))
            else:
                source_pinyin_exact_rows += 1
            if not example:
                blockers.append(issue("missing_example", "Missing curated example for XLSX row.", excel_row_number, hsk_key=key))
                continue
            if zh_example != example["example_zh"]:
                blockers.append(issue("zh_example_mismatch", "XLSX ZH example does not match curated example.", excel_row_number, expected=example["example_zh"], actual=zh_example, hsk_key=key))
            if zh_example_transcription != example["example_pinyin"]:
                blockers.append(issue("zh_example_transcription_mismatch", "XLSX ZH example transcription does not match curated example pinyin.", excel_row_number, expected=example["example_pinyin"], actual=zh_example_transcription, hsk_key=key))
            else:
                example_pinyin_exact_rows += 1

            for field, value in [
                ("ZH transcription", zh_transcription),
                ("ZH example transcription", zh_example_transcription),
            ]:
                if not value or not PINYIN_LATIN_RE.search(str(value)):
                    blockers.append(issue("missing_pinyin_letters", f"{field} is blank or has no pinyin letters.", excel_row_number, value=value, hsk_key=key))
                if HAN_RE.search(str(value)):
                    blockers.append(issue("han_in_pinyin", f"{field} contains Han characters.", excel_row_number, value=value, hsk_key=key))
                if TONE_NUMBER_RE.search(str(value)):
                    blockers.append(issue("tone_number_in_pinyin", f"{field} contains tone-number notation.", excel_row_number, value=value, hsk_key=key))
            if TONE_MARK_RE.search(str(zh_transcription)):
                source_rows_with_tone_marks += 1
            if TONE_MARK_RE.search(str(zh_example_transcription)):
                example_rows_with_tone_marks += 1
            else:
                blockers.append(issue("example_pinyin_no_tone_marks", "XLSX ZH example transcription has no tone marks.", excel_row_number, value=zh_example_transcription, hsk_key=key))

    if rows_checked != EXPECTED_ROWS:
        blockers.append(issue("row_count_mismatch", f"Expected {EXPECTED_ROWS} XLSX data rows, checked {rows_checked}."))

    report = {
        "release_id": RELEASE_ID,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": "blocked" if blockers else "ok",
        "xlsx": str(XLSX_PATH.relative_to(ROOT)),
        "sheet": SHEET_NAME,
        "rows_checked": rows_checked,
        "source_pinyin_exact_rows": source_pinyin_exact_rows,
        "example_pinyin_exact_rows": example_pinyin_exact_rows,
        "source_rows_with_tone_marks": source_rows_with_tone_marks,
        "example_rows_with_tone_marks": example_rows_with_tone_marks,
        "blockers": blockers,
        "warnings": warnings,
        "notes": [
            "This readback checks the actual XLSX buyer-facing Chinese transcription columns.",
            "It compares ZH transcription against the official source JSON and ZH example transcription against the curated HSK3 Level 6 v2 examples JSON.",
            "It does not check target-language transcriptions because HSK target-language columns intentionally have no transcription layer.",
        ],
    }
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "release_id": RELEASE_ID,
        "status": report["status"],
        "rows_checked": rows_checked,
        "source_pinyin_exact_rows": source_pinyin_exact_rows,
        "example_pinyin_exact_rows": example_pinyin_exact_rows,
        "blockers": len(blockers),
        "warnings": len(warnings),
        "report": str(REPORT_PATH.relative_to(ROOT)),
    }, ensure_ascii=False, indent=2))
    if blockers:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
