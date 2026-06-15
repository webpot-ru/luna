#!/usr/bin/env python3
import csv
import json
from pathlib import Path

from openpyxl import Workbook


ROOT = Path.cwd()
RELEASE_ID = "hsk3_level_6_1800_v2"
CSV_PATH = ROOT / "outputs/hsk/hsk3_level_6_1800_v2.csv"
XLSX_PATH = ROOT / "outputs/hsk/hsk3_level_6_1800_v2.xlsx"
LANGUAGES_PATH = ROOT / "config/language-order.json"
COURSE_METADATA_PATH = ROOT / "config/hsk3-level6-course-metadata.json"
SOURCE_PATH = ROOT / "outputs/hsk/source/hsk3_level_6_1800_v2.source.json"
REUSE_MAP_PATH = ROOT / "outputs/hsk/qa/hsk3_level_6_1800_v2_classic_reuse_map_20260612.json"
CONTRACT_PATH = ROOT / "config/hsk3-release-contract-v1.json"


def load_json(path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_csv_rows(path):
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader), list(reader.fieldnames or [])


def append_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(header, "") for header in headers])
        else:
            ws.append(row)


def hsk_key(source_row):
    return source_row.get("hsk_key") or f"{source_row['hsk_order']}:{source_row['source_word']}"


def main():
    rows, internal_headers = load_csv_rows(CSV_PATH)
    all_languages = load_json(LANGUAGES_PATH)
    course_metadata = load_json(COURSE_METADATA_PATH)
    source_rows = load_json(SOURCE_PATH)
    reuse_map = load_json(REUSE_MAP_PATH)
    contract = load_json(CONTRACT_PATH)

    configured_codes = [language["spreadsheetCode"] for language in all_languages]
    if "ZH" not in configured_codes:
        raise SystemExit("language-order.json must include ZH for HSK buyer-facing workbook")
    target_codes = [code for code in configured_codes if code != "ZH"]
    non_english_codes = [code for code in target_codes if code not in {"EN", "EN-GB"}]

    wb = Workbook(write_only=True)

    main_sheet = wb.create_sheet("HSK3 Level 6")
    main_headers = [
        "ZH",
        "ZH transcription",
        *target_codes,
        "ZH example",
        "ZH example transcription",
        *[f"{code} example" for code in target_codes],
    ]
    main_sheet.append(main_headers)
    for row in rows:
        main_sheet.append(
            [
                row.get("simplified", ""),
                row.get("pinyin", ""),
                *[row.get(code, "") for code in target_codes],
                row.get("example_zh", ""),
                row.get("example_pinyin", ""),
                *[row.get(f"example_{code}", "") for code in target_codes],
            ]
        )

    internal_sheet = wb.create_sheet("Internal Data")
    append_table(internal_sheet, internal_headers, rows)

    metadata_sheet = wb.create_sheet("Course Metadata")
    metadata_sheet.append(["", *target_codes])
    for field, label in [
        ("title", "Title"),
        ("description", "Description"),
        ("module", "Module"),
        ("category", "Category"),
    ]:
        metadata_sheet.append([label, *[course_metadata.get(code, {}).get(field, "") for code in target_codes]])

    languages_sheet = wb.create_sheet("Languages")
    languages_sheet.append(["order", "spreadsheet_code", "db_code", "language"])
    for index, language in enumerate([language for language in all_languages if language["spreadsheetCode"] != "ZH"], start=1):
        languages_sheet.append([index, language["spreadsheetCode"], language["dbCode"], language["language"]])

    reuse_headers = [
        "hsk3_order",
        "hsk3_source_word",
        "hsk3_simplified",
        "hsk3_pinyin",
        "hsk3_pos",
        "classic_reuse_class",
        "classic_reuse_allowed",
        "classic_reuse_notes",
        "first_classic_release_id",
        "first_classic_level",
        "first_classic_order",
        "first_classic_source_word",
        "first_classic_pinyin",
        "first_classic_en",
    ]
    reuse_sheet = wb.create_sheet("Classic Reuse Map")
    append_table(reuse_sheet, reuse_headers, reuse_map["rows"])

    translation_qa_sheet = wb.create_sheet("Translation QA")
    translation_qa_sheet.append(["order", "spreadsheet_code", "language", "word_filled", "example_filled", "word_missing", "example_missing", "status"])
    for index, language in enumerate([language for language in all_languages if language["spreadsheetCode"] != "ZH"], start=1):
        code = language["spreadsheetCode"]
        filled_words = sum(1 for row in rows if row.get(code))
        filled_examples = sum(1 for row in rows if row.get(f"example_{code}"))
        translation_qa_sheet.append(
            [
                index,
                code,
                language["language"],
                filled_words,
                filled_examples,
                len(rows) - filled_words,
                len(rows) - filled_examples,
                "english_pivot_ready" if code in {"EN", "EN-GB"} else "complete" if filled_words == len(rows) and filled_examples == len(rows) else "manual_hsk3_target_pending",
            ]
        )

    reuse_by_order = {row["hsk3_order"]: row for row in reuse_map["rows"]}
    source_audit_headers = [
        "hsk_order",
        "hsk_key",
        "source_word",
        "simplified",
        "pinyin",
        "source_pos",
        "raw_level",
        "classic_reuse_class",
        "classic_reuse_allowed",
        "first_classic_release_id",
        "first_classic_order",
        "first_classic_source_word",
        "first_classic_pinyin",
    ]
    source_audit_sheet = wb.create_sheet("Source Audit")
    source_audit_sheet.append(source_audit_headers)
    for source_row in source_rows:
        reuse = reuse_by_order.get(source_row["hsk_order"], {})
        source_audit_sheet.append(
            [
                source_row["hsk_order"],
                hsk_key(source_row),
                source_row["source_word"],
                source_row["simplified"],
                source_row["pinyin"],
                source_row.get("pos", ""),
                source_row.get("raw_level", ""),
                reuse.get("classic_reuse_class", ""),
                reuse.get("classic_reuse_allowed", ""),
                reuse.get("first_classic_release_id", ""),
                reuse.get("first_classic_order", ""),
                reuse.get("first_classic_source_word", ""),
                reuse.get("first_classic_pinyin", ""),
            ]
        )

    readme_sheet = wb.create_sheet("README")
    readme_rows = [
        ["LunaCards HSK 3.0 preparation workbook"],
        ["Release ID", RELEASE_ID],
        ["HSK version", "HSK 3.0"],
        ["HSK level", 6],
        ["Expected rows", 1800],
        ["Official source order", "3601-5400"],
        ["Internal language columns", len(all_languages)],
        ["Buyer-facing language columns", "54 including ZH"],
        ["Main sheet", "HSK3 Level 6"],
        ["Status", "complete source/chinese/reused-target/manual target-language preparation"],
        ["Chinese layer", "ready: Chinese word/example and pinyin are shown as ZH, ZH example, ZH transcription and ZH example transcription"],
        ["Official source", contract["source_policy"]["official_page_url"]],
        ["Official PDF", contract["source_policy"]["pdf_url"]],
        ["Local source snapshot", str(SOURCE_PATH.relative_to(ROOT))],
        ["Classic reuse map", str(REUSE_MAP_PATH.relative_to(ROOT))],
        ["Audio", "not in scope"],
        ["Target-language transcription", "not in scope"],
        ["Internal Data", "Technical HSK3 source/reuse/status fields are kept off the buyer-facing first sheet."],
        ["Contract", str(CONTRACT_PATH.relative_to(ROOT))],
        ["Documentation", "docs/hsk-3-release-plan.md"],
    ]
    for readme_row in readme_rows:
        readme_sheet.append(readme_row)

    sources_sheet = wb.create_sheet("Sources")
    sources_sheet.append(["source_id", "role", "url_or_path", "license_or_posture", "notes"])
    for source_row in [
        [
            "official-hsk3-syllabus-pdf",
            "Chinese base list for HSK 3.0 Level 6",
            contract["source_policy"]["pdf_url"],
            "official exam syllabus; use as source of truth for source rows",
            "Local snapshot is stored under outputs/hsk/source/.",
        ],
        [
            "hsk3-level6-v2-source-snapshot",
            "Normalized source rows",
            str(SOURCE_PATH.relative_to(ROOT)),
            "internal LunaCards normalization of official source",
            "Preserves official HSK 3.0 pinyin and row order 3601-5400.",
        ],
        [
            "hsk3-level6-v2-reused-target-translations",
            "Target-language cells copied from corrected Level 6 v2 reused target layer",
            "config/hsk3-level6-v2-reused-target-translations.json",
            "candidate reuse evidence only",
            "Covers legacy Level 6 target rows plus newly allowed corrected v2 reuse rows.",
        ],
        [
            "hsk3-level6-v2-manual-target-translations",
            "Target-language cells authored for corrected Level 6 v2 manual rows",
            "config/hsk3-level6-v2-manual-target-translations.json",
            "manual HSK3 target layer",
            "Complete for the 181 corrected v2 manual rows across all 51 non-English target languages.",
        ],
    ]:
        sources_sheet.append(source_row)

    filled_non_english = sum(1 for row in rows for code in non_english_codes if row.get(code))
    if len(rows) != 1800 or filled_non_english != 91800:
        raise SystemExit(f"Unexpected workbook shape: rows={len(rows)} filled_non_english={filled_non_english}")

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(json.dumps({"release_id": RELEASE_ID, "status": "ok", "xlsx": str(XLSX_PATH.relative_to(ROOT)), "rows": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
