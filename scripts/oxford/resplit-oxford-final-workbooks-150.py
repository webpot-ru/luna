#!/usr/bin/env python3
"""Build 150-row Oxford final workbooks from verified larger Oxford workbooks.

This is a workbook-preserving reslice: it does not regenerate translations,
examples, pronunciations, or support-language batches. It copies current final
US/UK workbook content, slices data rows and Card Metadata, then updates
deck-level metadata for the new 150-word delivery grid.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SCRIPT_VERSION = "2026-06-08.v2"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
FINAL_DIR = PROJECT_ROOT / "outputs/oxford-vocabulary/final"
EDITIONS_DIR = PROJECT_ROOT / "outputs/oxford-vocabulary/editions"
OUT_DIR = PROJECT_ROOT / "outputs/oxford-vocabulary/final-150"


@dataclass(frozen=True)
class SourceWorkbook:
    course_label: str
    course_slug: str
    level: str
    old_part: int
    edition: str
    edition_label: str
    source_variant: str
    primary_language_code: str
    path: Path
    row_count: int


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def normalize_path(path: Path) -> str:
    return str(path.relative_to(PROJECT_ROOT))


def main_sheet_name(wb, level: str, edition: str) -> str:
    suffix = "UK" if edition == "british" else "US"
    expected = f"Oxford {level} {suffix}"
    if expected in wb.sheetnames:
        return expected
    matches = [name for name in wb.sheetnames if name.startswith(f"Oxford {level} ")]
    if not matches:
        raise RuntimeError(f"Cannot find main Oxford sheet for {level}/{edition}: {wb.sheetnames}")
    return matches[0]


def count_data_rows(path: Path, level: str, edition: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheet = wb[main_sheet_name(wb, level, edition)]
    last_non_empty = 0
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value not in (None, ""):
            last_non_empty = row_index
    return max(0, last_non_empty - 1)


KNOWN_SOURCE_ROW_COUNTS = {
    ("oxford_3000_core", "A1", 1): 150,
    ("oxford_3000_core", "A1", 2): 300,
    ("oxford_3000_core", "A1", 3): 300,
    ("oxford_3000_core", "A1", 4): 147,
    ("oxford_3000_core", "A2", 1): 300,
    ("oxford_3000_core", "A2", 2): 300,
    ("oxford_3000_core", "A2", 3): 202,
    ("oxford_3000_core", "B1", 1): 300,
    ("oxford_3000_core", "B1", 2): 300,
    ("oxford_3000_core", "B1", 3): 100,
    ("oxford_3000_core", "B2", 1): 300,
    ("oxford_3000_core", "B2", 2): 300,
    ("oxford_5000_advanced_extension", "B2", 1): 300,
    ("oxford_5000_advanced_extension", "B2", 2): 300,
    ("oxford_5000_advanced_extension", "B2", 3): 98,
    ("oxford_5000_advanced_extension", "C1", 1): 300,
    ("oxford_5000_advanced_extension", "C1", 2): 300,
    ("oxford_5000_advanced_extension", "C1", 3): 300,
    ("oxford_5000_advanced_extension", "C1", 4): 300,
    ("oxford_5000_advanced_extension", "C1", 5): 97,
}


def parse_source_workbook(path: Path) -> SourceWorkbook | None:
    match = re.match(
        r"FlashcardsLuna_Oxford_(3000_Core|5000_Advanced_Extension)_"
        r"([A-Z]\d)_Part_(\d{3})_(US_English|British_English)\.xlsx$",
        path.name,
    )
    if not match:
        return None
    course_file, level, part_text, edition_file = match.groups()
    course_label = "Oxford 3000 Core" if course_file == "3000_Core" else "Oxford 5000 Advanced Extension"
    course_slug = "oxford_3000_core" if course_file == "3000_Core" else "oxford_5000_advanced_extension"
    edition = "british" if edition_file == "British_English" else "us"
    old_part = int(part_text)
    row_count = KNOWN_SOURCE_ROW_COUNTS.get((course_slug, level, old_part))
    if row_count is None:
        row_count = count_data_rows(path, level, edition)
    return SourceWorkbook(
        course_label=course_label,
        course_slug=course_slug,
        level=level,
        old_part=old_part,
        edition=edition,
        edition_label="British English" if edition == "british" else "US English",
        source_variant="EN-GB" if edition == "british" else "EN-US",
        primary_language_code="EN-GB" if edition == "british" else "EN",
        path=path,
        row_count=row_count,
    )


def source_workbooks() -> list[SourceWorkbook]:
    paths = sorted(FINAL_DIR.glob("FlashcardsLuna_Oxford_*_*English.xlsx"))
    paths.extend(sorted(EDITIONS_DIR.glob("FlashcardsLuna_Oxford_3000_Core_B1_Part_003_*English.xlsx")))
    rows: list[SourceWorkbook] = []
    for path in paths:
        parsed = parse_source_workbook(path)
        if parsed:
            rows.append(parsed)
    return sorted(rows, key=lambda item: (item.course_slug, item.level, item.edition, item.old_part))


def split_plan(sources: Iterable[SourceWorkbook], chunk_size: int) -> list[dict]:
    planned: list[dict] = []
    by_key: dict[tuple[str, str, str], list[SourceWorkbook]] = {}
    for source in sources:
        by_key.setdefault((source.course_slug, source.level, source.edition), []).append(source)

    for (course_slug, level, edition), group in sorted(by_key.items()):
        next_part = 1
        for source in sorted(group, key=lambda item: item.old_part):
            for offset in range(0, source.row_count, chunk_size):
                size = min(chunk_size, source.row_count - offset)
                planned.append(
                    {
                        "course_slug": course_slug,
                        "course_label": source.course_label,
                        "level": level,
                        "edition": edition,
                        "edition_label": source.edition_label,
                        "source_variant": source.source_variant,
                        "primary_language_code": source.primary_language_code,
                        "new_part": next_part,
                        "row_count": size,
                        "source_path": source.path,
                        "source_old_part": source.old_part,
                        "source_row_start": offset + 1,
                        "source_row_end": offset + size,
                    }
                )
                next_part += 1
    return planned


def set_metadata_value(ws, field: str, value):
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row=row_index, column=1).value == field:
            ws.cell(row=row_index, column=2).value = value
            return


def replace_part_number(value, old_part: int, new_part: int):
    if not isinstance(value, str):
        return value
    return re.sub(rf"(?<!\w){old_part}(?!\w)", str(new_part), value)


def replace_text(value, replacements: dict[str, str]):
    if not isinstance(value, str):
        return value
    next_value = value
    for old, new in replacements.items():
        next_value = next_value.replace(old, new)
    return next_value


def delete_data_outside_slice(ws, start: int, end: int):
    first_excel_row = start + 1
    last_excel_row = end + 1
    if ws.max_row > last_excel_row:
        ws.delete_rows(last_excel_row + 1, ws.max_row - last_excel_row)
    if first_excel_row > 2:
        ws.delete_rows(2, first_excel_row - 2)


def refresh_card_metadata(ws, new_set_id: str):
    for row_index in range(2, ws.max_row + 1):
        display_order = row_index - 1
        meaning_id = ws.cell(row=row_index, column=5).value
        ws.cell(row=row_index, column=1).value = row_index
        ws.cell(row=row_index, column=2).value = display_order
        if meaning_id:
            ws.cell(row=row_index, column=3).value = f"{new_set_id}::{meaning_id}"
        ws.cell(row=row_index, column=4).value = new_set_id


def update_repeated_text(wb, replacements: dict[str, str], old_part: int, new_part: int):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = replace_text(cell.value, replacements)
                value = replace_part_number(value, old_part, new_part)
                cell.value = value


def safe_slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def copy_workbook_slice(item: dict, out_path: Path, chunk_size: int) -> dict:
    source_path = item["source_path"]
    source_start = int(item["source_row_start"])
    source_end = int(item["source_row_end"])
    old_part = int(item["source_old_part"])
    new_part = int(item["new_part"])
    row_count = int(item["row_count"])
    level = item["level"]
    edition = item["edition"]
    edition_label = item["edition_label"]
    source_variant = item["source_variant"]
    primary_language_code = item["primary_language_code"]
    course_slug = item["course_slug"]
    course_label = item["course_label"]
    course_display = "Oxford 3000 Core" if course_slug == "oxford_3000_core" else "Oxford 5000 Advanced Extension"
    module_en = "Oxford 3000" if course_slug == "oxford_3000_core" else "Oxford 5000"
    part_text = f"Part {new_part:03d}"
    set_id = f"{course_slug}_{level.lower()}_part_{new_part:03d}_150_{edition}_english"
    title = f"FlashcardsLuna {course_display} {level} {part_text} (150 words) - {edition_label}"
    slug = safe_slug(title)
    main_sheet_suffix = "UK" if edition == "british" else "US"
    main_sheet = f"Oxford {level} {main_sheet_suffix}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, out_path)
    wb = load_workbook(out_path)

    source_main = main_sheet_name(wb, level, edition)
    if source_main != main_sheet:
        wb[source_main].title = main_sheet
    delete_data_outside_slice(wb[main_sheet], source_start, source_end)
    delete_data_outside_slice(wb["Card Metadata"], source_start, source_end)
    refresh_card_metadata(wb["Card Metadata"], set_id)

    replacements = {
        f"part_{old_part:03d}": f"part_{new_part:03d}",
        f"part-{old_part:03d}": f"part-{new_part:03d}",
        f"Part_{old_part:03d}": f"Part_{new_part:03d}",
        f"Part {old_part:03d}": f"Part {new_part:03d}",
        f"Part {old_part}": f"Part {new_part}",
        f"part {old_part}": f"part {new_part}",
    }
    update_repeated_text(wb, replacements, old_part, new_part)

    deck = wb["Deck Metadata"]
    set_metadata_value(deck, "set_id", set_id)
    set_metadata_value(deck, "slug", slug)
    set_metadata_value(deck, "target_item_count_min", row_count)
    set_metadata_value(deck, "target_item_count_max", chunk_size)
    set_metadata_value(deck, "actual_item_count", row_count)
    set_metadata_value(deck, "title_en", f"Oxford {level} Part {new_part} {main_sheet_suffix}.")
    set_metadata_value(
        deck,
        "description_en",
        f"{course_display} words. Part {new_part}. {edition_label}.",
    )
    set_metadata_value(deck, "module_en", module_en)
    set_metadata_value(deck, "area", level)
    set_metadata_value(deck, "category_en", f"{level} Part {new_part}.")
    set_metadata_value(deck, "learning_goal", f"Study {course_display} {level} vocabulary with {edition_label} as the primary English edition.")
    set_metadata_value(deck, "main_sheet", main_sheet)
    set_metadata_value(deck, "source_variant", source_variant)
    set_metadata_value(deck, "primary_language_code", primary_language_code)

    for sheet_name in ["_README", "_qa_status"]:
        ws = wb[sheet_name]
        for row_index in range(1, ws.max_row + 1):
            label = ws.cell(row=row_index, column=1).value
            if label in {"Set ID"}:
                ws.cell(row=row_index, column=2).value = set_id
            elif label == "Rows":
                ws.cell(row=row_index, column=2).value = row_count

    # Keep localized metadata intact, but correct the visible part number.
    for ws_name in ["Course Metadata"]:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = replace_part_number(cell.value, old_part, new_part)

    wb.save(out_path)
    return {
        "release_id": f"{course_slug}_{level.lower()}_part_{new_part:03d}_150_v2",
        "set_id": set_id,
        "title": title,
        "workbook_path": normalize_path(out_path),
        "workbook_sha256": sha256_file(out_path),
        "row_count": row_count,
        "source_workbook": normalize_path(source_path),
        "source_old_part": old_part,
        "source_rows": [source_start, source_end],
        "main_sheet": main_sheet,
        "edition": edition_label,
    }


def output_name(item: dict) -> str:
    course_file = "Oxford_3000_Core" if item["course_slug"] == "oxford_3000_core" else "Oxford_5000_Advanced_Extension"
    edition_file = "British_English" if item["edition"] == "british" else "US_English"
    return (
        f"FlashcardsLuna_{course_file}_{item['level']}_"
        f"Part_{int(item['new_part']):03d}_150_{edition_file}.xlsx"
    )


def write_reports(records: list[dict], chunk_size: int, out_dir: Path) -> None:
    manifest_path = out_dir / "oxford_150_resplit_manifest_v1.json"
    md_path = out_dir / "oxford_150_resplit_manifest_v1.md"
    by_release: dict[str, dict] = {}
    for record in records:
        key = record["release_id"]
        by_release.setdefault(
            key,
            {
                "release_id": key,
                "title_base": record["title"].rsplit(" - ", 1)[0],
                "row_count": record["row_count"],
                "editions": [],
            },
        )["editions"].append(
            {
                "edition": record["edition"],
                "workbook_path": record["workbook_path"],
                "workbook_sha256": record["workbook_sha256"],
                "source_workbook": record["source_workbook"],
                "source_rows": record["source_rows"],
            }
        )

    payload = {
        "manifest_id": "oxford_150_resplit_manifest_v1",
        "script_path": "scripts/oxford/resplit-oxford-final-workbooks-150.py",
        "script_version": SCRIPT_VERSION,
        "generated_at": utc_now(),
        "status": "local_workbooks_built_needs_google_sheet_upload_readback",
        "chunk_size": chunk_size,
        "principle": "Reslice current verified Oxford final workbooks into learner-smaller 150-row workbooks without regenerating language content.",
        "old_google_sheets_policy": "Do not overwrite old verified 300-row Google Sheets; upload new 150-row native Sheets and mark active only after readback.",
        "release_count": len(by_release),
        "workbook_count": len(records),
        "total_rows_by_edition": {
            edition: sum(record["row_count"] for record in records if record["edition"] == edition)
            for edition in sorted({record["edition"] for record in records})
        },
        "releases": list(by_release.values()),
    }
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        "# Oxford 150 Resplit Manifest",
        "",
        f"- Status: `{payload['status']}`",
        f"- Generated at: `{payload['generated_at']}`",
        f"- Release count: {payload['release_count']}",
        f"- Workbook count: {payload['workbook_count']}",
        f"- Chunk size: {chunk_size}",
        "",
        "| Release | Rows | Editions |",
        "| --- | ---: | --- |",
    ]
    for release in payload["releases"]:
        edition_text = "; ".join(f"{edition['edition']}: `{edition['workbook_path']}`" for edition in release["editions"])
        lines.append(f"| `{release['release_id']}` | {release['row_count']} | {edition_text} |")
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--chunk-size", type=int, default=150)
    parser.add_argument("--out-dir", default=str(OUT_DIR))
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    sources = source_workbooks()
    if not sources:
        raise RuntimeError("No Oxford source workbooks found.")

    plan = split_plan(sources, args.chunk_size)
    records: list[dict] = []
    for item in plan:
        out_path = out_dir / output_name(item)
        records.append(copy_workbook_slice(item, out_path, args.chunk_size))

    write_reports(records, args.chunk_size, out_dir)
    print(
        json.dumps(
            {
                "status": "built",
                "workbook_count": len(records),
                "release_count": len({record["release_id"] for record in records}),
                "manifest": normalize_path(out_dir / "oxford_150_resplit_manifest_v1.json"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
