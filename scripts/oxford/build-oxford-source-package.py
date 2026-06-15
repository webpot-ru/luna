#!/usr/bin/env python3
import argparse
import csv
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfReader


POS_MARKERS = [
    "indefinite article",
    "definite article",
    "modal verb",
    "auxiliary verb",
    "ordinal number",
    "number",
    "det.",
    "pron.",
    "prep.",
    "adv.",
    "adj.",
    "n.",
    "v.",
    "conj.",
    "exclam.",
    "prefix",
    "suffix",
]

LEVEL_RE = re.compile(r"\b(A1|A2|B1|B2|C1|C2)\b")

RAW_LINE_OVERRIDES = {
    # PDF text extraction drops the leading "s" on this Oxford 5000 row.
    "ecular adj. C1": "secular adj. C1",
    # The source line combines a B2 noun with a C1 adjective; this C1 release
    # needs the adjective row without carrying the B2 noun marker into headword.
    "terminal n B2, adj. C1": "terminal adj. C1",
}


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def safe_id(value):
    return re.sub(r"[^a-z0-9]+", "_", normalize_text(value).lower()).strip("_")


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_pos_start(line):
    lowered = line.lower()
    candidates = []
    for marker in POS_MARKERS:
        index = lowered.find(marker)
        if index > 0:
            candidates.append(index)
    return min(candidates) if candidates else -1


def parse_entry(line, order, source_id, source_label):
    text = normalize_text(line)
    if re.match(r"^(?:auxiliary\s+v\.|modal\s+v\.|det\.|pron\.|prep\.|adv\.|adj\.|n\.|v\.|conj\.|exclam\.)\s", text.lower()):
        return None
    if not LEVEL_RE.search(text):
        return None
    pos_start = find_pos_start(text)
    if pos_start <= 0:
        return None

    headword = normalize_text(text[:pos_start])
    pos_level_text = normalize_text(text[pos_start:])
    levels = LEVEL_RE.findall(pos_level_text)
    if not headword or not levels:
        return None

    pos_text = normalize_text(LEVEL_RE.sub("", pos_level_text).replace(" ,", ","))
    return {
        "source_id": source_id,
        "source_candidate_id": f"{source_id}::{order:04d}::{safe_id(headword)}",
        "oxford_order": order,
        "source_headword": headword,
        "normalized_headword": normalize_text(headword).lower(),
        "part_of_speech": pos_text,
        "cefr_levels": "|".join(sorted(set(levels), key=["A1", "A2", "B1", "B2", "C1", "C2"].index)),
        "level_min": sorted(set(levels), key=["A1", "A2", "B1", "B2", "C1", "C2"].index)[0],
        "level_max": sorted(set(levels), key=["A1", "A2", "B1", "B2", "C1", "C2"].index)[-1],
        "source_status": "user_reported_oup_permission_pending_written_evidence",
        "source_support": f"{source_label} PDF source row parsed from user-authorized Oxford source package.",
        "raw_line": text,
    }


def extract_rows(pdf_path, source_id="oxford_3000", source_label="Oxford 3000"):
    reader = PdfReader(str(pdf_path))
    rows = []
    seen = set()
    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = normalize_text(raw_line)
            line = RAW_LINE_OVERRIDES.get(line, line)
            if not line:
                continue
            if line.startswith("© Oxford University Press"):
                continue
            if line.startswith("The Oxford 3000") or line.startswith("The Oxford 5000"):
                continue
            if "most important words" in line:
                continue
            entry = parse_entry(line, len(rows) + 1, source_id, source_label)
            if not entry:
                continue
            key = (entry["normalized_headword"], entry["part_of_speech"], entry["cefr_levels"])
            if key in seen:
                continue
            seen.add(key)
            entry["source_page"] = page_index
            rows.append(entry)
    return rows


def write_jsonl(path, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with Path(path).open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "source_candidate_id",
        "oxford_order",
        "source_headword",
        "part_of_speech",
        "cefr_levels",
        "level_min",
        "level_max",
        "source_page",
        "source_status",
        "raw_line",
    ]
    with Path(path).open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def load_language_codes(path):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return [row["spreadsheetCode"] for row in data]


def make_candidate(
    row,
    index,
    release_id,
    course_id,
    source_variant,
    core_item_prefix,
    meaning_id_prefix,
    core_band,
    benchmark_membership,
    topic_context,
    source_label,
):
    headword_id = safe_id(row["source_headword"])
    pos_id = safe_id(row["part_of_speech"])
    return {
        "release_id": release_id,
        "course_id": course_id,
        "row_id": f"{release_id}::{index:03d}",
        "core_item_id": f"{core_item_prefix}_{index:04d}",
        "meaning_id": f"{meaning_id_prefix}_{headword_id}_{pos_id}_needs_review_01",
        "source_candidate_id": row["source_candidate_id"],
        "source_language": "EN",
        "source_variant": source_variant,
        "source_headword": row["source_headword"],
        "normalized_headword": row["normalized_headword"],
        "part_of_speech": row["part_of_speech"],
        "sense_no": "01_needs_review",
        "core_band": core_band,
        "level_min": row["level_min"],
        "level_max": row["level_max"],
        "oxford_order": row["oxford_order"],
        "benchmark_membership": benchmark_membership,
        "source_status": row["source_status"],
        "meaning_note": f"Needs row review: {source_label} source word '{row['source_headword']}'.",
        "semantic_scene": {
            "rule_version": "oxford-source-package-v0",
            "status": "needs_review",
            "target_display": row["source_headword"],
            "topic_context": topic_context,
        },
        "example_EN": "",
        "example_status": "needs_lunacards_authored_example",
        "qa_status": "source_snapshot_ready_needs_translation",
        "qa_notes": "Oxford source row selected; translations/examples/QA not generated yet.",
        "selection_decision": "selected",
        "required_next_gates": "written_permission_evidence,row_review,examples,translations,source_assisted_qa,workbook_delivery",
    }


def write_workbook(path, release_id, selected_rows, language_codes, manifest, main_sheet_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = main_sheet_name

    fixed = [
        "release_id",
        "course_id",
        "row_id",
        "core_item_id",
        "meaning_id",
        "source_language",
        "source_variant",
        "source_headword",
        "part_of_speech",
        "sense_no",
        "core_band",
        "level_min",
        "level_max",
        "oxford_order",
        "benchmark_membership",
        "source_status",
        "meaning_note",
        "semantic_scene",
    ]
    headers = fixed + language_codes + [f"example_{code}" for code in language_codes] + ["qa_status", "qa_notes"]
    sheet.append(headers)
    for row in selected_rows:
        record = dict(row)
        record["semantic_scene"] = json.dumps(row["semantic_scene"], ensure_ascii=False)
        for code in language_codes:
            record[code] = row["source_headword"] if code in {"EN", "EN-GB"} else ""
            record[f"example_{code}"] = ""
        sheet.append([record.get(header, "") for header in headers])

    readme = workbook.create_sheet("_README")
    for item in [
        ("Release ID", release_id),
        ("Status", "source_snapshot_ready_needs_translation"),
        ("Rows", len(selected_rows)),
        ("Oxford source data", "User reported OUP permission; written evidence still pending in docs."),
        ("Postgres changes", "false"),
        ("Google Sheet created", "false"),
        ("Definitions/examples/pronunciations copied", "false"),
    ]:
        readme.append(item)

    source_sheet = workbook.create_sheet("_source_contract")
    for key, value in manifest.items():
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        source_sheet.append([key, value])

    langs = workbook.create_sheet("_languages")
    langs.append(["spreadsheetCode"])
    for code in language_codes:
        langs.append([code])

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", default="outputs/oxford-vocabulary/source/raw/The_Oxford_3000.pdf")
    parser.add_argument("--languages", default="config/language-order.json")
    parser.add_argument("--release", default="oxford_3000_core_a1_part_001_150_v1")
    parser.add_argument("--course-id", default="oxford_3000_core")
    parser.add_argument("--source-id", default="oxford_3000")
    parser.add_argument("--source-label", default="Oxford 3000")
    parser.add_argument(
        "--source-url",
        default="https://www.oxfordlearnersdictionaries.com/us/external/pdf/wordlists/oxford-3000-5000/The_Oxford_3000.pdf",
    )
    parser.add_argument("--normalized-slug", default="oxford_3000")
    parser.add_argument("--core-item-prefix", default="oxford3000")
    parser.add_argument("--meaning-id-prefix", default="oxford3000")
    parser.add_argument("--core-band", default="oxford_3000")
    parser.add_argument("--benchmark-membership", default="oxford_3000")
    parser.add_argument("--topic-context", default="general_english_core_vocabulary")
    parser.add_argument("--source-variant", default="US English default over Oxford 3000 British/American source list")
    parser.add_argument("--levels", default="A1")
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--out-dir", default="outputs/oxford-vocabulary")
    parser.add_argument("--limit", type=int, default=150)
    parser.add_argument(
        "--workbook-name",
        default="FlashcardsLuna_Oxford_3000_Core_A1_Part_001_source_draft.xlsx",
    )
    parser.add_argument("--main-sheet-name", default="Oxford 3000 Core")
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    out_dir = Path(args.out_dir)
    rows = extract_rows(pdf_path, args.source_id, args.source_label)
    selected_levels = {level.strip() for level in args.levels.split(",") if level.strip()}
    if not selected_levels:
        raise ValueError("--levels must contain at least one CEFR level")
    level_rows = [row for row in rows if row["level_min"] in selected_levels]
    if args.offset < 0:
        raise ValueError("--offset must be >= 0")
    selected_source_rows = level_rows[args.offset : args.offset + args.limit]
    selected = [
        make_candidate(
            row,
            index + 1,
            args.release,
            args.course_id,
            args.source_variant,
            args.core_item_prefix,
            args.meaning_id_prefix,
            args.core_band,
            args.benchmark_membership,
            args.topic_context,
            args.source_label,
        )
        for index, row in enumerate(selected_source_rows)
    ]

    generated_at = datetime.now(timezone.utc).isoformat()
    normalized_path = out_dir / "source" / f"{args.release}.{args.normalized_slug}.normalized.jsonl"
    csv_path = out_dir / "source" / f"{args.release}.{args.normalized_slug}.normalized.csv"
    selected_path = out_dir / "candidate-pools" / f"{args.release}_candidate_pool.jsonl"
    summary_path = out_dir / "candidate-pools" / f"{args.release}_candidate_pool_summary.md"
    workbook_path = out_dir / "final" / args.workbook_name
    manifest_path = out_dir / "source" / f"{args.release}_source_snapshot_manifest.json"

    write_jsonl(normalized_path, rows)
    write_csv(csv_path, rows)
    write_jsonl(selected_path, selected)

    manifest = {
        "release_id": args.release,
        "course_id": args.course_id,
        "generated_at": generated_at,
        "source_id": args.source_id,
        "source_label": args.source_label,
        "source_url": args.source_url,
        "benchmark_membership": args.benchmark_membership,
        "source_path": str(pdf_path),
        "raw_file_sha256": sha256_file(pdf_path),
        "normalized_path": str(normalized_path),
        "normalized_rows": len(rows),
        "level_filter": sorted(selected_levels),
        "level_filtered_rows": len(level_rows),
        "offset": args.offset,
        "limit": args.limit,
        "selected_rows": len(selected),
        "selected_scope": f"{args.source_label} rows where level_min is one of {', '.join(sorted(selected_levels))}, offset {args.offset}, limit {args.limit}",
        "permission_status": "user_reported_all_allowed_pending_written_evidence",
        "approved_for_postgres_import": False,
        "google_sheet_created": False,
        "copied_definitions_examples_pronunciations": False,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary_path.write_text(
        "\n".join(
            [
                f"# Oxford Candidate Pool: {args.release}",
                "",
                f"- Normalized Oxford rows: {len(rows)}",
                f"- Level filter: {', '.join(sorted(selected_levels))}",
                f"- Level-filtered rows: {len(level_rows)}",
                f"- Offset: {args.offset}",
                f"- Limit: {args.limit}",
                f"- Selected rows: {len(selected)}",
                f"- Scope: {args.source_label} rows where level_min is one of {', '.join(sorted(selected_levels))}, offset {args.offset}, limit {args.limit}.",
                "- Status: source snapshot ready; translations/examples/QA not generated yet.",
                "- Permission: user reported all allowed; written OUP evidence still pending.",
                f"- Workbook: `{workbook_path}`",
                "",
            ]
        ),
        encoding="utf-8",
    )

    write_workbook(workbook_path, args.release, selected, load_language_codes(args.languages), manifest, args.main_sheet_name)
    print(f"Oxford source package built: rows={len(rows)} selected={len(selected)} workbook={workbook_path}")


if __name__ == "__main__":
    main()
