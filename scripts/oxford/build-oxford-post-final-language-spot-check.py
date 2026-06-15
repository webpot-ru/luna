#!/usr/bin/env python3
"""Build a post-final 5-per-language spot check for an Oxford release."""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONTRACT_PATH = PROJECT_ROOT / "config/oxford_3000_core_b1_part_002_300_v1_contract_v0.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs/oxford-vocabulary/qa"
REPORT_ID = "post_final_language_spot_check_v1"
SCRIPT_VERSION = "2026-05-31.v2"
SENTENCE_END_RE = re.compile(r"[.!?。！？؟।॥။։។៕]$")
IPA_RE = re.compile(r"^/[^/]+/(?:,\s*/[^/]+/)*$")

HARD_SCRIPT_REQUIREMENTS = {
    "ZH": re.compile(r"[\u3400-\u9FFF]"),
    "JA": re.compile(r"[\u3040-\u30FF\u3400-\u9FFF]"),
    "KO": re.compile(r"[\uAC00-\uD7AF]"),
    "TH": re.compile(r"[\u0E00-\u0E7F]"),
    "HI": re.compile(r"[\u0900-\u097F]"),
    "NE": re.compile(r"[\u0900-\u097F]"),
    "BN": re.compile(r"[\u0980-\u09FF]"),
    "MY": re.compile(r"[\u1000-\u109F]"),
    "KM": re.compile(r"[\u1780-\u17FF]"),
    "LO": re.compile(r"[\u0E80-\u0EFF]"),
    "SI": re.compile(r"[\u0D80-\u0DFF]"),
    "TA": re.compile(r"[\u0B80-\u0BFF]"),
    "TE": re.compile(r"[\u0C00-\u0C7F]"),
    "KN": re.compile(r"[\u0C80-\u0CFF]"),
    "ML": re.compile(r"[\u0D00-\u0D7F]"),
    "KA": re.compile(r"[\u10A0-\u10FF]"),
    "HY": re.compile(r"[\u0530-\u058F]"),
    "RU": re.compile(r"[\u0400-\u04FF]"),
    "BG": re.compile(r"[\u0400-\u04FF]"),
    "SR": re.compile(r"[\u0400-\u04FF]"),
    "KK": re.compile(r"[\u0400-\u04FF]"),
}


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def markdown_cell(value: object) -> str:
    return normalize_text(value).replace("|", "\\|")


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_jsonl(path: Path) -> list[dict]:
    rows = []
    for index, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as error:
            raise ValueError(f"Invalid JSONL at {path}:{index}: {error}") from error
    return rows


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(PROJECT_ROOT))


def by_row_id(rows: list[dict]) -> dict[str, dict]:
    return {row["row_id"]: row for row in rows}


def support_key(row_id: str, language_code: str) -> str:
    return f"{row_id}::{language_code}"


def row_position_from_row_id(row_id: str) -> int | None:
    match = re.search(r"::(\d+)$", row_id)
    return int(match.group(1)) if match else None


def support_transcription_field_is_forbidden(value: object) -> bool:
    if value is False or value is None:
        return False
    normalized = normalize_text(value).lower()
    return normalized not in {"", "false", "not_included_for_support_language", "excluded_for_support_language"}


def spread_sample(rows: list[dict], sample_size: int) -> list[dict]:
    if sample_size <= 0:
        raise ValueError("--sample-size must be positive")
    if sample_size >= len(rows):
        return rows
    if sample_size == 1:
        return [rows[0]]
    indexes = []
    for position in range(sample_size):
        index = int(position * (len(rows) - 1) / (sample_size - 1))
        if index not in indexes:
            indexes.append(index)
    return [rows[index] for index in indexes]


def load_workbook_rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(cell) for cell in rows[0]]
    records = []
    for raw in rows[1:]:
        records.append({headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))})
    return records


def build_expected_values(contract: dict, sample_rows: list[dict]) -> dict[str, dict[str, dict[str, str]]]:
    sample_row_ids = {row["row_id"] for row in sample_rows}
    expected = {row["row_id"]: {} for row in sample_rows}

    edition_layer_path = (
        contract.get("latest_edition_exports", {}).get("edition_layer_path")
        or contract.get("latest_edition_layer", {}).get("path")
    )
    if not edition_layer_path:
        raise ValueError("Contract is missing latest_edition_exports.edition_layer_path or latest_edition_layer.path")
    for row in read_jsonl(PROJECT_ROOT / edition_layer_path):
        row_id = row.get("row_id")
        if row_id not in sample_row_ids:
            continue
        expected[row_id]["EN"] = {
            "display": normalize_text(row.get("display_headword_EN_US")),
            "example": normalize_text(row.get("example_EN_US")),
        }
        expected[row_id]["EN-GB"] = {
            "display": normalize_text(row.get("display_headword_EN_GB")),
            "example": normalize_text(row.get("example_EN_GB")),
        }

    for batch in contract.get("latest_support_translation_batches", []):
        for row in read_jsonl(PROJECT_ROOT / batch["path"]):
            row_id = row.get("row_id")
            if row_id not in sample_row_ids:
                continue
            for language_code in batch.get("languages", []):
                expected[row_id][language_code] = {
                    "display": normalize_text(row.get(language_code)),
                    "example": normalize_text(row.get(f"example_{language_code}")),
                }

    return expected


def build_support_rows_by_language(contract: dict) -> dict[str, dict[str, dict]]:
    rows_by_language: dict[str, dict[str, dict]] = {}
    for batch in contract.get("latest_support_translation_batches", []):
        rows_by_id = by_row_id(read_jsonl(PROJECT_ROOT / batch["path"]))
        for language_code in batch.get("languages", []):
            rows_by_language[language_code] = rows_by_id
    return rows_by_language


def build_audit_indexes(contract: dict) -> tuple[dict[str, dict], dict[str, dict], dict[str, int]]:
    source_audit_path = contract.get("latest_support_translation_source_backed_audit", {}).get("path")
    example_audit_path = contract.get("latest_support_example_quality_audit", {}).get("path")
    source_payload = read_json(PROJECT_ROOT / source_audit_path) if source_audit_path else {"cells": []}
    example_payload = read_json(PROJECT_ROOT / example_audit_path) if example_audit_path else {"cells": []}
    source_by_cell = {
        support_key(cell["row_id"], cell["language_code"]): cell for cell in source_payload.get("cells", [])
    }
    example_by_cell = {
        support_key(cell["row_id"], cell["language_code"]): cell for cell in example_payload.get("cells", [])
    }
    return source_by_cell, example_by_cell, {
        "source_audit_total_cells": len(source_by_cell),
        "example_audit_total_cells": len(example_by_cell),
    }


def check_transcription_rows(
    sample_rows: list[dict],
    expected_values: dict[str, dict[str, dict[str, str]]],
    en_us_by_id: dict[str, dict],
    en_gb_by_id: dict[str, dict],
) -> tuple[list[dict], list[dict]]:
    blockers = []
    entries = []
    variants = [
        ("EN", "EN-US", en_us_by_id, "transcription_EN", "example_transcription_EN", "example_EN"),
        ("EN-GB", "EN-GB", en_gb_by_id, "transcription_EN_GB", "example_transcription_EN_GB", "example_EN_GB"),
    ]
    for source_row in sample_rows:
        row_id = source_row["row_id"]
        for language_code, variant, rows_by_id, transcription_field, example_transcription_field, example_field in variants:
            row = rows_by_id.get(row_id)
            expected_display = expected_values[row_id][language_code]["display"]
            expected_example = expected_values[row_id][language_code]["example"]
            entry_blockers = []
            if not row:
                entry_blockers.append("missing_pronunciation_row")
            else:
                transcription = normalize_text(row.get(transcription_field) or row.get("transcription"))
                example_transcription = normalize_text(
                    row.get(example_transcription_field) or row.get("example_transcription")
                )
                if normalize_text(row.get("reviewed_display_headword")) != expected_display:
                    entry_blockers.append("display_mismatch_against_edition_layer")
                if normalize_text(row.get(example_field) or row.get("example")) != expected_example:
                    entry_blockers.append("example_mismatch_against_edition_layer")
                if not transcription or not IPA_RE.match(transcription):
                    entry_blockers.append("missing_or_invalid_display_ipa")
                if not example_transcription or not IPA_RE.match(example_transcription):
                    entry_blockers.append("missing_or_invalid_example_ipa")
                if not normalize_text(row.get("transcription_status")).startswith("source_backed_"):
                    entry_blockers.append("display_transcription_not_source_backed")
                if not normalize_text(row.get("example_transcription_status")).startswith("source_backed_"):
                    entry_blockers.append("example_transcription_not_source_backed")
                if row.get("does_not_use_oxford_pronunciation") is not True:
                    entry_blockers.append("oxford_pronunciation_reuse_not_explicitly_denied")
                if not row.get("source_ids"):
                    entry_blockers.append("missing_pronunciation_source_ids")
            if entry_blockers:
                blockers.append(
                    {
                        "row_id": row_id,
                        "source_headword": source_row["source_headword"],
                        "language_code": language_code,
                        "variant": variant,
                        "reasons": entry_blockers,
                    }
                )
            entries.append(
                {
                    "row_id": row_id,
                    "source_headword": source_row["source_headword"],
                    "language_code": language_code,
                    "variant": variant,
                    "transcription": normalize_text((row or {}).get(transcription_field) or (row or {}).get("transcription")),
                    "example_transcription": normalize_text(
                        (row or {}).get(example_transcription_field) or (row or {}).get("example_transcription")
                    ),
                    "transcription_status": normalize_text((row or {}).get("transcription_status")),
                    "example_transcription_status": normalize_text((row or {}).get("example_transcription_status")),
                    "status": "pass" if not entry_blockers else "blocker",
                }
            )
    return entries, blockers


def expected_pronunciation_row(edition: dict, row_id: str, en_us_by_id: dict[str, dict], en_gb_by_id: dict[str, dict]) -> dict:
    return en_us_by_id[row_id] if edition.get("primary_language_code") == "EN" else en_gb_by_id[row_id]


def check_final_workbooks(
    manifest: dict,
    sample_rows: list[dict],
    expected_values: dict[str, dict[str, dict[str, str]]],
    en_us_by_id: dict[str, dict],
    en_gb_by_id: dict[str, dict],
) -> dict:
    blockers = []
    editions = []
    cell_status: dict[str, str] = {}
    display_cells_checked = 0
    example_cells_checked = 0
    transcription_cells_checked = 0
    sample_by_meaning_id = {row["meaning_id"]: row for row in sample_rows}

    for edition in manifest.get("editions", []):
        workbook_path = PROJECT_ROOT / edition["path"]
        main_records = load_workbook_rows(workbook_path, edition["sheet"])
        metadata_records = load_workbook_rows(workbook_path, "Card Metadata")
        metadata_by_meaning_id = {
            normalize_text(row.get("meaning_id")): row
            for row in metadata_records
            if normalize_text(row.get("meaning_id")) in sample_by_meaning_id
        }
        metadata_by_display_order = {}
        for row in metadata_records:
            try:
                display_order = int(row.get("display_order") or 0)
            except (TypeError, ValueError):
                display_order = 0
            if display_order:
                metadata_by_display_order[display_order] = row
        edition_display_cells_checked = 0
        edition_example_cells_checked = 0
        edition_transcription_cells_checked = 0
        for meaning_id, source_row in sample_by_meaning_id.items():
            row_id = source_row["row_id"]
            metadata_row = metadata_by_meaning_id.get(meaning_id)
            if not metadata_row:
                row_position = row_position_from_row_id(row_id)
                metadata_row = metadata_by_display_order.get(row_position or 0)
            main_sheet_row = int(metadata_row.get("main_sheet_row") or 0) if metadata_row else 0
            workbook_index = main_sheet_row - 2
            workbook_row = main_records[workbook_index] if 0 <= workbook_index < len(main_records) else None
            if not metadata_row or not workbook_row:
                blockers.append(
                    {
                        "edition_id": edition.get("edition_id"),
                        "row_id": row_id,
                        "meaning_id": meaning_id,
                        "reason": "sample_row_missing_from_final_workbook_metadata_or_main_sheet",
                    }
                )
                continue
            for language_code in edition.get("main_sheet_language_codes", []):
                expected_display = expected_values[row_id][language_code]["display"]
                actual_display = normalize_text(workbook_row.get(language_code))
                display_cells_checked += 1
                edition_display_cells_checked += 1
                display_key = f"{edition['edition_id']}::{row_id}::{language_code}::display"
                if expected_display != actual_display:
                    cell_status[display_key] = "blocker"
                    blockers.append(
                        {
                            "edition_id": edition.get("edition_id"),
                            "row_id": row_id,
                            "language_code": language_code,
                            "reason": "final_workbook_display_value_mismatch",
                            "expected": expected_display,
                            "actual": actual_display,
                        }
                    )
                else:
                    cell_status[display_key] = "pass"
                expected_example = expected_values[row_id][language_code]["example"]
                actual_example = normalize_text(workbook_row.get(f"{language_code} example"))
                example_cells_checked += 1
                edition_example_cells_checked += 1
                example_key = f"{edition['edition_id']}::{row_id}::{language_code}::example"
                if expected_example != actual_example:
                    cell_status[example_key] = "blocker"
                    blockers.append(
                        {
                            "edition_id": edition.get("edition_id"),
                            "row_id": row_id,
                            "language_code": language_code,
                            "reason": "final_workbook_example_value_mismatch",
                            "expected": expected_example,
                            "actual": actual_example,
                        }
                    )
                else:
                    cell_status[example_key] = "pass"

            pronunciation_row = expected_pronunciation_row(edition, row_id, en_us_by_id, en_gb_by_id)
            transcription_header, example_transcription_header = edition["main_sheet_pronunciation_headers"]
            expected_transcription = normalize_text(pronunciation_row.get("transcription"))
            expected_example_transcription = normalize_text(pronunciation_row.get("example_transcription"))
            actual_transcription = normalize_text(workbook_row.get(transcription_header))
            actual_example_transcription = normalize_text(workbook_row.get(example_transcription_header))
            transcription_cells_checked += 2
            edition_transcription_cells_checked += 2
            if expected_transcription != actual_transcription:
                blockers.append(
                    {
                        "edition_id": edition.get("edition_id"),
                        "row_id": row_id,
                        "language_code": edition.get("primary_language_code"),
                        "reason": "final_workbook_display_transcription_mismatch",
                        "expected": expected_transcription,
                        "actual": actual_transcription,
                    }
                )
            if expected_example_transcription != actual_example_transcription:
                blockers.append(
                    {
                        "edition_id": edition.get("edition_id"),
                        "row_id": row_id,
                        "language_code": edition.get("primary_language_code"),
                        "reason": "final_workbook_example_transcription_mismatch",
                        "expected": expected_example_transcription,
                        "actual": actual_example_transcription,
                    }
                )

        editions.append(
            {
                "edition_id": edition.get("edition_id"),
                "path": edition.get("path"),
                "sheet": edition.get("sheet"),
                "primary_language_code": edition.get("primary_language_code"),
                "sample_rows_checked": len(sample_rows),
                "display_cells_checked": edition_display_cells_checked,
                "example_cells_checked": edition_example_cells_checked,
                "transcription_cells_checked": edition_transcription_cells_checked,
            }
        )

    return {
        "status": "pass" if not blockers else "blocker",
        "editions": editions,
        "display_cells_checked": display_cells_checked,
        "example_cells_checked": example_cells_checked,
        "transcription_cells_checked": transcription_cells_checked,
        "cell_status": cell_status,
        "blockers": blockers,
    }


def workbook_status_for_entry(final_workbook_check: dict, row_id: str, language_code: str) -> str:
    statuses = []
    for edition in final_workbook_check.get("editions", []):
        edition_id = edition["edition_id"]
        display_key = f"{edition_id}::{row_id}::{language_code}::display"
        example_key = f"{edition_id}::{row_id}::{language_code}::example"
        if display_key in final_workbook_check["cell_status"] or example_key in final_workbook_check["cell_status"]:
            statuses.append(
                f"{edition_id}:"
                f"{final_workbook_check['cell_status'].get(display_key, 'not_checked')}/"
                f"{final_workbook_check['cell_status'].get(example_key, 'not_checked')}"
            )
    return ";".join(statuses) if statuses else "not_in_main_sheet"


def build_review(contract_path: Path, output_dir: Path, sample_size: int) -> tuple[Path, Path, Path, dict]:
    contract = read_json(contract_path)
    release_id = contract["latest_source_snapshot"]["release_id"]
    generated_at = datetime.now(timezone.utc).isoformat()

    language_rows = read_json(PROJECT_ROOT / "config/language-order.json")
    support_language_codes = [
        row["spreadsheetCode"] for row in language_rows if row["spreadsheetCode"] not in {"EN", "EN-GB"}
    ]
    review_language_codes = ["EN", "EN-GB", *support_language_codes]
    language_names = {row["spreadsheetCode"]: row["language"] for row in language_rows}
    language_names["EN-GB"] = "British English"

    row_reviews = read_jsonl(PROJECT_ROOT / contract["latest_row_review"]["path"])
    sample_rows = spread_sample(row_reviews, sample_size)
    expected_values = build_expected_values(contract, sample_rows)
    support_rows_by_language = build_support_rows_by_language(contract)
    source_audit_by_cell, example_audit_by_cell, audit_index_counts = build_audit_indexes(contract)

    en_us_contract = next(
        item for item in contract.get("latest_edition_pronunciations", []) if item.get("source_variant") == "EN-US"
    )
    en_gb_contract = next(
        item for item in contract.get("latest_edition_pronunciations", []) if item.get("source_variant") == "EN-GB"
    )
    en_us_by_id = by_row_id(read_jsonl(PROJECT_ROOT / en_us_contract["path"]))
    en_gb_by_id = by_row_id(read_jsonl(PROJECT_ROOT / en_gb_contract["path"]))
    edition_exports = contract["latest_edition_exports"]
    final_manifest_path = edition_exports.get("manifest_path")
    final_manifest = read_json(PROJECT_ROOT / final_manifest_path) if final_manifest_path else edition_exports

    final_workbook_check = check_final_workbooks(final_manifest, sample_rows, expected_values, en_us_by_id, en_gb_by_id)
    transcription_entries, transcription_blockers = check_transcription_rows(
        sample_rows, expected_values, en_us_by_id, en_gb_by_id
    )

    blockers = []
    warnings = []
    sample_entries = []
    summary_by_language = {}
    for language_code in review_language_codes:
        language_blocker_count = 0
        language_warning_count = 0
        for source_row in sample_rows:
            row_id = source_row["row_id"]
            display = expected_values.get(row_id, {}).get(language_code, {}).get("display", "")
            example = expected_values.get(row_id, {}).get(language_code, {}).get("example", "")
            entry_blockers = []
            entry_warnings = []
            source_audit_status = ""
            example_quality_status = ""
            source_warning_codes = []
            example_warning_codes = []
            if not display:
                entry_blockers.append("missing_display")
            if not example:
                entry_blockers.append("missing_example")
            if example and not SENTENCE_END_RE.search(example):
                entry_blockers.append("example_missing_sentence_punctuation")
            requirement = HARD_SCRIPT_REQUIREMENTS.get(language_code)
            if requirement and display and not requirement.search(display):
                entry_blockers.append("display_missing_expected_script")
            if requirement and example and not requirement.search(example):
                entry_blockers.append("example_missing_expected_script")
            if language_code not in {"EN", "EN-GB"}:
                support_row = support_rows_by_language.get(language_code, {}).get(row_id)
                if not support_row:
                    entry_blockers.append("missing_support_language_row")
                else:
                    transcription_like_keys = [
                        key
                        for key in support_row
                        if "transcription" in key.lower()
                        and support_transcription_field_is_forbidden(support_row.get(key))
                    ]
                    if transcription_like_keys:
                        entry_blockers.append(
                            f"support_language_transcription_fields_present:{','.join(transcription_like_keys)}"
                        )
                source_audit = source_audit_by_cell.get(support_key(row_id, language_code))
                example_audit = example_audit_by_cell.get(support_key(row_id, language_code))
                if not source_audit:
                    entry_blockers.append("missing_source_backed_display_audit_cell")
                else:
                    source_audit_status = normalize_text(source_audit.get("status"))
                    source_warning_codes = source_audit.get("warning_codes", []) or []
                    if source_audit_status != "pass" or source_audit.get("blocker_codes"):
                        entry_blockers.append("source_backed_display_audit_not_pass")
                if not example_audit:
                    entry_blockers.append("missing_support_example_quality_audit_cell")
                else:
                    example_quality_status = normalize_text(example_audit.get("status"))
                    example_warning_codes = example_audit.get("warning_codes", []) or []
                    if example_quality_status != "pass" or example_audit.get("blocker_codes"):
                        entry_blockers.append("support_example_quality_audit_not_pass")
                entry_warnings.extend([f"source_audit:{code}" for code in source_warning_codes])
                entry_warnings.extend([f"example_quality:{code}" for code in example_warning_codes])
            if entry_blockers:
                language_blocker_count += len(entry_blockers)
                blockers.append(
                    {
                        "row_id": row_id,
                        "source_headword": source_row["source_headword"],
                        "language_code": language_code,
                        "reasons": entry_blockers,
                    }
                )
            if entry_warnings:
                language_warning_count += len(entry_warnings)
                warnings.append(
                    {
                        "row_id": row_id,
                        "source_headword": source_row["source_headword"],
                        "language_code": language_code,
                        "warnings": entry_warnings,
                    }
                )
            sample_entries.append(
                {
                    "language_code": language_code,
                    "language": language_names.get(language_code, language_code),
                    "row_id": row_id,
                    "source_headword": source_row["source_headword"],
                    "meaning_id": source_row["meaning_id"],
                    "meaning_note": source_row.get("meaning_note", ""),
                    "display": display,
                    "example": example,
                    "source_audit_status": source_audit_status,
                    "source_audit_warning_codes": ",".join(source_warning_codes),
                    "example_quality_status": example_quality_status,
                    "example_quality_warning_codes": ",".join(example_warning_codes),
                    "workbook_status": workbook_status_for_entry(final_workbook_check, row_id, language_code),
                    "status": "pass" if not entry_blockers else "blocker",
                    "blockers": ",".join(entry_blockers),
                    "advisory_warnings": ",".join(entry_warnings),
                }
            )
        summary_by_language[language_code] = {
            "language": language_names.get(language_code, language_code),
            "sample_rows_checked": len(sample_rows),
            "display_cells_checked": len(sample_rows),
            "example_cells_checked": len(sample_rows),
            "blocker_count": language_blocker_count,
            "advisory_warning_signal_count": language_warning_count,
            "status": "pass" if language_blocker_count == 0 else "blocker",
        }

    blockers.extend(transcription_blockers)
    blockers.extend(final_workbook_check["blockers"])
    status = "pass" if not blockers else "blocker"
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"{release_id}_{REPORT_ID}.json"
    report_csv_path = output_dir / f"{release_id}_{REPORT_ID}.csv"
    report_md_path = output_dir / f"{release_id}_{REPORT_ID}.md"

    report = {
        "release_id": release_id,
        "report_id": REPORT_ID,
        "script_version": SCRIPT_VERSION,
        "generated_at": generated_at,
        "status": status,
        "sample_method": "deterministic_even_spread_across_row_review_order",
        "sample_size_per_language": len(sample_rows),
        "sample_row_ids": [row["row_id"] for row in sample_rows],
        "sample_source_headwords": [row["source_headword"] for row in sample_rows],
        "languages_reviewed": len(review_language_codes),
        "support_languages_reviewed": len(support_language_codes),
        "source_package_display_cells_checked": len(sample_entries),
        "source_package_example_cells_checked": len(sample_entries),
        "support_source_audit_cells_available": audit_index_counts["source_audit_total_cells"],
        "support_example_audit_cells_available": audit_index_counts["example_audit_total_cells"],
        "final_workbook_check": final_workbook_check,
        "transcription_entries": transcription_entries,
        "transcription_rows_checked": len(transcription_entries),
        "blocker_count": len(blockers),
        "advisory_warning_signal_count": sum(len(item["warnings"]) for item in warnings),
        "blockers": blockers,
        "advisory_warnings": warnings,
        "summary_by_language": summary_by_language,
        "sample_entries": sample_entries,
        "inputs": {
            "contract_path": relative(contract_path),
            "row_review_path": contract["latest_row_review"]["path"],
            "edition_layer_path": (
                contract.get("latest_edition_exports", {}).get("edition_layer_path")
                or contract.get("latest_edition_layer", {}).get("path")
            ),
            "final_manifest_path": final_manifest_path or "embedded_contract.latest_edition_exports",
            "en_us_pronunciation_path": en_us_contract["path"],
            "en_gb_pronunciation_path": en_gb_contract["path"],
            "source_audit_path": contract.get("latest_support_translation_source_backed_audit", {}).get("path"),
            "example_audit_path": contract.get("latest_support_example_quality_audit", {}).get("path"),
        },
        "does_not_replace": "native-speaker certification or a full row-by-row linguistic audit",
    }

    write_json(report_path, report)
    with report_csv_path.open("w", encoding="utf-8", newline="") as handle:
        fieldnames = [
            "language_code",
            "language",
            "source_headword",
            "row_id",
            "meaning_id",
            "meaning_note",
            "display",
            "example",
            "source_audit_status",
            "source_audit_warning_codes",
            "example_quality_status",
            "example_quality_warning_codes",
            "workbook_status",
            "status",
            "blockers",
            "advisory_warnings",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fieldnames} for row in sample_entries)

    markdown_lines = [
        f"# Oxford Post-Final Language Spot Check: {release_id}",
        "",
        f"- Report id: `{REPORT_ID}`",
        f"- Status: `{status}`",
        f"- Generated at: `{generated_at}`",
        "- Method: deterministic even-spread sample from the reviewed release row order.",
        f"- Sample size: {len(sample_rows)} rows per language variant",
        f"- Language variants checked: {len(review_language_codes)}",
        f"- Source-package display cells checked: {len(sample_entries)}",
        f"- Source-package example cells checked: {len(sample_entries)}",
        f"- Final workbook display cells checked: {final_workbook_check['display_cells_checked']}",
        f"- Final workbook example cells checked: {final_workbook_check['example_cells_checked']}",
        f"- Final workbook EN/EN-GB transcription cells checked: {final_workbook_check['transcription_cells_checked']}",
        f"- EN/EN-GB source-backed transcription rows checked: {len(transcription_entries)}",
        f"- Blockers: {len(blockers)}",
        f"- Advisory warning signals from existing support audits: {report['advisory_warning_signal_count']}",
        "- Native-speaker certification: not claimed.",
        "",
        "## Sample Rows",
        "",
        "| # | Source headword | Row id | Meaning note |",
        "| --- | --- | --- | --- |",
    ]
    for index, row in enumerate(sample_rows, start=1):
        markdown_lines.append(
            f"| {index} | `{markdown_cell(row['source_headword'])}` | `{markdown_cell(row['row_id'])}` | {markdown_cell(row.get('meaning_note'))} |"
        )
    markdown_lines.extend(
        [
            "",
            "## EN/EN-GB Transcriptions",
            "",
            "| Source headword | Variant | Word IPA | Example IPA | Status |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for entry in transcription_entries:
        markdown_lines.append(
            f"| `{markdown_cell(entry['source_headword'])}` | `{entry['variant']}` | `{markdown_cell(entry['transcription'])}` | `{markdown_cell(entry['example_transcription'])}` | `{entry['status']}` |"
        )
    markdown_lines.extend(
        [
            "",
            "## Artifacts",
            "",
            f"- JSON: `{relative(report_path)}`",
            f"- CSV: `{relative(report_csv_path)}`",
            f"- Markdown: `{relative(report_md_path)}`",
            "",
        ]
    )
    report_md_path.write_text("\n".join(markdown_lines) + "\n", encoding="utf-8")

    contract["latest_post_final_language_spot_check"] = {
        "report_id": REPORT_ID,
        "script_path": relative(Path(__file__)),
        "path": relative(report_path),
        "csv_path": relative(report_csv_path),
        "markdown_path": relative(report_md_path),
        "status": status,
        "generated_at": generated_at,
        "sample_method": report["sample_method"],
        "sample_size_per_language": report["sample_size_per_language"],
        "languages_reviewed": report["languages_reviewed"],
        "support_languages_reviewed": report["support_languages_reviewed"],
        "sample_source_headwords": report["sample_source_headwords"],
        "source_package_display_cells_checked": report["source_package_display_cells_checked"],
        "source_package_example_cells_checked": report["source_package_example_cells_checked"],
        "final_workbook_display_cells_checked": final_workbook_check["display_cells_checked"],
        "final_workbook_example_cells_checked": final_workbook_check["example_cells_checked"],
        "final_workbook_transcription_cells_checked": final_workbook_check["transcription_cells_checked"],
        "transcription_rows_checked": report["transcription_rows_checked"],
        "blocker_count": report["blocker_count"],
        "advisory_warning_signal_count": report["advisory_warning_signal_count"],
        "does_not_replace": report["does_not_replace"],
    }
    write_json(contract_path, contract)

    return report_path, report_csv_path, report_md_path, report


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a post-final Oxford language spot-check report.")
    parser.add_argument("--contract", default=str(DEFAULT_CONTRACT_PATH))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--sample-size", type=int, default=5)
    args = parser.parse_args()
    report_path, report_csv_path, report_md_path, report = build_review(
        Path(args.contract), Path(args.output_dir), args.sample_size
    )
    print(
        "Oxford post-final language spot check "
        f"status={report['status']} "
        f"languages={report['languages_reviewed']} "
        f"sample_per_language={report['sample_size_per_language']} "
        f"source_display_cells={report['source_package_display_cells_checked']} "
        f"source_example_cells={report['source_package_example_cells_checked']} "
        f"workbook_display_cells={report['final_workbook_check']['display_cells_checked']} "
        f"workbook_example_cells={report['final_workbook_check']['example_cells_checked']} "
        f"workbook_transcription_cells={report['final_workbook_check']['transcription_cells_checked']} "
        f"transcription_rows={report['transcription_rows_checked']} "
        f"blockers={report['blocker_count']} "
        f"json={relative(report_path)} "
        f"csv={relative(report_csv_path)} "
        f"markdown={relative(report_md_path)}"
    )


if __name__ == "__main__":
    main()
