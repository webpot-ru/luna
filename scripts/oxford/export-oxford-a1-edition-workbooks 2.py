#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_VERSION = "2026-05-17.v1"
SUPPORT_STATUS = "draft_support_language_aid_needs_final_source_approval"

EDITION_CONFIG = {
    "us_english": {
        "deck_id": "oxford_3000_core_a1_part_001_us_english_draft",
        "deck_title": "FlashcardsLuna Oxford 3000 Core A1 Part 001 - US English",
        "deck_edition": "US English",
        "source_variant": "EN-US",
        "primary_language_code": "EN",
        "display_field": "display_headword_EN_US",
        "example_field": "example_EN_US",
        "edition_note_field": "edition_note_EN_US",
        "pronunciation_artifact_id": "en_us_edition_pronunciations_v1",
        "sheet_name": "Oxford A1 US",
        "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_US_English_draft.xlsx",
    },
    "british_english": {
        "deck_id": "oxford_3000_core_a1_part_001_british_english_draft",
        "deck_title": "FlashcardsLuna Oxford 3000 Core A1 Part 001 - British English",
        "deck_edition": "British English",
        "source_variant": "EN-GB",
        "primary_language_code": "EN-GB",
        "display_field": "display_headword_EN_GB",
        "example_field": "example_EN_GB",
        "edition_note_field": "edition_note_EN_GB",
        "pronunciation_artifact_id": "en_gb_edition_pronunciations_v1",
        "sheet_name": "Oxford A1 UK",
        "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_British_English_draft.xlsx",
    },
}

MAIN_HEADERS = [
    "release_id",
    "deck_id",
    "deck_title",
    "deck_edition",
    "source_variant",
    "row_id",
    "core_item_id",
    "meaning_id",
    "source_headword",
    "display_headword",
    "part_of_speech",
    "sense_no",
    "core_band",
    "level_min",
    "level_max",
    "benchmark_membership",
    "source_status",
    "meaning_note",
    "semantic_scene",
    "example",
    "transcription",
    "example_transcription",
    "pronunciation_status",
    "pronunciation_source",
    "variant_note",
]


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def read_jsonl(path):
    rows = []
    for index, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as error:
            raise ValueError(f"Invalid JSONL at {path}:{index}: {error}") from error
    return rows


def load_language_codes(path):
    data = read_json(path)
    return [row["spreadsheetCode"] for row in data]


def load_support_rows(contract, support_codes):
    support_by_row = {}
    seen_codes = set()
    source_paths = []
    for batch in contract.get("latest_support_translation_batches", []):
        path = Path(batch["path"])
        rows = read_jsonl(path)
        source_paths.append(str(path))
        languages = batch.get("languages", [])
        seen_codes.update(languages)
        for row in rows:
            row_id = row["row_id"]
            target = support_by_row.setdefault(row_id, {})
            for code in languages:
                value = normalize_text(row.get(code))
                example = normalize_text(row.get(f"example_{code}"))
                if not value:
                    raise ValueError(f"Missing support translation {code} for {row_id} in {path}")
                if not example:
                    raise ValueError(f"Missing support example {code} for {row_id} in {path}")
                if code in target:
                    raise ValueError(f"Duplicate support translation {code} for {row_id}")
                target[code] = value
                target[f"example_{code}"] = example

    missing_codes = sorted(set(support_codes) - seen_codes)
    extra_codes = sorted(seen_codes - set(support_codes))
    if missing_codes or extra_codes:
        raise ValueError(f"Support language coverage mismatch missing={missing_codes} extra={extra_codes}")
    return support_by_row, source_paths


def load_pronunciations(path):
    rows = read_jsonl(path)
    by_row = {row["row_id"]: row for row in rows}
    if len(by_row) != len(rows):
        raise ValueError(f"Duplicate pronunciation rows in {path}")
    return by_row


def as_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def build_main_row(layer_row, support_cells, pronunciation_row, edition, support_codes):
    cfg = EDITION_CONFIG[edition]
    row = {
        "release_id": layer_row["release_id"],
        "deck_id": cfg["deck_id"],
        "deck_title": cfg["deck_title"],
        "deck_edition": cfg["deck_edition"],
        "source_variant": cfg["source_variant"],
        "row_id": layer_row["row_id"],
        "core_item_id": layer_row["core_item_id"],
        "meaning_id": layer_row["meaning_id"],
        "source_headword": layer_row["source_headword"],
        "display_headword": layer_row[cfg["display_field"]],
        "part_of_speech": layer_row["reviewed_part_of_speech"],
        "sense_no": layer_row["sense_no"],
        "core_band": layer_row["core_band"],
        "level_min": layer_row["level_min"],
        "level_max": layer_row["level_max"],
        "benchmark_membership": layer_row["benchmark_membership"],
        "source_status": layer_row["source_status"],
        "meaning_note": layer_row["meaning_note"],
        "semantic_scene": layer_row["semantic_scene"],
        "example": layer_row[cfg["example_field"]],
        "transcription": pronunciation_row["transcription"],
        "example_transcription": pronunciation_row["example_transcription"],
        "pronunciation_status": pronunciation_row["transcription_status"],
        "pronunciation_source": pronunciation_row["pronunciation_source"],
        "variant_note": layer_row[cfg["edition_note_field"]],
        "qa_status": SUPPORT_STATUS,
        "qa_notes": "Local draft workbook only: no Postgres import, no Google Sheet delivery, final legal/allowed-fields/delivery gates still blocked.",
    }
    for code in support_codes:
        row[code] = support_cells[code]
        row[f"example_{code}"] = support_cells[f"example_{code}"]
    return row


def style_sheet(sheet, headers):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    widths = {
        "A": 22,
        "B": 40,
        "C": 52,
        "D": 18,
        "E": 14,
        "F": 42,
        "G": 18,
        "H": 42,
        "I": 22,
        "J": 24,
        "K": 22,
        "L": 12,
        "M": 16,
        "N": 12,
        "O": 12,
        "P": 20,
        "Q": 32,
        "R": 52,
        "S": 60,
        "T": 36,
        "U": 24,
        "V": 36,
        "W": 34,
        "X": 48,
        "Y": 52,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for index in range(26, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 24


def add_pair_sheet(workbook, name, pairs):
    sheet = workbook.create_sheet(name)
    sheet.append(["key", "value"])
    for key, value in pairs:
        sheet.append([key, as_cell(value)])
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 90
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)


def write_workbook(path, edition, layer_rows, support_by_row, pronunciation_by_row, support_codes, source_paths, generated_at):
    cfg = EDITION_CONFIG[edition]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = cfg["sheet_name"]
    headers = MAIN_HEADERS + support_codes + [f"example_{code}" for code in support_codes] + ["qa_status", "qa_notes"]
    sheet.append(headers)

    for layer_row in layer_rows:
        row_id = layer_row["row_id"]
        support_cells = support_by_row.get(row_id)
        if not support_cells:
            raise ValueError(f"Missing support cells for {row_id}")
        pronunciation_row = pronunciation_by_row.get(row_id)
        if not pronunciation_row:
            raise ValueError(f"Missing pronunciation row for {row_id}")
        main_row = build_main_row(layer_row, support_cells, pronunciation_row, edition, support_codes)
        sheet.append([as_cell(main_row.get(header, "")) for header in headers])

    style_sheet(sheet, headers)
    add_pair_sheet(
        workbook,
        "_README",
        [
            ("deck_title", cfg["deck_title"]),
            ("deck_id", cfg["deck_id"]),
            ("deck_edition", cfg["deck_edition"]),
            ("source_variant", cfg["source_variant"]),
            ("primary_language_code", cfg["primary_language_code"]),
            ("rows", len(layer_rows)),
            ("columns", len(headers)),
            ("status", "local_draft_not_final_delivery"),
            ("postgres_import", False),
            ("google_sheet_created", False),
            ("generated_at", generated_at),
        ],
    )
    add_pair_sheet(
        workbook,
        "_edition_contract",
        [
            ("principle", "one canonical Oxford A1 source package, two buyer-facing edition exports"),
            ("shared_identity_fields", ["release_id", "row_id", "core_item_id", "meaning_id"]),
            ("edition_specific_fields", ["deck_id", "deck_title", "deck_edition", "source_variant", "display_headword", "example", "transcription", "example_transcription"]),
            ("support_language_columns", support_codes),
            ("final_delivery_ready", False),
            ("remaining_blockers", ["permission_evidence_or_project_evidence_decision", "allowed_fields_review", "final_delivery_approval_check"]),
        ],
    )
    languages = workbook.create_sheet("_languages")
    languages.append(["role", "spreadsheetCode"])
    languages.append(["primary_english", cfg["primary_language_code"]])
    for code in support_codes:
        languages.append(["support_language", code])
    languages.column_dimensions["A"].width = 24
    languages.column_dimensions["B"].width = 20
    add_pair_sheet(
        workbook,
        "_source_artifacts",
        [
            ("edition_layer", "outputs/oxford-vocabulary/edition-layers/oxford_3000_core_a1_part_001_150_v1_us_uk_edition_layer_v1.jsonl"),
            ("pronunciation_artifact_id", cfg["pronunciation_artifact_id"]),
            ("support_translation_batch_paths", source_paths),
            ("copied_oxford_definitions_examples_pronunciations_audio", False),
        ],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return {"path": str(path), "rows": len(layer_rows), "columns": len(headers), "sheet": cfg["sheet_name"]}


def workbook_shape(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return {"rows": sheet.max_row - 1, "columns": sheet.max_column, "main_sheet": sheet.title, "sheets": workbook.sheetnames}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--contract", default="config/oxford-vocabulary-release-contract-v0.json")
    parser.add_argument(
        "--edition-layer",
        default="outputs/oxford-vocabulary/edition-layers/oxford_3000_core_a1_part_001_150_v1_us_uk_edition_layer_v1.jsonl",
    )
    parser.add_argument(
        "--en-us-pronunciations",
        default="outputs/oxford-vocabulary/pronunciations/oxford_3000_core_a1_part_001_150_v1_en_us_edition_pronunciations_v1.jsonl",
    )
    parser.add_argument(
        "--en-gb-pronunciations",
        default="outputs/oxford-vocabulary/pronunciations/oxford_3000_core_a1_part_001_150_v1_en_gb_edition_pronunciations_v1.jsonl",
    )
    parser.add_argument("--languages", default="config/language-order.json")
    parser.add_argument("--out-dir", default="outputs/oxford-vocabulary/editions")
    parser.add_argument("--export-id", default="edition_exports_v1")
    args = parser.parse_args()

    contract = read_json(Path(args.contract))
    layer_path = Path(args.edition_layer)
    layer_rows = read_jsonl(layer_path)
    if not layer_rows:
        raise ValueError("Edition layer artifact is empty")
    release_id = layer_rows[0]["release_id"]
    language_codes = load_language_codes(Path(args.languages))
    support_codes = [code for code in language_codes if code not in {"EN", "EN-GB"}]
    if len(support_codes) != 52:
        raise ValueError(f"Expected 52 support languages, found {len(support_codes)}")

    support_by_row, source_paths = load_support_rows(contract, support_codes)
    en_us_by_row = load_pronunciations(Path(args.en_us_pronunciations))
    en_gb_by_row = load_pronunciations(Path(args.en_gb_pronunciations))
    generated_at = datetime.now(timezone.utc).isoformat()
    out_dir = Path(args.out_dir)

    editions = []
    for edition, pronunciation_rows in [
        ("us_english", en_us_by_row),
        ("british_english", en_gb_by_row),
    ]:
        path = out_dir / EDITION_CONFIG[edition]["file_name"]
        info = write_workbook(path, edition, layer_rows, support_by_row, pronunciation_rows, support_codes, source_paths, generated_at)
        info.update(
            {
                "edition_id": edition,
                "deck_id": EDITION_CONFIG[edition]["deck_id"],
                "deck_title": EDITION_CONFIG[edition]["deck_title"],
                "source_variant": EDITION_CONFIG[edition]["source_variant"],
                "primary_language_code": EDITION_CONFIG[edition]["primary_language_code"],
                "shape": workbook_shape(path),
            }
        )
        editions.append(info)

    manifest = {
        "export_id": args.export_id,
        "script_version": SCRIPT_VERSION,
        "status": "local_draft_edition_workbooks_not_final_delivery",
        "release_id": release_id,
        "generated_at": generated_at,
        "principle": "one canonical source package, two buyer-facing US/UK English edition exports",
        "edition_layer_path": str(layer_path),
        "support_language_count": len(support_codes),
        "support_display_cells_per_workbook": len(layer_rows) * len(support_codes),
        "support_example_cells_per_workbook": len(layer_rows) * len(support_codes),
        "postgres_import": False,
        "google_sheet_created": False,
        "final_delivery_ready": False,
        "remaining_blockers": [
            "permission_evidence_or_project_evidence_decision",
            "allowed_fields_review",
            "final_delivery_approval_check",
        ],
        "editions": editions,
    }
    manifest_path = out_dir / f"{release_id}_{args.export_id}.json"
    summary_path = out_dir / f"{release_id}_{args.export_id}_summary.md"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    summary = [
        f"# Oxford A1 US/UK Edition Workbook Exports: {release_id}",
        "",
        f"- Script version: `{SCRIPT_VERSION}`",
        f"- Status: `{manifest['status']}`",
        f"- Editions: {len(editions)}",
        f"- Rows per edition: {len(layer_rows)}",
        f"- Support languages per edition: {len(support_codes)}",
        f"- Support display cells per edition: {manifest['support_display_cells_per_workbook']}",
        f"- Support example cells per edition: {manifest['support_example_cells_per_workbook']}",
        "- Postgres import: false",
        "- Google Sheet created: false",
        "- Final delivery ready: false",
        "",
    ]
    for edition in editions:
        summary.append(f"- `{edition['deck_title']}`: `{edition['path']}`")
    summary_path.write_text("\n".join(summary) + "\n", encoding="utf-8")

    print(
        f"Oxford edition workbooks exported: editions={len(editions)} rows={len(layer_rows)} "
        f"manifest={manifest_path} summary={summary_path}"
    )


if __name__ == "__main__":
    main()
