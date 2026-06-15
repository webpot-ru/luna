#!/usr/bin/env python3
"""Fast read-only audit for Oxford 150-word XLSX workbooks.

This script audits the current 150-word delivery layer without mutating local
workbooks, Google Sheets, DB state or contracts. It is deterministic QA, not
native-speaker certification.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_SERVICE_SHEETS = {
    "Course Metadata",
    "Deck Metadata",
    "Card Metadata",
    "_README",
    "_qa_status",
    "_languages",
}
LANGUAGE_COUNT = 53
MAIN_COLUMN_COUNT = 108
FILENAME_RE = re.compile(
    r"^FlashcardsLuna_Oxford_(?P<course>3000_Core|5000_Advanced_Extension)_"
    r"(?P<level>A1|A2|B1|B2|C1)_Part_(?P<part>\d{3})_"
    r"(?P<rows>\d+)_(?P<edition>US|British)_English\.xlsx$"
)

SCRIPT_PATTERNS = {
    "RU": re.compile(r"[\u0400-\u04FF]"),
    "BG": re.compile(r"[\u0400-\u04FF]"),
    "SR": re.compile(r"[\u0400-\u04FF]"),
    "ZH": re.compile(r"[\u4E00-\u9FFF]"),
    "JA": re.compile(r"[\u3040-\u30FF\u4E00-\u9FFF]"),
    "KO": re.compile(r"[\uAC00-\uD7AF]"),
    "TH": re.compile(r"[\u0E00-\u0E7F]"),
    "MY": re.compile(r"[\u1000-\u109F]"),
    "KM": re.compile(r"[\u1780-\u17FF]"),
    "LO": re.compile(r"[\u0E80-\u0EFF]"),
    "HI": re.compile(r"[\u0900-\u097F]"),
    "NE": re.compile(r"[\u0900-\u097F]"),
    "BN": re.compile(r"[\u0980-\u09FF]"),
    "SI": re.compile(r"[\u0D80-\u0DFF]"),
    "TA": re.compile(r"[\u0B80-\u0BFF]"),
    "TE": re.compile(r"[\u0C00-\u0C7F]"),
    "KN": re.compile(r"[\u0C80-\u0CFF]"),
    "ML": re.compile(r"[\u0D00-\u0D7F]"),
    "KA": re.compile(r"[\u10A0-\u10FF]"),
    "HY": re.compile(r"[\u0530-\u058F]"),
}

BAD_ANCHOR_PATTERNS = [
    re.compile(r"\bThe AIDS changed the plan\b", re.I),
    re.compile(r"\bThe decision was audio\b", re.I),
    re.compile(r"\bThe decision was biological\b", re.I),
    re.compile(r"\bThe report mentioned an agriculture\b", re.I),
    re.compile(r"\bThey [a-z]+ the issue carefully\b", re.I),
]


def is_documented_b1_extension_row(match: re.Match[str], expected_level: str, actual_level: str) -> bool:
    return (
        match.group("course") == "5000_Advanced_Extension"
        and expected_level == "B2"
        and actual_level == "B1"
    )


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def rows(wb, sheet: str) -> list[list[str]]:
    return [[text(value) for value in row] for row in wb[sheet].iter_rows(values_only=True)]


def get(row: list[str], index: int) -> str:
    return row[index] if index < len(row) else ""


def finding(
    out: list[dict[str, Any]],
    severity: str,
    code: str,
    workbook: Path,
    detail: str,
    sheet: str = "",
    row: int | None = None,
    language: str = "",
) -> None:
    out.append(
        {
            "severity": severity,
            "code": code,
            "workbook": workbook.name,
            "sheet": sheet,
            "row": row,
            "language": language,
            "detail": detail,
        }
    )


def kv_metadata(sheet_rows: list[list[str]]) -> dict[str, str]:
    return {get(row, 0): get(row, 1) for row in sheet_rows[1:] if get(row, 0)}


def audit_one(path: Path, manifest_rows: dict[str, int]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    findings: list[dict[str, Any]] = []
    match = FILENAME_RE.match(path.name)
    info = {"workbook": path.name, "row_count": 0, "course": "", "level": "", "edition": ""}
    if not match:
        finding(findings, "blocker", "filename_shape", path, "Unexpected filename.")
        return info, findings

    expected_course = "Oxford 3000 Core" if match.group("course") == "3000_Core" else "Oxford 5000 Advanced Extension"
    expected_level = match.group("level")
    expected_rows = manifest_rows.get(str(path), manifest_rows.get(path.name, int(match.group("rows"))))
    expected_edition = "US English" if match.group("edition") == "US" else "British English"
    info.update({"row_count": expected_rows, "course": expected_course, "level": expected_level, "edition": expected_edition})

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in sorted(EXPECTED_SERVICE_SHEETS - set(wb.sheetnames)):
        finding(findings, "blocker", "missing_service_sheet", path, f"Missing sheet {sheet}.")

    main_name = wb.sheetnames[0]
    main = rows(wb, main_name)
    header = main[0] if main else []
    if len(header) != MAIN_COLUMN_COUNT:
        finding(findings, "blocker", "main_column_count", path, f"Expected {MAIN_COLUMN_COUNT}, got {len(header)}.", main_name)
    if len(main) - 1 != expected_rows:
        finding(findings, "blocker", "main_row_count", path, f"Expected {expected_rows}, got {len(main) - 1}.", main_name)
    if expected_level not in main_name:
        finding(findings, "blocker", "main_sheet_level_mismatch", path, f"Main sheet {main_name!r} lacks {expected_level}.", main_name)

    word_langs = header[:LANGUAGE_COUNT]
    example_langs = [h.removesuffix(" example") for h in header[LANGUAGE_COUNT : LANGUAGE_COUNT * 2]]
    if word_langs != example_langs:
        finding(findings, "blocker", "word_example_language_order", path, "Word/example headers do not match.", main_name)
    edition_word_header = "EN-GB" if expected_edition == "British English" else "EN"
    expected_transcription_header = f"{edition_word_header} transcription"
    expected_example_transcription_header = f"{edition_word_header} example transcription"
    if get(header, 106) != expected_transcription_header or get(header, 107) != expected_example_transcription_header:
        finding(findings, "blocker", "english_transcription_headers", path, "Bad EN transcription headers.", main_name)

    for excel_row, row in enumerate(main[1:], start=2):
        en_word = get(row, 0)
        en_example = get(row, LANGUAGE_COUNT)
        for pattern in BAD_ANCHOR_PATTERNS:
            if pattern.search(en_example):
                finding(findings, "blocker", "known_bad_english_anchor", path, en_example, main_name, excel_row, "EN")
        for offset, lang in enumerate(word_langs):
            word = get(row, offset)
            example = get(row, LANGUAGE_COUNT + offset)
            if not word:
                finding(findings, "blocker", "empty_word_cell", path, "Empty word/display cell.", main_name, excel_row, lang)
            if not example:
                finding(findings, "blocker", "empty_example_cell", path, "Empty example cell.", main_name, excel_row, lang)
            if lang not in {"EN", "EN-GB"}:
                if word and word.casefold() == en_word.casefold():
                    finding(findings, "warning", "exact_english_word_fallback", path, word, main_name, excel_row, lang)
                if example and example.casefold() == en_example.casefold():
                    finding(findings, "blocker", "exact_english_example_fallback", path, "Target example equals EN example.", main_name, excel_row, lang)
            script = SCRIPT_PATTERNS.get(lang)
            if script and word and not script.search(word):
                finding(findings, "warning", "expected_script_missing_word", path, word, main_name, excel_row, lang)
            if script and example and not script.search(example):
                finding(findings, "warning", "expected_script_missing_example", path, example, main_name, excel_row, lang)

    deck = kv_metadata(rows(wb, "Deck Metadata")) if "Deck Metadata" in wb.sheetnames else {}
    if deck.get("domain") != expected_course:
        finding(findings, "blocker", "deck_metadata_domain_mismatch", path, f"Expected {expected_course}, got {deck.get('domain')}.", "Deck Metadata")
    if deck.get("area") != expected_level:
        finding(findings, "blocker", "deck_metadata_area_mismatch", path, f"Expected {expected_level}, got {deck.get('area')}.", "Deck Metadata")
    if deck.get("deck_edition") != expected_edition:
        finding(findings, "blocker", "deck_metadata_edition_mismatch", path, f"Expected {expected_edition}, got {deck.get('deck_edition')}.", "Deck Metadata")

    if "Card Metadata" in wb.sheetnames:
        card = rows(wb, "Card Metadata")
        ch = card[0] if card else []
        ix = {name: i for i, name in enumerate(ch)}
        if len(card) - 1 != expected_rows:
            finding(findings, "blocker", "card_metadata_row_count", path, f"Expected {expected_rows}, got {len(card) - 1}.", "Card Metadata")
        for required in ["main_sheet_row", "card_key", "set_id", "meaning_id", "level"]:
            if required not in ix:
                finding(findings, "blocker", "card_metadata_missing_column", path, f"Missing {required}.", "Card Metadata")
        for excel_row, row in enumerate(card[1:], start=2):
            if "main_sheet_row" in ix:
                try:
                    main_row = int(get(row, ix["main_sheet_row"]))
                except ValueError:
                    main_row = -1
                if main_row != excel_row:
                    finding(findings, "blocker", "card_metadata_main_row_mismatch", path, f"Expected {excel_row}, got {main_row}.", "Card Metadata", excel_row)
            if "set_id" in ix and deck.get("set_id") and get(row, ix["set_id"]) != deck["set_id"]:
                finding(findings, "blocker", "card_metadata_set_id_mismatch", path, "set_id differs from Deck Metadata.", "Card Metadata", excel_row)
            if "level" in ix:
                actual_level = get(row, ix["level"])
                if actual_level != expected_level:
                    if is_documented_b1_extension_row(match, expected_level, actual_level):
                        finding(
                            findings,
                            "warning",
                            "documented_b1_extension_row_in_b2_extension",
                            path,
                            f"Workbook level is {expected_level}; retained source row level is {actual_level}.",
                            "Card Metadata",
                            excel_row,
                        )
                    else:
                        finding(findings, "blocker", "card_metadata_level_mismatch", path, f"Expected {expected_level}, got {actual_level}.", "Card Metadata", excel_row)

    if "Course Metadata" in wb.sheetnames:
        course = rows(wb, "Course Metadata")
        course_langs = course[0][1:] if course else []
        if course_langs != word_langs:
            finding(findings, "blocker", "course_metadata_language_order", path, "Language order differs from main sheet.", "Course Metadata")
        labels = {get(row, 0) for row in course[1:]}
        for label in ["Title", "Description", "Module", "Category"]:
            if label not in labels:
                finding(findings, "blocker", "course_metadata_missing_row", path, f"Missing {label}.", "Course Metadata")
        for row_number, row in enumerate(course[1:], start=2):
            label = get(row, 0)
            for col, lang in enumerate(course_langs, start=1):
                value = get(row, col)
                if not value:
                    finding(findings, "blocker", "course_metadata_empty_cell", path, f"Empty {label}.", "Course Metadata", row_number, lang)
                if label == "Title" and len(value) > 30:
                    finding(findings, "warning", "course_metadata_title_long", path, value, "Course Metadata", row_number, lang)
                if label == "Description" and len(value) > 80:
                    finding(findings, "warning", "course_metadata_description_long", path, value, "Course Metadata", row_number, lang)
                script = SCRIPT_PATTERNS.get(lang)
                if script and label in {"Title", "Description", "Category"} and value and not script.search(value):
                    finding(findings, "warning", "course_metadata_expected_script_missing", path, value, "Course Metadata", row_number, lang)

    return info, findings


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default="outputs/oxford-vocabulary/final-150")
    parser.add_argument("--manifest", default="outputs/oxford-vocabulary/final-150/oxford_150_resplit_manifest_v1.json")
    parser.add_argument("--out", default="outputs/oxford-vocabulary/qa/oxford_150_workbook_quality_audit_v1.json")
    parser.add_argument("--max-findings", type=int, default=500)
    args = parser.parse_args()

    workbooks = sorted(Path(args.input_dir).glob("*.xlsx"))
    manifest_rows: dict[str, int] = {}
    manifest_path = Path(args.manifest)
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        for release in manifest.get("releases", []):
            row_count = int(release.get("row_count") or 0)
            for edition in release.get("editions", []):
                workbook_path = str(edition.get("workbook_path") or "")
                if workbook_path:
                    manifest_rows[workbook_path] = row_count
                    manifest_rows[Path(workbook_path).name] = row_count
    all_findings: list[dict[str, Any]] = []
    infos: list[dict[str, Any]] = []
    for workbook in workbooks:
        info, findings = audit_one(workbook, manifest_rows)
        infos.append(info)
        all_findings.extend(findings)

    severity = Counter(f["severity"] for f in all_findings)
    codes = Counter(f["code"] for f in all_findings)
    blocked_workbooks: dict[str, int] = defaultdict(int)
    for item in all_findings:
        if item["severity"] == "blocker":
            blocked_workbooks[item["workbook"]] += 1

    report = {
        "report_id": "oxford_150_workbook_quality_audit_v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_dir": args.input_dir,
        "manifest": args.manifest,
        "status": "blocked" if severity["blocker"] else "pass",
        "native_speaker_certification": False,
        "scope": "Read-only deterministic audit of all Oxford 150-word local XLSX workbooks.",
        "workbook_count": len(workbooks),
        "checked_rows": sum(int(info["row_count"]) for info in infos),
        "severity_counts": dict(severity),
        "code_counts": dict(codes),
        "blocked_workbook_count": len(blocked_workbooks),
        "blocked_workbooks": dict(sorted(blocked_workbooks.items())),
        "workbooks": infos,
        "finding_count": len(all_findings),
        "findings_sample": all_findings[: args.max_findings],
    }

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    md = out.with_suffix(".md")
    lines = [
        "# Oxford 150 Workbook Quality Audit",
        "",
        f"- Status: `{report['status']}`",
        f"- Workbooks: {report['workbook_count']}",
        f"- Checked rows: {report['checked_rows']}",
        f"- Native-speaker certification: `{report['native_speaker_certification']}`",
        f"- Severity counts: `{report['severity_counts']}`",
        f"- Blocked workbooks: {report['blocked_workbook_count']}",
        "",
        "## Finding Codes",
        "",
    ]
    for code, count in codes.most_common():
        lines.append(f"- `{code}`: {count}")
    lines += ["", "## Blocked Workbooks", ""]
    if blocked_workbooks:
        for workbook, count in sorted(blocked_workbooks.items()):
            lines.append(f"- `{workbook}`: {count}")
    else:
        lines.append("- None")
    lines += ["", "## Sample Findings", ""]
    for item in all_findings[:50]:
        lines.append(
            f"- `{item['severity']}` `{item['code']}` `{item['workbook']}` "
            f"{item['sheet']} row={item.get('row') or ''} lang={item.get('language') or ''}: {item['detail']}"
        )
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(json.dumps({k: report[k] for k in ["status", "workbook_count", "checked_rows", "severity_counts", "blocked_workbook_count"]}, ensure_ascii=False))
    return 1 if report["status"] == "blocked" else 0


if __name__ == "__main__":
    raise SystemExit(main())
