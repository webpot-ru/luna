#!/usr/bin/env python3
"""Compare a local Oxford edition workbook with a Drive-exported XLSX.

This is a fallback for the JS artifact-tool readback path when workbook import
hangs. It preserves the same cell coverage used by
check-oxford-edition-google-sheet-readback.mjs.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def normalized_cell(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return unicodedata.normalize("NFC", str(value if value is not None else "")).strip()


def col_name(index: int) -> str:
    n = index + 1
    name = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def workbook_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main_sheet_for_manifest(manifest: dict) -> str:
    name = str(manifest.get("workbook_file") or manifest.get("workbook_path") or "")
    level_match = re.search(r"(?:Core|Advanced_Extension)_([A-Z]\d)_Part_", name)
    level = level_match.group(1) if level_match else "A1"
    return f"Oxford {level} UK" if "British_English" in name else f"Oxford {level} US"


def infer_data_rows(workbook, main_sheet: str) -> int:
    sheet = workbook[main_sheet]
    last_non_empty = 0
    for row_index in range(1, 2001):
        if normalized_cell(sheet.cell(row=row_index, column=1).value):
            last_non_empty = row_index
    if last_non_empty <= 1:
        raise RuntimeError(f"Could not infer data rows from {main_sheet}!A:A")
    return last_non_empty - 1


def ranges_for_main_sheet(main_sheet: str, data_rows: int) -> list[dict]:
    rows_with_header = data_rows + 1
    return [
        {"name": main_sheet, "rows": rows_with_header, "cols": 112},
        {"name": "Course Metadata", "rows": 5, "cols": 55},
        {"name": "Deck Metadata", "rows": 42, "cols": 3},
        {"name": "Card Metadata", "rows": rows_with_header, "cols": 28},
        {"name": "_README", "rows": 24, "cols": 2},
        {"name": "_qa_status", "rows": 66, "cols": 10},
        {"name": "_languages", "rows": 55, "cols": 5},
    ]


def compare_workbooks(local_path: Path, remote_path: Path, main_sheet: str) -> tuple[int, list[dict]]:
    local_wb = load_workbook(local_path, read_only=False, data_only=False)
    remote_wb = load_workbook(remote_path, read_only=False, data_only=False)
    blockers: list[dict] = []
    checked = 0
    data_rows = infer_data_rows(local_wb, main_sheet)

    for spec in ranges_for_main_sheet(main_sheet, data_rows):
        name = spec["name"]
        if name not in local_wb.sheetnames or name not in remote_wb.sheetnames:
            blockers.append(
                {
                    "sheet": name,
                    "cell": "sheet",
                    "expected": "present",
                    "actual": "missing",
                }
            )
            break
        local_sheet = local_wb[name]
        remote_sheet = remote_wb[name]
        for row_index in range(1, spec["rows"] + 1):
            for col_index in range(1, spec["cols"] + 1):
                checked += 1
                local_value = normalized_cell(local_sheet.cell(row=row_index, column=col_index).value)
                remote_value = normalized_cell(remote_sheet.cell(row=row_index, column=col_index).value)
                if local_value != remote_value:
                    blockers.append(
                        {
                            "sheet": name,
                            "cell": f"{col_name(col_index - 1)}{row_index}",
                            "expected": local_value,
                            "actual": remote_value,
                        }
                    )
                    if len(blockers) >= 80:
                        return checked, blockers
    return checked, blockers


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", required=True)
    parser.add_argument("--remote-workbook", required=True)
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    remote_workbook = Path(args.remote_workbook)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    local_workbook = Path(manifest["workbook_path"])
    main_sheet = main_sheet_for_manifest(manifest)

    checked, blockers = compare_workbooks(local_workbook, remote_workbook, main_sheet)
    manifest.update(
        {
            "google_sheet_readback_status": "verified" if not blockers else "failed",
            "google_sheet_readback_verified_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "google_sheet_readback_sample_count": checked,
            "google_sheet_readback_workbook_sha256": workbook_sha256(local_workbook),
            "google_sheet_readback_errors": blockers,
            "google_sheet_readback_method": "drive_export_xlsx_full_oxford_edition_workbook_values_openpyxl_fallback",
        }
    )
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if blockers:
        print(json.dumps({"status": "failed", "checked_cells": checked, "blockers": blockers[:5]}, ensure_ascii=False))
        return 1
    print(json.dumps({"status": "verified", "checked_cells": checked, "manifest": str(manifest_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
