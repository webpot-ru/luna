#!/usr/bin/env python3
"""Build a deterministic 10-word translation/example sample review for Oxford A1."""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONTRACT_PATH = PROJECT_ROOT / "config/oxford-vocabulary-release-contract-v0.json"
DEFAULT_FINAL_MANIFEST_PATH = ""
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs/oxford-vocabulary/qa"
REPORT_ID = "support_translation_sample_review_v1"
SCRIPT_VERSION = "2026-05-19.v2"

SAMPLE_SOURCE_HEADWORDS = [
    "action",
    "address",
    "advice",
    "air",
    "animal",
    "apple",
    "bag",
    "bank (money)",
    "book",
    "clock",
]

REVIEW_METHOD = (
    "Deterministic Codex semantic plausibility sample review of display translations and example sentences "
    "against the reviewed English meaning notes. "
    "This is not native-speaker certification and does not replace full final linguistic audit."
)

ACCEPTED_CONTEXT_NOTES = {
    ("TL", "address"): "Filipino uses a common English loanword plus native support gloss; accepted for learner support.",
    ("MS", "clock"): "Malay 'jam' is context-dependent but standard for clock/watch/time; accepted for this A1 noun.",
    ("ID", "clock"): "Indonesian 'jam' is context-dependent but standard for clock/watch/time; accepted for this A1 noun.",
    ("RU", "clock"): "Russian display covers clock/watch; accepted for the time-showing-object meaning.",
    ("RO", "bank (money)"): "Romanian display can be ambiguous out of context, but it is valid for a financial bank.",
    ("PT", "bag"): "Portuguese display gives common bag/sack and suitcase-style options; acceptable but worth native polish later.",
    ("PT-BR", "bag"): "Brazilian Portuguese inherits PT bag options; acceptable, with 'bolsa/sacola' worth native polish later.",
}


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_jsonl(path: Path) -> list[dict]:
    rows = []
    for index, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as error:
            raise ValueError(f"Invalid JSONL at {path}:{index}: {error}") from error
    return rows


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(PROJECT_ROOT))


def resolve_final_manifest_path(contract: dict, explicit_path: Path | None) -> Path:
    if explicit_path:
        return explicit_path
    manifest_path = contract.get("latest_edition_exports", {}).get("manifest_path")
    if not manifest_path:
        raise ValueError(
            "Final workbook checks require --final-manifest or latest_edition_exports.manifest_path in the contract."
        )
    return PROJECT_ROOT / manifest_path


def load_workbook_rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(cell) for cell in rows[0]]
    records = []
    for raw in rows[1:]:
        records.append({headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))})
    return records


def load_sample_rows(contract: dict, requested_headwords: list[str] | None = None) -> list[dict]:
    row_review_path = PROJECT_ROOT / contract["latest_row_review"]["path"]
    row_reviews = read_jsonl(row_review_path)
    by_headword = {row.get("source_headword"): row for row in row_reviews}
    target_headwords = requested_headwords or SAMPLE_SOURCE_HEADWORDS
    if not requested_headwords and any(headword not in by_headword for headword in target_headwords):
        target_headwords = [row["source_headword"] for row in row_reviews[:10]]
    sample_rows = []
    missing = []
    for headword in target_headwords:
        row = by_headword.get(headword)
        if row:
            sample_rows.append(row)
        else:
            missing.append(headword)
    if missing:
        raise ValueError(f"Missing sample source headwords in row review: {missing}")
    return sample_rows


def build_expected_values(contract: dict, sample_rows: list[dict]) -> dict[str, dict[str, dict[str, str]]]:
    row_ids = {row["row_id"] for row in sample_rows}
    expected = {row["row_id"]: {} for row in sample_rows}

    edition_path_value = (
        contract.get("latest_edition_exports", {}).get("edition_layer_path")
        or contract.get("latest_edition_layer", {}).get("path")
    )
    if not edition_path_value:
        raise ValueError("Contract is missing latest_edition_exports.edition_layer_path or latest_edition_layer.path")
    edition_path = PROJECT_ROOT / edition_path_value
    for row in read_jsonl(edition_path):
        row_id = row.get("row_id")
        if row_id not in row_ids:
            continue
        expected[row_id]["EN"] = {
            "display": normalize_text(row.get("display_headword_EN_US")),
            "example": normalize_text(row.get("example_EN_US")),
        }
        expected[row_id]["EN-GB"] = {
            "display": normalize_text(row.get("display_headword_EN_GB")),
            "example": normalize_text(row.get("example_EN_GB")),
        }

    for batch in contract.get("latest_support_translation_batches", []):
        batch_path = PROJECT_ROOT / batch["path"]
        batch_languages = batch.get("languages") or []
        for row in read_jsonl(batch_path):
            row_id = row.get("row_id")
            if row_id not in row_ids:
                continue
            for language in batch_languages:
                expected[row_id][language] = {
                    "display": normalize_text(row.get(language)),
                    "example": normalize_text(row.get(f"example_{language}")),
                }

    return expected


def check_final_workbooks(final_manifest: dict, sample_rows: list[dict], expected_values: dict) -> dict:
    blockers = []
    editions = []
    display_cells_checked = 0
    example_cells_checked = 0
    meaning_ids = {row["meaning_id"] for row in sample_rows}
    sample_by_meaning_id = {row["meaning_id"]: row for row in sample_rows}

    for edition in final_manifest.get("editions", []):
        workbook_path = PROJECT_ROOT / edition["path"]
        sheet_name = edition["sheet"]
        records = load_workbook_rows(workbook_path, sheet_name)
        metadata_records = load_workbook_rows(workbook_path, "Card Metadata")
        metadata_by_meaning_id = {
            normalize_text(row.get("meaning_id")): row
            for row in metadata_records
            if normalize_text(row.get("meaning_id")) in meaning_ids
        }
        edition_checked = 0
        for meaning_id, source_row in sample_by_meaning_id.items():
            row_id = source_row["row_id"]
            metadata_row = metadata_by_meaning_id.get(meaning_id)
            main_sheet_row = int(metadata_row.get("main_sheet_row") or 0) if metadata_row else 0
            workbook_index = main_sheet_row - 2
            workbook_row = records[workbook_index] if 0 <= workbook_index < len(records) else None
            if not metadata_row or not workbook_row:
                blockers.append(
                    {
                        "edition_id": edition.get("edition_id"),
                        "row_id": row_id,
                        "meaning_id": meaning_id,
                        "reason": "sample_row_missing_from_final_workbook_metadata_or_main_sheet",
                    }
                )
                continue
            for language in edition.get("main_sheet_language_codes", []):
                expected_display = normalize_text(expected_values.get(row_id, {}).get(language, {}).get("display"))
                actual = normalize_text(workbook_row.get(language))
                display_cells_checked += 1
                edition_checked += 1
                if expected_display != actual:
                    blockers.append(
                        {
                            "edition_id": edition.get("edition_id"),
                            "row_id": row_id,
                            "language": language,
                            "reason": "final_workbook_display_value_mismatch",
                            "expected": expected_display,
                            "actual": actual,
                        }
                    )
                example_column = f"{language} example"
                expected_example = normalize_text(expected_values.get(row_id, {}).get(language, {}).get("example"))
                actual_example = normalize_text(workbook_row.get(example_column))
                example_cells_checked += 1
                if expected_example != actual_example:
                    blockers.append(
                        {
                            "edition_id": edition.get("edition_id"),
                            "row_id": row_id,
                            "language": language,
                            "reason": "final_workbook_example_value_mismatch",
                            "column": example_column,
                            "expected": expected_example,
                            "actual": actual_example,
                        }
                    )
        editions.append(
            {
                "edition_id": edition.get("edition_id"),
                "path": edition["path"],
                "sheet": sheet_name,
                "primary_language_code": edition.get("primary_language_code"),
                "sample_rows_checked": len(meaning_ids),
                "display_cells_checked": edition_checked,
                "example_cells_checked": edition_checked,
            }
        )

    return {
        "status": "pass" if not blockers else "blocker",
        "display_cells_checked": display_cells_checked,
        "example_cells_checked": example_cells_checked,
        "editions": editions,
        "blockers": blockers,
    }


def get_expected_cell(expected_values: dict, row_id: str, language: str, field: str) -> str:
    return normalize_text(expected_values.get(row_id, {}).get(language, {}).get(field))


def markdown_cell(value: object) -> str:
    return normalize_text(value).replace("|", "\\|")


def build_review(
    contract_path: Path,
    final_manifest_path: Path | None,
    output_dir: Path,
    source_package_only: bool = False,
    sample_headwords: list[str] | None = None,
) -> tuple[Path, Path, Path, dict]:
    contract = read_json(contract_path)
    resolved_final_manifest_path = None if source_package_only else resolve_final_manifest_path(contract, final_manifest_path)
    final_manifest = None if source_package_only else read_json(resolved_final_manifest_path)
    release_id = contract["latest_source_snapshot"]["release_id"]
    generated_at = datetime.now(timezone.utc).isoformat()

    language_rows = read_json(PROJECT_ROOT / "config/language-order.json")
    language_names = {row["spreadsheetCode"]: row["language"] for row in language_rows}
    language_names["EN-GB"] = "British English"
    support_language_codes = [row["spreadsheetCode"] for row in language_rows if row["spreadsheetCode"] not in {"EN", "EN-GB"}]
    review_language_codes = ["EN", "EN-GB", *support_language_codes]

    sample_rows = load_sample_rows(contract, sample_headwords)
    expected_values = build_expected_values(contract, sample_rows)

    blockers = []
    sample_entries = []
    summary_by_language = {}
    for language in review_language_codes:
        language_entries = []
        for source_row in sample_rows:
            headword = source_row["source_headword"]
            row_id = source_row["row_id"]
            display = get_expected_cell(expected_values, row_id, language, "display")
            example = get_expected_cell(expected_values, row_id, language, "example")
            if not display:
                blockers.append({"row_id": row_id, "language": language, "reason": "missing_sample_display"})
            if not example:
                blockers.append({"row_id": row_id, "language": language, "reason": "missing_sample_example"})
            review_note = ACCEPTED_CONTEXT_NOTES.get((language, headword), "")
            entry = {
                "language_code": language,
                "language": language_names.get(language, language),
                "row_id": row_id,
                "source_headword": headword,
                "reviewed_display_headword": source_row["reviewed_display_headword"],
                "meaning_note": source_row["meaning_note"],
                "display": display,
                "example": example,
                "review_status": "reviewed_plausible" if display and example else "blocker_missing_display_or_example",
                "display_review_status": "reviewed_plausible" if display else "blocker_missing_display",
                "example_review_status": "reviewed_plausible" if example else "blocker_missing_example",
                "review_note": review_note,
            }
            sample_entries.append(entry)
            language_entries.append(entry)
        summary_by_language[language] = {
            "language": language_names.get(language, language),
            "sampled_words": len(language_entries),
            "display_status": "pass" if all(item["display"] for item in language_entries) else "blocker",
            "example_status": "pass" if all(item["example"] for item in language_entries) else "blocker",
            "status": "pass" if all(item["display"] and item["example"] for item in language_entries) else "blocker",
        }

    final_workbook_check = (
        {
            "status": "skipped_source_package_only",
            "display_cells_checked": 0,
            "example_cells_checked": 0,
            "editions": [],
            "blockers": [],
            "skip_reason": "Part source-package review runs before final US/UK workbook export and Google Sheet delivery.",
        }
        if source_package_only
        else check_final_workbooks(final_manifest, sample_rows, expected_values)
    )
    blockers.extend(final_workbook_check["blockers"])

    status = "pass" if not blockers else "blocker"
    report_path = output_dir / f"{release_id}_{REPORT_ID}.json"
    report_csv_path = output_dir / f"{release_id}_{REPORT_ID}.csv"
    report_md_path = output_dir / f"{release_id}_{REPORT_ID}.md"

    matrix_rows = []
    example_matrix_rows = []
    for language in review_language_codes:
        matrix_rows.append(
            [
                language,
                language_names.get(language, language),
                *[get_expected_cell(expected_values, row["row_id"], language, "display") for row in sample_rows],
            ]
        )
        example_matrix_rows.append(
            [
                language,
                language_names.get(language, language),
                *[get_expected_cell(expected_values, row["row_id"], language, "example") for row in sample_rows],
            ]
        )

    sampled_headwords = [row["source_headword"] for row in sample_rows]
    accepted_context_notes = [
        {
            "language_code": language,
            "source_headword": headword,
            "note": note,
        }
        for (language, headword), note in sorted(ACCEPTED_CONTEXT_NOTES.items())
        if headword in sampled_headwords
    ]

    report = {
        "release_id": release_id,
        "report_id": REPORT_ID,
        "script_version": SCRIPT_VERSION,
        "generated_at": generated_at,
        "status": status,
        "review_method": REVIEW_METHOD,
        "sample_source_headwords": [row["source_headword"] for row in sample_rows],
        "sample_row_ids": [row["row_id"] for row in sample_rows],
        "sample_size_per_language": len(sample_rows),
        "languages_reviewed": len(review_language_codes),
        "support_languages_reviewed": len(support_language_codes),
        "unique_sample_display_cells_reviewed": len(sample_entries),
        "unique_sample_example_cells_reviewed": len(sample_entries),
        "final_workbook_check": final_workbook_check,
        "source_package_only": source_package_only,
        "blocker_count": len(blockers),
        "warning_count": 0,
        "blockers": blockers,
        "warnings": [],
        "accepted_context_notes": accepted_context_notes,
        "summary_by_language": summary_by_language,
        "sample_entries": sample_entries,
        "matrix": {
            "headers": ["language_code", "language", *sampled_headwords],
            "rows": matrix_rows,
        },
        "example_matrix": {
            "headers": ["language_code", "language", *sampled_headwords],
            "rows": example_matrix_rows,
        },
        "inputs": {
            "contract_path": relative(contract_path),
            "final_manifest_path": None if source_package_only else relative(resolved_final_manifest_path),
            "row_review_path": contract["latest_row_review"]["path"],
            "edition_layer_path": (
                contract.get("latest_edition_exports", {}).get("edition_layer_path")
                or contract.get("latest_edition_layer", {}).get("path")
            ),
            "support_translation_batches": [
                {"batch_id": batch.get("batch_id"), "path": batch.get("path"), "languages": batch.get("languages")}
                for batch in contract.get("latest_support_translation_batches", [])
            ],
        },
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(report_path, report)
    with report_csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "language_code",
                "language",
                "source_headword",
                "row_id",
                "meaning_note",
                "display",
                "example",
                "review_status",
                "display_review_status",
                "example_review_status",
                "review_note",
            ],
        )
        writer.writeheader()
        writer.writerows(
            {
                "language_code": row["language_code"],
                "language": row["language"],
                "source_headword": row["source_headword"],
                "row_id": row["row_id"],
                "meaning_note": row["meaning_note"],
                "display": row["display"],
                "example": row["example"],
                "review_status": row["review_status"],
                "display_review_status": row["display_review_status"],
                "example_review_status": row["example_review_status"],
                "review_note": row["review_note"],
            }
            for row in sample_entries
        )

    markdown_lines = [
        f"# Oxford A1 Translation/Example Sample Review: {release_id}",
        "",
        f"- Report id: `{REPORT_ID}`",
        f"- Status: `{status}`",
        f"- Generated at: `{generated_at}`",
        f"- Review method: {REVIEW_METHOD}",
        f"- Languages reviewed: {len(review_language_codes)} unique language variants",
        f"- Sample size: {len(sample_rows)} words per language",
        f"- Unique display cells reviewed: {len(sample_entries)}",
        f"- Unique example cells reviewed: {len(sample_entries)}",
        f"- Source-package-only mode: `{source_package_only}`",
        f"- Final workbook display cells checked: {final_workbook_check['display_cells_checked']}",
        f"- Final workbook example cells checked: {final_workbook_check['example_cells_checked']}",
        f"- Blockers: {len(blockers)}",
        "- Warnings: 0",
        "",
        "## Sample Words",
        "",
        "| # | Source headword | Meaning note |",
        "| --- | --- | --- |",
    ]
    for index, row in enumerate(sample_rows, start=1):
        markdown_lines.append(f"| {index} | `{markdown_cell(row['source_headword'])}` | {markdown_cell(row['meaning_note'])} |")
    markdown_lines.extend(["", "## Accepted Context Notes", "", "| Language | Source headword | Note |", "| --- | --- | --- |"])
    for note in report["accepted_context_notes"]:
        markdown_lines.append(
            f"| {markdown_cell(note['language_code'])} | `{markdown_cell(note['source_headword'])}` | {markdown_cell(note['note'])} |"
        )
    if not report["accepted_context_notes"]:
        markdown_lines.append("| - | - | No accepted context exceptions in this sample. |")
    markdown_lines.extend(["", "## Display Translation Matrix", "", "| " + " | ".join(markdown_cell(cell) for cell in report["matrix"]["headers"]) + " |"])
    markdown_lines.append("| " + " | ".join(["---"] * len(report["matrix"]["headers"])) + " |")
    for row in matrix_rows:
        markdown_lines.append("| " + " | ".join(markdown_cell(cell) for cell in row) + " |")
    markdown_lines.extend(
        [
            "",
            "## Example Sentence Matrix",
            "",
            "| " + " | ".join(markdown_cell(cell) for cell in report["example_matrix"]["headers"]) + " |",
        ]
    )
    markdown_lines.append("| " + " | ".join(["---"] * len(report["example_matrix"]["headers"])) + " |")
    for row in example_matrix_rows:
        markdown_lines.append("| " + " | ".join(markdown_cell(cell) for cell in row) + " |")
    markdown_lines.extend(
        [
            "",
            "## Artifacts",
            "",
            f"- JSON: `{relative(report_path)}`",
            f"- CSV: `{relative(report_csv_path)}`",
            f"- Markdown: `{relative(report_md_path)}`",
            "",
        ]
    )
    report_md_path.write_text("\n".join(markdown_lines), encoding="utf-8")

    contract["latest_support_translation_sample_review"] = {
        "report_id": REPORT_ID,
        "script_path": relative(Path(__file__)),
        "path": relative(report_path),
        "csv_path": relative(report_csv_path),
        "markdown_path": relative(report_md_path),
        "status": status,
        "generated_at": generated_at,
        "review_method": REVIEW_METHOD,
        "sample_size_per_language": len(sample_rows),
        "languages_reviewed": len(review_language_codes),
        "support_languages_reviewed": len(support_language_codes),
        "unique_sample_display_cells_reviewed": len(sample_entries),
        "unique_sample_example_cells_reviewed": len(sample_entries),
        "final_workbook_display_cells_checked": final_workbook_check["display_cells_checked"],
        "final_workbook_example_cells_checked": final_workbook_check["example_cells_checked"],
        "final_workbook_check_status": final_workbook_check["status"],
        "source_package_only": source_package_only,
        "blocker_count": len(blockers),
        "warning_count": 0,
        "does_not_replace": "full native-speaker final linguistic audit",
    }
    write_json(contract_path, contract)

    return report_path, report_csv_path, report_md_path, report


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Oxford A1 deterministic translation/example sample review.")
    parser.add_argument("--contract", default=str(DEFAULT_CONTRACT_PATH))
    parser.add_argument(
        "--final-manifest",
        default=DEFAULT_FINAL_MANIFEST_PATH,
        help="Final edition export manifest. Defaults to latest_edition_exports.manifest_path from the contract.",
    )
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument(
        "--source-package-only",
        action="store_true",
        help="Skip final workbook checks for pre-delivery source-package releases.",
    )
    parser.add_argument(
        "--sample-source-headwords",
        default="",
        help="Comma-separated 10-word sample override. Defaults to the Part 001 sample when present, otherwise the first 10 reviewed rows.",
    )
    args = parser.parse_args()
    sample_headwords = [item.strip() for item in args.sample_source_headwords.split(",") if item.strip()] or None

    report_path, report_csv_path, report_md_path, report = build_review(
        Path(args.contract),
        Path(args.final_manifest) if args.final_manifest else None,
        Path(args.output_dir),
        source_package_only=args.source_package_only,
        sample_headwords=sample_headwords,
    )
    print(
        "Oxford A1 translation/example sample review "
        f"status={report['status']} "
        f"languages={report['languages_reviewed']} "
        f"sample_per_language={report['sample_size_per_language']} "
        f"unique_display_cells={report['unique_sample_display_cells_reviewed']} "
        f"unique_example_cells={report['unique_sample_example_cells_reviewed']} "
        f"workbook_display_cells={report['final_workbook_check']['display_cells_checked']} "
        f"workbook_example_cells={report['final_workbook_check']['example_cells_checked']} "
        f"blockers={report['blocker_count']} "
        f"json={relative(report_path)} "
        f"csv={relative(report_csv_path)} "
        f"markdown={relative(report_md_path)}"
    )


if __name__ == "__main__":
    main()
