#!/usr/bin/env python3
import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


REPORT_ID = "oxford_english_learning_gates_v1"
SCRIPT_VERSION = "2026-05-31.v9"
LEVEL_ORDER = ["A1", "A2", "B1", "B2", "C1", "C2"]
FORBIDDEN_CLAIM_RE = re.compile(r"\b(official|certified|approved by oxford|endorsed by oxford)\b", re.I)
PLACEHOLDER_RE = re.compile(r"\b(needs_review|needs row review|needs_lunacards_authored_example)\b", re.I)
SENTENCE_END_RE = re.compile(r"[.!?]$")
IPA_VALUE_RE = re.compile(r"^/[^/]+/$")
REQUIRED_EN_PRONUNCIATION_SOURCE_IDS = {
    "ipa-focused-english-us-cmudict-dict",
    "ipa-focused-english-us-cmudict-phones",
    "ipa-focused-english-us-cmudict-symbols",
}
OXFORD_EDITION_PRONUNCIATION_HEADERS = [
    "EN transcription",
    "EN-GB transcription",
    "EN example transcription",
    "EN-GB example transcription",
]
REQUIRED_CARD_METADATA_HEADERS = {
    "main_sheet_row",
    "display_order",
    "card_key",
    "set_id",
    "meaning_id",
    "canonical_english",
    "english_with_article",
    "part_of_speech",
    "level",
    "frequency_band",
    "priority_band",
    "domain",
    "area",
    "category",
    "context_domain",
    "context_area",
    "context_category",
    "countability",
    "plural_form_en",
    "semantic_class",
    "tags",
    "meaning_note",
    "context_note",
    "context_example_id",
    "context_example_key",
    "context_example_quality_status",
    "meaning_quality_status",
    "membership_quality_status",
}
POST_FINAL_SPOT_CHECK_REPORT_ID = "post_final_language_spot_check_v1"
POST_FINAL_SPOT_CHECK_MIN_SAMPLE_SIZE = 5
POST_FINAL_WARNING_BUDGET = {
    "max_total_warning_codes": 250,
    "max_source_no_candidate_codes": 120,
    "max_weak_language_codes": 80,
    "max_target_example_lexical_anchor_codes": 30,
    "max_exact_english_surface_loan_review_codes": 0,
}


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def read_jsonl(path):
    rows = []
    for index, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as error:
            raise ValueError(f"Invalid JSONL at {path}:{index}: {error}") from error
    return rows


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rel(path):
    return str(Path(path).resolve().relative_to(Path.cwd()))


def add_gate(results, layer, gate, status, message, **detail):
    item = {"layer": layer, "gate": gate, "status": status, "message": message}
    if detail:
        item.update(detail)
    results.append(item)


def main_sheet_language_codes_for_primary(primary_language_code, language_codes):
    support_codes = [code for code in language_codes if code not in {"EN", "EN-GB"}]
    return [primary_language_code, *support_codes]


def pronunciation_headers_for_primary(primary_language_code):
    return [f"{primary_language_code} transcription", f"{primary_language_code} example transcription"]


def opposite_english_code(primary_language_code):
    return "EN-GB" if primary_language_code == "EN" else "EN"


def infer_part_number(release_id):
    match = re.search(r"_part_(\d+)_", release_id)
    if not match:
        return "1"
    return str(int(match.group(1)))


def infer_course_level(release_id):
    match = re.search(r"_(?:core|advanced)_([abc]\d)_(?:extension_)?part_", release_id, flags=re.IGNORECASE)
    if not match:
        return "A1"
    return match.group(1).upper()


def infer_wordlist_token(release_id):
    return "5000" if "oxford_5000" in release_id else "3000"


def row_sample(rows, field="row_id", limit=8):
    sample = []
    for row in rows[:limit]:
        sample.append(row.get(field) or row.get("core_item_id") or row.get("source_headword") or "?")
    return sample


def load_workbook_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    main_sheet_name = workbook.sheetnames[0]
    sheet = workbook[main_sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return workbook, main_sheet_name, [], []
    headers = [normalize_text(cell) for cell in rows[0]]
    data = []
    for raw in rows[1:]:
        data.append({headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))})
    return workbook, main_sheet_name, headers, data


def has_forbidden_claim(value):
    return bool(FORBIDDEN_CLAIM_RE.search(normalize_text(value)))


def word_count(sentence):
    return len(re.findall(r"[A-Za-z0-9]+(?:'[A-Za-z0-9]+)?", normalize_text(sentence)))


def valid_ipa_cell(value):
    parts = [part.strip() for part in normalize_text(value).split(",") if part.strip()]
    return bool(parts) and all(IPA_VALUE_RE.match(part) for part in parts)


def iter_advisory_warning_codes(report):
    for item in report.get("advisory_warnings") or []:
        for code in item.get("warnings") or []:
            yield normalize_text(code)


def summarize_advisory_warning_codes(report):
    counts = {}
    for code in iter_advisory_warning_codes(report):
        counts[code] = counts.get(code, 0) + 1
    return counts


def main():
    parser = argparse.ArgumentParser(
        description="Check Oxford / English-learning release-specific gates without touching ordinary polyglot gates."
    )
    parser.add_argument("--contract", default="config/oxford-vocabulary-release-contract-v0.json")
    parser.add_argument("--release", default="")
    parser.add_argument("--output-dir", default="outputs/oxford-vocabulary/qa")
    args = parser.parse_args()

    contract_path = Path(args.contract)
    contract = read_json(contract_path)
    snapshot = contract["latest_source_snapshot"]
    release_id = args.release or snapshot["release_id"]
    output_dir = Path(args.output_dir)
    report_json_path = output_dir / f"{release_id}_{REPORT_ID}.json"
    report_md_path = output_dir / f"{release_id}_{REPORT_ID}.md"
    release_part_number = infer_part_number(release_id)
    release_course_level = infer_course_level(release_id)
    release_wordlist_token = infer_wordlist_token(release_id)

    gate_results = []
    checked_artifacts = []

    def check_path(gate, artifact_path):
        path = Path(artifact_path or "")
        checked_artifacts.append({"gate": gate, "path": str(path)})
        if not artifact_path:
            add_gate(gate_results, "artifact", gate, "blocker", "Contract is missing artifact path.")
            return None
        if not path.exists():
            add_gate(gate_results, "artifact", gate, "blocker", "Artifact path does not exist.", path=str(path))
            return None
        return path

    manifest_path = check_path("source_list_snapshot", snapshot.get("manifest_path"))
    normalized_path = check_path("source_list_snapshot", snapshot.get("normalized_path"))
    candidate_path = check_path("source_list_snapshot", snapshot.get("candidate_pool_path"))
    workbook_path = check_path("workbook_contract_check", snapshot.get("workbook_path"))
    permission_request_path = check_path("source_license_review", contract.get("permission_request", {}).get("request_artifact"))
    row_review_path_value = contract.get("latest_row_review", {}).get("path")
    row_review_path = check_path("learner_sense_pos_check", row_review_path_value) if row_review_path_value else None
    english_examples_path_value = contract.get("latest_english_examples", {}).get("path")
    english_examples_path = (
        check_path("english_example_quality_check", english_examples_path_value)
        if english_examples_path_value
        else None
    )
    english_pronunciation_path_value = contract.get("latest_english_pronunciation", {}).get("path")
    english_pronunciation_path = (
        check_path("english_pronunciation_source_check", english_pronunciation_path_value)
        if english_pronunciation_path_value
        else None
    )
    support_translation_paths = []
    for batch in contract.get("latest_support_translation_batches", []):
        batch_path = check_path("support_translation_meaning_check", batch.get("path"))
        if batch_path:
            support_translation_paths.append((batch, batch_path))

    manifest = read_json(manifest_path) if manifest_path else {}
    normalized_rows = read_jsonl(normalized_path) if normalized_path else []
    candidate_rows = read_jsonl(candidate_path) if candidate_path else []
    row_review_rows = read_jsonl(row_review_path) if row_review_path else []
    row_review_by_id = {row.get("row_id"): row for row in row_review_rows}
    english_example_rows = read_jsonl(english_examples_path) if english_examples_path else []
    english_pronunciation_rows = read_jsonl(english_pronunciation_path) if english_pronunciation_path else []
    support_translation_batches = [
        {"contract": batch, "path": path, "rows": read_jsonl(path)} for batch, path in support_translation_paths
    ]
    language_order = read_json(Path("config/language-order.json"))
    language_codes = [row["spreadsheetCode"] for row in language_order]
    target_language_codes = [code for code in language_codes if code not in {"EN", "EN-GB"}]

    workbook = None
    workbook_sheet = ""
    workbook_headers = []
    workbook_rows = []
    if workbook_path:
        workbook, workbook_sheet, workbook_headers, workbook_rows = load_workbook_rows(workbook_path)

    source_policy = contract.get("source_policy", {})
    permission_status = contract.get("permission_request", {}).get("status", "")
    written_or_project_evidence = permission_status in {
        "written_evidence_stored",
        "project_evidence_decision_accepted",
        "written_evidence_stored_and_reviewed",
    }
    user_reported_permission = "user_reported" in permission_status or written_or_project_evidence

    if user_reported_permission:
        add_gate(
            gate_results,
            "source_and_legal",
            "source_license_review",
            "pass",
            "User-reported OUP permission state is recorded for local source-package work.",
            permission_status=permission_status,
        )
    else:
        add_gate(
            gate_results,
            "source_and_legal",
            "source_license_review",
            "blocker",
            "No user-reported or written Oxford permission state is recorded.",
            permission_status=permission_status,
        )

    if written_or_project_evidence:
        add_gate(
            gate_results,
            "source_and_legal",
            "permission_evidence_or_project_evidence_decision",
            "pass",
            "Written OUP evidence or explicit project evidence decision is recorded.",
        )
    else:
        add_gate(
            gate_results,
            "source_and_legal",
            "permission_evidence_or_project_evidence_decision",
            "blocker",
            "Written OUP evidence or explicit project evidence decision is not stored; final learner-facing delivery remains blocked.",
            permission_status=permission_status,
        )

    allowed_fields = source_policy.get("allowed_fields") or source_policy.get("licensed_allowed_fields")
    if allowed_fields:
        add_gate(
            gate_results,
            "source_and_legal",
            "allowed_fields_review",
            "pass",
            "Allowed Oxford fields are recorded.",
            allowed_fields=allowed_fields,
        )
    else:
        add_gate(
            gate_results,
            "source_and_legal",
            "allowed_fields_review",
            "blocker",
            "Allowed Oxford fields are not recorded as a reviewed grant; only requested fields are listed.",
        )

    copied_flag = bool(manifest.get("copied_definitions_examples_pronunciations")) or bool(
        snapshot.get("copied_definitions_examples_pronunciations")
    )
    forbidden_data_fields = []
    for row in candidate_rows:
        for field in row:
            if any(token in field.lower() for token in ["definition", "pronunciation", "audio"]):
                forbidden_data_fields.append(field)
    if copied_flag or forbidden_data_fields:
        add_gate(
            gate_results,
            "source_and_legal",
            "no_oxford_content_copy_check",
            "blocker",
            "Candidate/source data appears to include copied Oxford definitions, pronunciations or audio fields.",
            copied_flag=copied_flag,
            fields=sorted(set(forbidden_data_fields))[:20],
        )
    else:
        add_gate(
            gate_results,
            "source_and_legal",
            "no_oxford_content_copy_check",
            "pass",
            "No copied Oxford definitions/examples/pronunciation/audio fields are present in the source draft artifacts.",
        )

    claim_hits = []
    for value in [
        contract.get("course", {}).get("product_safe_working_title", ""),
        contract.get("permission_request", {}).get("user_intent", ""),
        manifest.get("selected_scope", ""),
    ]:
        if has_forbidden_claim(value):
            claim_hits.append(value)
    attribution_wording = source_policy.get("attribution_no_endorsement_wording")
    if claim_hits:
        add_gate(
            gate_results,
            "source_and_legal",
            "attribution_no_endorsement_review",
            "blocker",
            "Forbidden Oxford endorsement/official claim found.",
            claims=claim_hits,
        )
    elif attribution_wording:
        add_gate(
            gate_results,
            "source_and_legal",
            "attribution_no_endorsement_review",
            "pass",
            "Final attribution/no-endorsement wording is recorded and no forbidden endorsement claim was found.",
            attribution_no_endorsement_wording=attribution_wording,
        )
    else:
        add_gate(
            gate_results,
            "source_and_legal",
            "attribution_no_endorsement_review",
            "warning",
            "No forbidden endorsement claim found, but final attribution/no-endorsement wording is not yet recorded.",
        )

    artifact_errors = [item for item in gate_results if item["layer"] == "artifact" and item["status"] == "blocker"]
    if not artifact_errors:
        source_path = Path(manifest.get("source_path", ""))
        raw_sha_ok = source_path.exists() and sha256_file(source_path) == manifest.get("raw_file_sha256")
        counts_ok = (
            manifest.get("release_id") == release_id
            and manifest.get("course_id") == contract.get("course", {}).get("course_id")
            and manifest.get("normalized_rows") == len(normalized_rows)
            and manifest.get("selected_rows") == len(candidate_rows)
            and snapshot.get("selected_rows") == len(candidate_rows)
            and raw_sha_ok
        )
        add_gate(
            gate_results,
            "english_learning_core",
            "source_list_snapshot",
            "pass" if counts_ok else "blocker",
            "Source manifest, normalized source, candidate pool and raw file hash are consistent."
            if counts_ok
            else "Source manifest, normalized source, candidate pool or raw file hash is inconsistent.",
            normalized_rows=len(normalized_rows),
            selected_rows=len(candidate_rows),
            manifest_selected_rows=manifest.get("selected_rows"),
            raw_sha_ok=raw_sha_ok,
        )

    expected_release = release_id
    expected_course = contract.get("course", {}).get("course_id")
    row_ids = [row.get("row_id") for row in candidate_rows]
    core_ids = [row.get("core_item_id") for row in candidate_rows]
    meaning_ids = [row.get("meaning_id") for row in candidate_rows]
    surface_pos_keys = [
        (normalize_text(row.get("normalized_headword") or row.get("source_headword")).lower(), normalize_text(row.get("part_of_speech")))
        for row in candidate_rows
    ]
    identity_ok = (
        len(row_ids) == len(set(row_ids))
        and len(core_ids) == len(set(core_ids))
        and len(meaning_ids) == len(set(meaning_ids))
        and len(surface_pos_keys) == len(set(surface_pos_keys))
        and all(row.get("release_id") == expected_release for row in candidate_rows)
        and all(row.get("course_id") == expected_course for row in candidate_rows)
    )
    add_gate(
        gate_results,
        "english_learning_core",
        "row_identity_check",
        "pass" if identity_ok else "blocker",
        "Candidate rows have unique row/core/meaning ids and no duplicate surface+POS keys."
        if identity_ok
        else "Candidate rows have identity mismatch or duplicate ids/surface+POS keys.",
        rows=len(candidate_rows),
    )

    level_filter = set(manifest.get("level_filter") or snapshot.get("level_filter") or [])
    selected_level_min_values = {row.get("level_min") for row in candidate_rows}
    selected_level_values = selected_level_min_values | {row.get("level_max") for row in candidate_rows}
    broader_level_rows = [row for row in candidate_rows if row.get("level_min") != row.get("level_max")]
    level_values_ok = selected_level_min_values <= level_filter and bool(level_filter)
    valid_levels = selected_level_values <= set(LEVEL_ORDER)
    cefr_status = "pass" if level_values_ok and valid_levels else "blocker"
    add_gate(
        gate_results,
        "english_learning_core",
        "cefr_level_fit_check",
        cefr_status,
        "All selected source rows have level_min inside the intended CEFR level filter."
        if cefr_status == "pass"
        else "Selected source rows include level_min values outside the intended filter or invalid CEFR labels.",
        level_filter=sorted(level_filter),
        selected_level_min_values=sorted(value for value in selected_level_min_values if value),
        selected_level_values=sorted(value for value in selected_level_values if value),
        broader_level_span_rows=len(broader_level_rows),
    )
    if broader_level_rows and cefr_status == "pass":
        reviewed_broader_rows = [
            row
            for row in broader_level_rows
            if normalize_text(row_review_by_id.get(row.get("row_id"), {}).get("level_span_decision"))
            in {"reviewed_beginner_safe_sense_selected", "reviewed_no_issue", "reviewed"}
        ]
        span_status = "pass" if len(reviewed_broader_rows) == len(broader_level_rows) else "warning"
        add_gate(
            gate_results,
            "english_learning_core",
            "cefr_level_span_review",
            span_status,
            "All broader level-span rows have explicit level-appropriate learner-sense review."
            if span_status == "pass"
            else "Some selected rows have broader level_max values because the source row combines multiple POS/sense levels; learner_sense_pos_check must choose the level-appropriate learner sense.",
            count=len(broader_level_rows),
            reviewed=len(reviewed_broader_rows),
            sample=row_sample(broader_level_rows),
        )

    row_review_artifact_ok = False
    if row_review_rows:
        review_ids = [row.get("row_id") for row in row_review_rows]
        candidate_by_id = {row.get("row_id"): row for row in candidate_rows}
        mismatches = []
        for row in row_review_rows:
            source_row = candidate_by_id.get(row.get("row_id"))
            if not source_row:
                mismatches.append({"row_id": row.get("row_id"), "reason": "unknown_row_id"})
                continue
            for field in ["release_id", "course_id", "core_item_id", "source_headword"]:
                if normalize_text(row.get(field)) != normalize_text(source_row.get(field)):
                    mismatches.append({"row_id": row.get("row_id"), "reason": f"{field}_mismatch"})
        row_review_artifact_ok = (
            len(row_review_rows) == len(candidate_rows)
            and len(review_ids) == len(set(review_ids))
            and len([row.get("meaning_id") for row in row_review_rows]) == len({row.get("meaning_id") for row in row_review_rows})
            and not mismatches
        )
        add_gate(
            gate_results,
            "english_learning_core",
            "row_review_artifact_check",
            "pass" if row_review_artifact_ok else "blocker",
            "Row review artifact covers every selected candidate row and matches source identity."
            if row_review_artifact_ok
            else "Row review artifact is missing rows, has duplicates or does not match source identity.",
            rows=len(row_review_rows),
            expected=len(candidate_rows),
            mismatches=mismatches[:10],
        )

    rows_for_sense_review = row_review_rows if row_review_artifact_ok else candidate_rows
    placeholder_rows = [
        row
        for row in rows_for_sense_review
        if PLACEHOLDER_RE.search(
            " ".join(
                [
                    normalize_text(row.get("meaning_id")),
                    normalize_text(row.get("sense_no")),
                    normalize_text(row.get("meaning_note")),
                    normalize_text(row.get("reviewed_part_of_speech")),
                    json.dumps(row.get("semantic_scene", ""), ensure_ascii=False),
                ]
            )
        )
    ]
    missing_sense_fields = [
        row
        for row in rows_for_sense_review
        if not all(
            normalize_text(row.get(field))
            for field in ["meaning_id", "sense_no", "meaning_note", "reviewed_part_of_speech", "semantic_scene"]
        )
    ]
    if placeholder_rows or missing_sense_fields or not row_review_artifact_ok:
        add_gate(
            gate_results,
            "english_learning_core",
            "learner_sense_pos_check",
            "blocker",
            "Rows still have placeholder or missing sense/POS/meaning review markers.",
            placeholder_count=len(placeholder_rows),
            missing_field_count=len(missing_sense_fields),
            sample=row_sample(placeholder_rows or missing_sense_fields),
        )
    else:
        add_gate(
            gate_results,
            "english_learning_core",
            "learner_sense_pos_check",
            "pass",
            "Rows no longer contain placeholder sense/POS/meaning review markers.",
        )

    learner_value_reviewed = [
        row
        for row in rows_for_sense_review
        if normalize_text(row.get("learner_value_status")) in {"reviewed", "approved", "generated_checked"}
    ]
    if len(learner_value_reviewed) == len(candidate_rows) and candidate_rows and row_review_artifact_ok:
        add_gate(
            gate_results,
            "english_learning_core",
            "learner_value_check",
            "pass",
            "All candidate rows have explicit learner-value review status.",
        )
    else:
        add_gate(
            gate_results,
            "english_learning_core",
            "learner_value_check",
            "blocker",
            "Rows do not yet have explicit English learner-value review status.",
            reviewed=len(learner_value_reviewed),
            expected=len(candidate_rows),
        )

    en_variant_ok = contract.get("course", {}).get("source_language_code") == "EN" and contract.get("course", {}).get(
        "regional_variant_code"
    ) == "EN-GB"
    add_gate(
        gate_results,
        "english_learning_core",
        "english_variant_check",
        "pass" if en_variant_ok else "blocker",
        "EN is configured as the source/default layer and EN-GB as a documented buyer-facing edition variant."
        if en_variant_ok
        else "EN/EN-GB course variant settings are not configured as expected.",
    )

    pronunciation_artifact_ok = False
    if english_pronunciation_rows:
        pronunciation_ids = [row.get("row_id") for row in english_pronunciation_rows]
        example_by_id = {row.get("row_id"): row for row in english_example_rows}
        pronunciation_mismatches = []
        pronunciation_quality_problems = []
        source_path = Path("reference-sources/raw/ipa-focused/english-us/cmudict.dict")
        for row in english_pronunciation_rows:
            example_row = example_by_id.get(row.get("row_id"))
            if not example_row:
                pronunciation_mismatches.append({"row_id": row.get("row_id"), "reason": "unknown_row_id"})
                continue
            for field in ["release_id", "course_id", "core_item_id", "meaning_id", "source_headword"]:
                if normalize_text(row.get(field)) != normalize_text(example_row.get(field)):
                    pronunciation_mismatches.append({"row_id": row.get("row_id"), "reason": f"{field}_mismatch"})
            if not valid_ipa_cell(row.get("transcription_EN")):
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "invalid_transcription_EN"})
            if not valid_ipa_cell(row.get("example_transcription_EN")):
                pronunciation_quality_problems.append(
                    {"row_id": row.get("row_id"), "reason": "invalid_example_transcription_EN"}
                )
            if normalize_text(row.get("transcription_status")) != "source_backed_cmudict_component_exact":
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "bad_transcription_status"})
            if normalize_text(row.get("example_transcription_status")) != "source_backed_cmudict_component_exact":
                pronunciation_quality_problems.append(
                    {"row_id": row.get("row_id"), "reason": "bad_example_transcription_status"}
                )
            source_ids = set(row.get("source_ids") or [])
            if not REQUIRED_EN_PRONUNCIATION_SOURCE_IDS <= source_ids:
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "missing_cmudict_source_ids"})
            if not source_path.exists() or normalize_text(row.get("source_path")) != str(source_path):
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "bad_cmudict_source_path"})
            if row.get("does_not_use_oxford_pronunciation") is not True:
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "oxford_pronunciation_not_excluded"})
            evidence = list(row.get("display_pronunciation_evidence") or []) + list(
                row.get("example_pronunciation_evidence") or []
            )
            if not evidence:
                pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "missing_cmudict_evidence"})
            for item in evidence:
                if (
                    not normalize_text(item.get("token"))
                    or not normalize_text(item.get("arpabet"))
                    or not normalize_text(item.get("ipa"))
                    or item.get("source_id") != "ipa-focused-english-us-cmudict-dict"
                ):
                    pronunciation_quality_problems.append({"row_id": row.get("row_id"), "reason": "malformed_cmudict_evidence"})
                    break
        pronunciation_artifact_ok = (
            len(english_pronunciation_rows) == len(candidate_rows)
            and len(pronunciation_ids) == len(set(pronunciation_ids))
            and english_example_rows
            and not pronunciation_mismatches
            and not pronunciation_quality_problems
        )
        add_gate(
            gate_results,
            "english_learning_core",
            "english_pronunciation_source_check",
            "pass" if pronunciation_artifact_ok else "blocker",
            "Source-backed EN-US pronunciation/IPA evidence exists for every reviewed row and uses local CMUdict evidence."
            if pronunciation_artifact_ok
            else "EN pronunciation artifact is missing rows, mismatches examples or lacks valid CMUdict evidence.",
            rows=len(english_pronunciation_rows),
            expected=len(candidate_rows),
            mismatches=pronunciation_mismatches[:10],
            quality_problems=pronunciation_quality_problems[:10],
        )
    else:
        add_gate(
            gate_results,
            "english_learning_core",
            "english_pronunciation_source_check",
            "blocker",
            "No source-backed EN pronunciation/IPA evidence exists yet.",
        )

    example_artifact_ok = False
    if english_example_rows:
        example_ids = [row.get("row_id") for row in english_example_rows]
        review_by_id = {row.get("row_id"): row for row in row_review_rows}
        example_mismatches = []
        example_quality_problems = []
        for row in english_example_rows:
            review_row = review_by_id.get(row.get("row_id"))
            if not review_row:
                example_mismatches.append({"row_id": row.get("row_id"), "reason": "unknown_row_id"})
                continue
            for field in ["release_id", "course_id", "core_item_id", "meaning_id", "source_headword"]:
                if normalize_text(row.get(field)) != normalize_text(review_row.get(field)):
                    example_mismatches.append({"row_id": row.get("row_id"), "reason": f"{field}_mismatch"})
            example = normalize_text(row.get("example_EN"))
            if not example:
                example_quality_problems.append({"row_id": row.get("row_id"), "reason": "missing_example"})
            elif not SENTENCE_END_RE.search(example):
                example_quality_problems.append({"row_id": row.get("row_id"), "reason": "missing_sentence_punctuation"})
            elif word_count(example) > 10:
                example_quality_problems.append(
                    {"row_id": row.get("row_id"), "reason": "example_too_long", "word_count": word_count(example)}
                )
            if normalize_text(row.get("example_source")) != "lunacards_authored_not_oxford":
                example_quality_problems.append({"row_id": row.get("row_id"), "reason": "example_source_not_lunacards"})
            if normalize_text(row.get("example_quality_status")) not in {"reviewed", "approved", "generated_checked"}:
                example_quality_problems.append({"row_id": row.get("row_id"), "reason": "example_not_reviewed"})
        example_artifact_ok = (
            len(english_example_rows) == len(candidate_rows)
            and len(example_ids) == len(set(example_ids))
            and row_review_artifact_ok
            and not example_mismatches
            and not example_quality_problems
        )
        add_gate(
            gate_results,
            "english_learning_core",
            "english_example_artifact_check",
            "pass" if example_artifact_ok else "blocker",
            "English example artifact covers every reviewed row and matches row-review identity."
            if example_artifact_ok
            else "English example artifact is missing rows, has duplicates, mismatches row review or has quality problems.",
            rows=len(english_example_rows),
            expected=len(candidate_rows),
            mismatches=example_mismatches[:10],
            quality_problems=example_quality_problems[:10],
        )

    example_rows = [row for row in english_example_rows if normalize_text(row.get("example_EN"))]
    if len(example_rows) == len(candidate_rows) and candidate_rows and example_artifact_ok:
        add_gate(
            gate_results,
            "english_learning_core",
            "english_example_quality_check",
            "pass",
            "All rows have short reviewed LunaCards-authored English examples.",
        )
    else:
        add_gate(
            gate_results,
            "english_learning_core",
            "english_example_quality_check",
            "blocker",
            "LunaCards-authored English examples are missing for selected rows.",
            filled=len(example_rows),
            expected=len(candidate_rows),
        )

    expected_benchmark_membership = normalize_text(
        manifest.get("benchmark_membership")
        or snapshot.get("benchmark_membership")
        or ("oxford_5000_extension" if expected_course == "oxford_5000_advanced_extension" else "oxford_3000")
    )
    benchmark_ok = all(row.get("benchmark_membership") == expected_benchmark_membership for row in candidate_rows)
    add_gate(
        gate_results,
        "english_learning_core",
        "benchmark_audit",
        "pass" if benchmark_ok else "blocker",
        f"All candidate rows retain {expected_benchmark_membership} benchmark membership metadata."
        if benchmark_ok
        else f"Some candidate rows lack {expected_benchmark_membership} benchmark membership metadata.",
    )

    support_role_ok = (
        contract.get("course", {}).get("target_language_role")
        == "support_language_aids_for_english_learning_not_ordinary_polyglot_deck_contract"
    )
    add_gate(
        gate_results,
        "support_language",
        "support_language_role_check",
        "pass" if support_role_ok else "blocker",
        "Support-language source-package role is scoped as an English-learning aid; US/UK edition workbooks use the final Oxford English-learning sheet contract with no empty support transcription columns."
        if support_role_ok
        else "Support-language role is not explicitly scoped.",
    )

    filled_target_display_keys = set()
    filled_target_example_keys = set()
    if workbook_rows and workbook_headers:
        for row in workbook_rows:
            row_id = normalize_text(row.get("row_id"))
            for code in target_language_codes:
                if normalize_text(row.get(code)):
                    filled_target_display_keys.add((row_id, code))
                if normalize_text(row.get(f"example_{code}")):
                    filled_target_example_keys.add((row_id, code))
    support_batch_problems = []
    support_batch_languages = set()
    candidate_row_ids = {row.get("row_id") for row in candidate_rows}
    expected_row_count = len(candidate_rows)
    for batch in support_translation_batches:
        batch_contract = batch["contract"]
        batch_rows = batch["rows"]
        batch_languages = batch_contract.get("languages") or []
        batch_id = batch_contract.get("batch_id") or batch["path"].stem
        support_batch_languages.update(batch_languages)
        if len(batch_rows) != expected_row_count:
            support_batch_problems.append(
                {"batch": batch_id, "reason": "wrong_row_count", "rows": len(batch_rows), "expected": expected_row_count}
            )
        seen_batch_row_ids = set()
        for row in batch_rows:
            row_id = row.get("row_id")
            if row_id not in candidate_row_ids:
                support_batch_problems.append({"batch": batch_id, "row_id": row_id, "reason": "unknown_row_id"})
                continue
            if row_id in seen_batch_row_ids:
                support_batch_problems.append({"batch": batch_id, "row_id": row_id, "reason": "duplicate_row_id"})
            seen_batch_row_ids.add(row_id)
            if row.get("release_id") != release_id:
                support_batch_problems.append({"batch": batch_id, "row_id": row_id, "reason": "release_id_mismatch"})
            for code in batch_languages:
                if code not in target_language_codes:
                    support_batch_problems.append({"batch": batch_id, "language": code, "reason": "unsupported_language"})
                    continue
                if f"transcription_{code}" in row or f"example_transcription_{code}" in row:
                    support_batch_problems.append(
                        {"batch": batch_id, "row_id": row_id, "language": code, "reason": "target_transcription_forbidden"}
                    )
                if normalize_text(row.get(code)):
                    filled_target_display_keys.add((row_id, code))
                else:
                    support_batch_problems.append(
                        {"batch": batch_id, "row_id": row_id, "language": code, "reason": "missing_display"}
                    )
                if normalize_text(row.get(f"example_{code}")):
                    filled_target_example_keys.add((row_id, code))
                else:
                    support_batch_problems.append(
                        {"batch": batch_id, "row_id": row_id, "language": code, "reason": "missing_example"}
                    )
        missing_batch_rows = candidate_row_ids - seen_batch_row_ids
        if missing_batch_rows:
            support_batch_problems.append(
                {
                    "batch": batch_id,
                    "reason": "missing_rows",
                    "count": len(missing_batch_rows),
                    "sample": sorted(missing_batch_rows)[:8],
                }
            )
    filled_target_cells = len(filled_target_display_keys)
    filled_target_examples = len(filled_target_example_keys)
    expected_target_cells = len(candidate_rows) * len(target_language_codes)
    support_all_filled = (
        filled_target_cells == expected_target_cells
        and filled_target_examples == expected_target_cells
        and expected_target_cells > 0
        and not support_batch_problems
    )
    if support_all_filled:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_meaning_check",
            "pass",
            "Support-language display translations are filled for every configured support language and row.",
            filled_display_cells=filled_target_cells,
            expected_each=expected_target_cells,
            support_batch_languages=sorted(support_batch_languages),
        )
        add_gate(
            gate_results,
            "support_language",
            "support_example_scene_check",
            "pass",
            "Support-language examples are filled for every configured support language and row.",
            filled_example_cells=filled_target_examples,
            expected=expected_target_cells,
            support_batch_languages=sorted(support_batch_languages),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_meaning_check",
            "blocker",
            "Support-language translations are only partially filled or still have batch-shape problems.",
            filled_display_cells=filled_target_cells,
            filled_example_cells=filled_target_examples,
            expected_each=expected_target_cells,
            support_batch_languages=sorted(support_batch_languages),
            batch_problem_sample=support_batch_problems[:10],
        )
        add_gate(
            gate_results,
            "support_language",
            "support_example_scene_check",
            "blocker",
            "Support-language examples are only partially filled, so full scene-preservation checks cannot pass yet.",
            filled_example_cells=filled_target_examples,
            expected=expected_target_cells,
            support_batch_languages=sorted(support_batch_languages),
            batch_problem_sample=support_batch_problems[:10],
        )

    weak_review_artifact = contract.get("latest_weak_language_targeted_review")
    if weak_review_artifact and Path(weak_review_artifact.get("path", "")).exists():
        add_gate(
            gate_results,
            "support_language",
            "weak_language_targeted_review",
            "pass",
            "Weak-language targeted review artifact exists.",
            path=weak_review_artifact.get("path"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "weak_language_targeted_review",
            "blocker",
            "No weak-language targeted review artifact exists yet.",
        )

    sample_review_artifact = contract.get("latest_support_translation_sample_review") or {}
    sample_review_path_value = sample_review_artifact.get("path") or ""
    sample_review_path = Path(sample_review_path_value) if sample_review_path_value else None
    sample_review = (
        read_json(sample_review_path)
        if sample_review_path_value and sample_review_path.exists()
        else {}
    )
    checked_artifacts.append({"gate": "support_translation_sample_review", "path": sample_review_path_value})
    expected_sample_size = 10
    expected_unique_cells = expected_sample_size * len(language_codes)
    sample_review_ok = (
        sample_review_artifact.get("report_id") == "support_translation_sample_review_v1"
        and sample_review_artifact.get("status") == "pass"
        and bool(sample_review_path_value)
        and sample_review_path.exists()
        and sample_review.get("status") == "pass"
        and sample_review.get("sample_size_per_language") == expected_sample_size
        and sample_review.get("languages_reviewed") == len(language_codes)
        and sample_review.get("unique_sample_display_cells_reviewed") == expected_unique_cells
        and sample_review.get("blocker_count") == 0
    )
    if sample_review_ok:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_sample_review",
            "pass",
            "Deterministic 10-word-per-language translation sample review is recorded and has no blockers.",
            path=sample_review_path_value,
            languages_reviewed=sample_review.get("languages_reviewed"),
            sample_size_per_language=sample_review.get("sample_size_per_language"),
            unique_sample_display_cells_reviewed=sample_review.get("unique_sample_display_cells_reviewed"),
            final_workbook_display_cells_checked=sample_review_artifact.get("final_workbook_display_cells_checked"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_sample_review",
            "blocker",
            "Deterministic 10-word-per-language translation sample review is missing, stale or blocked.",
            path=sample_review_path_value,
            report_id=sample_review_artifact.get("report_id"),
            artifact_status=sample_review_artifact.get("status"),
            report_status=sample_review.get("status"),
            languages_reviewed=sample_review.get("languages_reviewed"),
            sample_size_per_language=sample_review.get("sample_size_per_language"),
            unique_sample_display_cells_reviewed=sample_review.get("unique_sample_display_cells_reviewed"),
            expected_unique_sample_display_cells=expected_unique_cells,
            blocker_count=sample_review.get("blocker_count"),
        )

    all_fields_review_artifact = contract.get("latest_all_fields_sample_review") or {}
    all_fields_review_path_value = all_fields_review_artifact.get("path") or ""
    all_fields_review_path = Path(all_fields_review_path_value) if all_fields_review_path_value else None
    all_fields_review = (
        read_json(all_fields_review_path)
        if all_fields_review_path_value and all_fields_review_path.exists()
        else {}
    )
    checked_artifacts.append({"gate": "all_fields_sample_review", "path": all_fields_review_path_value})
    expected_language_row_records = expected_sample_size * len(language_codes)
    all_fields_review_ok = (
        all_fields_review_artifact.get("report_id") == "all_fields_sample_review_v1"
        and all_fields_review_artifact.get("status") == "pass"
        and bool(all_fields_review_path_value)
        and all_fields_review_path.exists()
        and all_fields_review.get("status") == "pass"
        and all_fields_review.get("sample_size_per_language") == expected_sample_size
        and all_fields_review.get("languages_reviewed") == len(language_codes)
        and all_fields_review.get("support_languages_reviewed") == len(target_language_codes)
        and all_fields_review.get("language_row_records_checked") == expected_language_row_records
        and all_fields_review.get("checked_field_count", 0) > expected_language_row_records
        and all_fields_review.get("blocker_count") == 0
    )
    if all_fields_review_ok:
        add_gate(
            gate_results,
            "support_language",
            "all_fields_sample_review",
            "pass",
            "All-fields 10-word-per-language sample review is recorded and has no blockers.",
            path=all_fields_review_path_value,
            languages_reviewed=all_fields_review.get("languages_reviewed"),
            support_languages_reviewed=all_fields_review.get("support_languages_reviewed"),
            sample_size_per_language=all_fields_review.get("sample_size_per_language"),
            language_row_records_checked=all_fields_review.get("language_row_records_checked"),
            checked_field_count=all_fields_review.get("checked_field_count"),
            warning_signal_count=all_fields_review.get("warning_signal_count"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "all_fields_sample_review",
            "blocker",
            "All-fields 10-word-per-language sample review is missing, stale or blocked.",
            path=all_fields_review_path_value,
            report_id=all_fields_review_artifact.get("report_id"),
            artifact_status=all_fields_review_artifact.get("status"),
            report_status=all_fields_review.get("status"),
            languages_reviewed=all_fields_review.get("languages_reviewed"),
            expected_languages_reviewed=len(language_codes),
            support_languages_reviewed=all_fields_review.get("support_languages_reviewed"),
            expected_support_languages_reviewed=len(target_language_codes),
            sample_size_per_language=all_fields_review.get("sample_size_per_language"),
            expected_sample_size_per_language=expected_sample_size,
            language_row_records_checked=all_fields_review.get("language_row_records_checked"),
            expected_language_row_records=expected_language_row_records,
            checked_field_count=all_fields_review.get("checked_field_count"),
            blocker_count=all_fields_review.get("blocker_count"),
        )

    source_backed_audit_artifact = contract.get("latest_support_translation_source_backed_audit") or {}
    source_backed_audit_path_value = source_backed_audit_artifact.get("path") or ""
    source_backed_audit_path = Path(source_backed_audit_path_value)
    source_backed_audit = (
        read_json(source_backed_audit_path)
        if source_backed_audit_path_value and source_backed_audit_path.exists()
        else {}
    )
    checked_artifacts.append(
        {"gate": "support_translation_source_backed_audit", "path": str(source_backed_audit_path)}
    )
    expected_source_backed_cells = len(candidate_rows) * len(target_language_codes)
    source_backed_audit_ok = (
        source_backed_audit_artifact.get("report_id") == "support_translation_source_backed_audit_v1"
        and source_backed_audit_artifact.get("status") == "pass"
        and bool(source_backed_audit_path_value)
        and source_backed_audit_path.exists()
        and source_backed_audit.get("status") == "pass"
        and source_backed_audit.get("display_cells_checked") == expected_source_backed_cells
        and source_backed_audit.get("support_languages_checked") == len(target_language_codes)
        and source_backed_audit.get("blocker_count") == 0
        and source_backed_audit.get("source_candidate_rows", 0) > 0
        and bool(source_backed_audit.get("source_hash"))
        and source_backed_audit_artifact.get("source_hash") == source_backed_audit.get("source_hash")
    )
    if source_backed_audit_ok:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_source_backed_audit",
            "pass",
            "Full support-language display translation source-backed audit is recorded and has no blockers.",
            path=str(source_backed_audit_path),
            display_cells_checked=source_backed_audit.get("display_cells_checked"),
            support_languages_checked=source_backed_audit.get("support_languages_checked"),
            source_candidate_rows=source_backed_audit.get("source_candidate_rows"),
            source_exact_or_normalized_match_rows=source_backed_audit.get("source_exact_or_normalized_match_rows"),
            source_partial_only_rows=source_backed_audit.get("source_partial_only_rows"),
            no_source_candidate_rows=source_backed_audit.get("no_source_candidate_rows"),
            warning_count=source_backed_audit.get("warning_count"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_translation_source_backed_audit",
            "blocker",
            "Full support-language display translation source-backed audit is missing, stale, incomplete or blocked.",
            path=str(source_backed_audit_path),
            report_id=source_backed_audit_artifact.get("report_id"),
            artifact_status=source_backed_audit_artifact.get("status"),
            report_status=source_backed_audit.get("status"),
            display_cells_checked=source_backed_audit.get("display_cells_checked"),
            expected_display_cells=expected_source_backed_cells,
            support_languages_checked=source_backed_audit.get("support_languages_checked"),
            expected_support_languages=len(target_language_codes),
            source_candidate_rows=source_backed_audit.get("source_candidate_rows"),
            blocker_count=source_backed_audit.get("blocker_count"),
            source_hash_present=bool(source_backed_audit.get("source_hash")),
            source_hash_matches=source_backed_audit_artifact.get("source_hash") == source_backed_audit.get("source_hash"),
        )

    support_example_audit_artifact = contract.get("latest_support_example_quality_audit") or {}
    support_example_audit_path_value = support_example_audit_artifact.get("path") or ""
    support_example_audit_path = Path(support_example_audit_path_value)
    support_example_audit = (
        read_json(support_example_audit_path)
        if support_example_audit_path_value and support_example_audit_path.exists()
        else {}
    )
    checked_artifacts.append({"gate": "support_example_quality_audit", "path": str(support_example_audit_path)})
    support_example_audit_ok = (
        support_example_audit_artifact.get("report_id") == "support_example_quality_audit_v1"
        and support_example_audit_artifact.get("status") == "pass"
        and bool(support_example_audit_path_value)
        and support_example_audit_path.exists()
        and support_example_audit.get("status") == "pass"
        and support_example_audit.get("example_cells_checked") == expected_target_cells
        and support_example_audit.get("support_languages_checked") == len(target_language_codes)
        and support_example_audit.get("blocker_count") == 0
        and bool(support_example_audit.get("source_hash"))
        and support_example_audit_artifact.get("source_hash") == support_example_audit.get("source_hash")
    )
    if support_example_audit_ok:
        add_gate(
            gate_results,
            "support_language",
            "support_example_quality_audit",
            "pass",
            "Full support-language example quality audit is recorded and has no blockers.",
            path=str(support_example_audit_path),
            example_cells_checked=support_example_audit.get("example_cells_checked"),
            support_languages_checked=support_example_audit.get("support_languages_checked"),
            warning_count=support_example_audit.get("warning_count"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_example_quality_audit",
            "blocker",
            "Full support-language example quality audit is missing, stale, incomplete or blocked.",
            path=str(support_example_audit_path),
            report_id=support_example_audit_artifact.get("report_id"),
            artifact_status=support_example_audit_artifact.get("status"),
            report_status=support_example_audit.get("status"),
            example_cells_checked=support_example_audit.get("example_cells_checked"),
            expected_example_cells=expected_target_cells,
            support_languages_checked=support_example_audit.get("support_languages_checked"),
            expected_support_languages=len(target_language_codes),
            blocker_count=support_example_audit.get("blocker_count"),
            source_hash_present=bool(support_example_audit.get("source_hash")),
            source_hash_matches=support_example_audit_artifact.get("source_hash") == support_example_audit.get("source_hash"),
        )

    article_repair = contract.get("latest_support_article_display_repair") or {}
    changed_batches = article_repair.get("changed_batches") or []
    checked_batches = article_repair.get("checked_batches") or []
    changed_batch_ids = {batch.get("batch_id") for batch in changed_batches if batch.get("batch_id")}
    checked_batch_ids = {batch.get("batch_id") for batch in checked_batches if batch.get("batch_id")}
    latest_batch_ids = {
        batch["contract"].get("batch_id")
        for batch in support_translation_batches
        if batch["contract"].get("batch_id")
    }
    article_batch_paths = [Path(batch.get("path")) for batch in [*changed_batches, *checked_batches] if batch.get("path")]
    missing_article_batch_paths = [str(path) for path in article_batch_paths if not path.exists()]
    expected_article_languages = {"ES", "FR", "DE", "IT", "PT", "NL", "SV", "NO", "DA", "RO", "ES-419", "PT-BR"}
    article_languages = set(article_repair.get("target_languages") or [])
    zero_change_check_ok = (
        article_repair.get("check_status") == "pass"
        and article_repair.get("display_cells_changed", 0) == 0
        and checked_batch_ids
        and checked_batch_ids.issubset(latest_batch_ids)
    )
    changed_repair_ok = (
        article_repair.get("display_cells_changed", 0) > 0
        and changed_batch_ids
        and changed_batch_ids.issubset(latest_batch_ids)
    )
    article_repair_ok = (
        article_repair.get("repair_id") == "support_article_display_repair_v1"
        and article_languages == expected_article_languages
        and (changed_repair_ok or zero_change_check_ok)
        and not missing_article_batch_paths
    )
    checked_artifacts.extend(
        {"gate": "support_article_display_repair_check", "path": str(path)} for path in article_batch_paths
    )
    if article_repair_ok:
        add_gate(
            gate_results,
            "support_language",
            "support_article_display_repair_check",
            "pass",
            "Support-language article display repair/check is recorded and all article-language batches are contract-backed.",
            repair_id=article_repair.get("repair_id"),
            display_cells_changed=article_repair.get("display_cells_changed"),
            target_languages=sorted(article_languages),
            changed_batch_ids=sorted(changed_batch_ids),
            checked_batch_ids=sorted(checked_batch_ids),
            check_status=article_repair.get("check_status"),
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_article_display_repair_check",
            "blocker",
            "Support-language article display repair is missing, stale or not wired into the current support batches.",
            repair_id=article_repair.get("repair_id"),
            display_cells_changed=article_repair.get("display_cells_changed"),
            target_languages=sorted(article_languages),
            changed_batch_ids=sorted(changed_batch_ids),
            checked_batch_ids=sorted(checked_batch_ids),
            check_status=article_repair.get("check_status"),
            latest_batch_ids=sorted(latest_batch_ids),
            missing_changed_batch_paths=missing_article_batch_paths,
        )

    workbook_required_fixed = [
        "release_id",
        "course_id",
        "row_id",
        "core_item_id",
        "meaning_id",
        "source_language",
        "source_variant",
        "source_headword",
        "part_of_speech",
        "sense_no",
        "core_band",
        "level_min",
        "level_max",
        "oxford_order",
        "benchmark_membership",
        "source_status",
        "meaning_note",
        "semantic_scene",
    ]
    expected_headers = workbook_required_fixed + language_codes + [f"example_{code}" for code in language_codes] + [
        "qa_status",
        "qa_notes",
    ]
    service_sheets_expected = {"_README", "_source_contract", "_languages"}
    service_sheets_future = {"_benchmark_audit", "_source_assisted_qa", "_source_decisions", "_release_metadata"}
    edition_details = []
    edition_workbook_ok = True
    edition_export = contract.get("latest_edition_exports") or {}
    if edition_export:
        expected_edition_service_sheets = {"Course Metadata", "Deck Metadata", "Card Metadata", "_README", "_qa_status", "_languages"}
        for edition in edition_export.get("editions", []):
            primary_language_code = normalize_text(edition.get("primary_language_code"))
            expected_edition_language_codes = main_sheet_language_codes_for_primary(primary_language_code, language_codes)
            expected_pronunciation_headers = pronunciation_headers_for_primary(primary_language_code)
            expected_edition_headers = (
                expected_edition_language_codes
                + [f"{code} example" for code in expected_edition_language_codes]
                + expected_pronunciation_headers
            )
            opposite_code = opposite_english_code(primary_language_code)
            path = Path(edition.get("path") or "")
            checked_artifacts.append({"gate": "workbook_contract_check", "path": str(path)})
            detail = {
                "edition_id": edition.get("edition_id"),
                "path": str(path),
                "expected_rows": len(candidate_rows),
                "expected_columns": len(expected_edition_headers),
            }
            if not path.exists():
                detail["exists"] = False
                detail["status"] = "missing"
                edition_workbook_ok = False
                edition_details.append(detail)
                continue
            edition_workbook = load_workbook(path, read_only=True, data_only=True)
            main_sheet = edition_workbook[edition_workbook.sheetnames[0]]
            header_rows = list(main_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [normalize_text(cell) for cell in header_rows[0]] if header_rows else []
            header_index = {header: index for index, header in enumerate(headers)}
            support_transcription_headers_present = [
                header
                for header in headers
                if header.endswith(" transcription") and header not in set(OXFORD_EDITION_PRONUNCIATION_HEADERS)
            ]
            opposite_english_headers_present = [
                header
                for header in [
                    opposite_code,
                    f"{opposite_code} example",
                    f"{opposite_code} transcription",
                    f"{opposite_code} example transcription",
                ]
                if header in header_index
            ]
            pronunciation_blockers = []
            if not support_transcription_headers_present and all(
                header in header_index for header in expected_pronunciation_headers
            ):
                for row_number, main_row in enumerate(
                    main_sheet.iter_rows(min_row=2, max_row=main_sheet.max_row, values_only=True), start=2
                ):
                    row_values = {
                        header: normalize_text(main_row[index] if index < len(main_row) else "")
                        for header, index in header_index.items()
                    }
                    for header in expected_pronunciation_headers:
                        if not valid_ipa_cell(row_values.get(header)):
                            pronunciation_blockers.append({"row": row_number, "column": header, "reason": "invalid_ipa"})
                    if len(pronunciation_blockers) >= 10:
                        break
            else:
                missing_pronunciation_headers = [
                    header for header in expected_pronunciation_headers if header not in header_index
                ]
                pronunciation_blockers.extend(
                    [
                        {"reason": "missing_pronunciation_headers", "headers": missing_pronunciation_headers},
                        {"reason": "support_transcription_headers_present", "headers": support_transcription_headers_present[:10]},
                    ]
                )
            pronunciation_block_ok = not pronunciation_blockers
            deck_metadata = {}
            if "Deck Metadata" in edition_workbook.sheetnames:
                for row in edition_workbook["Deck Metadata"].iter_rows(min_row=2, max_col=2, values_only=True):
                    if row and normalize_text(row[0]):
                        deck_metadata[normalize_text(row[0])] = normalize_text(row[1])
            course_metadata_ok = False
            course_metadata_detail = {}
            if "Course Metadata" in edition_workbook.sheetnames:
                course_rows = list(edition_workbook["Course Metadata"].iter_rows(values_only=True))
                course_header = [normalize_text(cell) for cell in course_rows[0][1:]] if course_rows else []
                course_map = {
                    normalize_text(row[0]): [normalize_text(cell) for cell in row[1:]]
                    for row in course_rows[1:]
                    if row and normalize_text(row[0])
                }
                titles = course_map.get("Title", [])
                descriptions = course_map.get("Description", [])
                modules = course_map.get("Module", [])
                categories = course_map.get("Category", [])
                course_metadata_ok = (
                    course_header == expected_edition_language_codes
                    and len(titles) == len(expected_edition_language_codes)
                    and len(descriptions) == len(expected_edition_language_codes)
                    and len(modules) == len(expected_edition_language_codes)
                    and len(categories) == len(expected_edition_language_codes)
                    and all(titles)
                    and all(descriptions)
                    and all(modules)
                    and all(categories)
                    and all(release_course_level in title and release_part_number in title for title in titles)
                    and all(
                        release_wordlist_token in description and release_part_number in description
                        for description in descriptions
                    )
                    and len(set(titles)) >= 20
                    and len(set(descriptions)) >= 20
                )
                course_metadata_detail = {
                    "header_ok": course_header == expected_edition_language_codes,
                    "title_sample": titles[:6],
                    "description_sample": descriptions[:6],
                    "unique_titles": len(set(titles)),
                    "unique_descriptions": len(set(descriptions)),
                }
            card_metadata_ok = False
            card_metadata_detail = {}
            if "Card Metadata" in edition_workbook.sheetnames:
                card_rows = list(edition_workbook["Card Metadata"].iter_rows(values_only=True))
                card_headers = [normalize_text(cell) for cell in card_rows[0]] if card_rows else []
                card_header_index = {header: index for index, header in enumerate(card_headers)}
                missing_card_headers = sorted(REQUIRED_CARD_METADATA_HEADERS - set(card_headers))
                card_blockers = []
                if not missing_card_headers:
                    for card_row in card_rows[1:]:
                        row_values = {
                            header: normalize_text(card_row[index] if index < len(card_row) else "")
                            for header, index in card_header_index.items()
                        }
                        for required_value in [
                            "main_sheet_row",
                            "display_order",
                            "card_key",
                            "set_id",
                            "meaning_id",
                            "context_example_key",
                            "context_example_quality_status",
                            "meaning_quality_status",
                            "membership_quality_status",
                        ]:
                            if not row_values.get(required_value):
                                card_blockers.append(
                                    {
                                        "row": row_values.get("main_sheet_row"),
                                        "reason": f"missing_{required_value}",
                                    }
                                )
                        if len(card_blockers) >= 10:
                            break
                card_metadata_ok = not missing_card_headers and not card_blockers and len(card_rows) - 1 == len(candidate_rows)
                card_metadata_detail = {
                    "columns": len(card_headers),
                    "missing_headers": missing_card_headers,
                    "blockers": card_blockers[:10],
                }
            missing_service_sheets = sorted(expected_edition_service_sheets - set(edition_workbook.sheetnames))
            metadata_ok = (
                deck_metadata.get("set_id") == normalize_text(edition.get("deck_id"))
                and deck_metadata.get("source_variant") == normalize_text(edition.get("source_variant"))
                and deck_metadata.get("primary_language_code") == normalize_text(edition.get("primary_language_code"))
            )
            detail.update(
                {
                    "exists": True,
                    "main_sheet": main_sheet.title,
                    "rows": main_sheet.max_row - 1,
                    "columns": main_sheet.max_column,
                    "sheets": edition_workbook.sheetnames,
                    "missing_service_sheets": missing_service_sheets,
                    "main_column_blocks_ok": headers == expected_edition_headers
                    and not support_transcription_headers_present
                    and not opposite_english_headers_present
                    and pronunciation_block_ok,
                    "primary_language_code": primary_language_code,
                    "main_sheet_language_count": len(expected_edition_language_codes),
                    "opposite_english_headers_present": opposite_english_headers_present,
                    "support_transcription_headers_present": support_transcription_headers_present,
                    "pronunciation_block_ok": pronunciation_block_ok,
                    "pronunciation_blockers": pronunciation_blockers[:10],
                    "deck_metadata_ok": metadata_ok,
                    "course_metadata_ok": course_metadata_ok,
                    "course_metadata": course_metadata_detail,
                    "card_metadata_ok": card_metadata_ok,
                    "card_metadata": card_metadata_detail,
                    "deck_metadata": {
                        "set_id": deck_metadata.get("set_id"),
                        "source_variant": deck_metadata.get("source_variant"),
                        "primary_language_code": deck_metadata.get("primary_language_code"),
                    },
                }
            )
            detail["status"] = "pass" if (
                detail["rows"] == len(candidate_rows)
                and detail["columns"] == len(expected_edition_headers)
                and detail["main_column_blocks_ok"]
                and not missing_service_sheets
                and metadata_ok
                and course_metadata_ok
                and card_metadata_ok
            ) else "blocker"
            if detail["status"] != "pass":
                edition_workbook_ok = False
            edition_details.append(detail)
    if workbook:
        workbook_shape_ok = (
            workbook_headers == expected_headers
            and len(workbook_rows) == len(candidate_rows)
            and service_sheets_expected <= set(workbook.sheetnames)
        )
        status = "pass" if workbook_shape_ok and edition_workbook_ok else "blocker"
        add_gate(
            gate_results,
            "workbook_and_delivery",
            "workbook_contract_check",
            status,
            "Workbook source-draft shape and local US/UK ordinary-format edition workbook shapes match current Oxford contracts."
            if status == "pass"
            else "Source-draft or US/UK edition workbook shape does not match current Oxford contracts.",
            sheet=workbook_sheet,
            rows=len(workbook_rows),
            columns=len(workbook_headers),
            missing_service_sheets=sorted(service_sheets_expected - set(workbook.sheetnames)),
            future_service_sheets_missing=sorted(service_sheets_future - set(workbook.sheetnames)),
            edition_workbooks=edition_details,
        )

    add_gate(
        gate_results,
        "workbook_and_delivery",
        "release_specific_audit_script_check",
        "pass",
        "Oxford/English-learning executable gate checker exists and produced this report.",
        script="scripts/oxford/check-oxford-english-learning-gates.py",
        script_version=SCRIPT_VERSION,
    )

    semantic_smell_artifact = contract.get("latest_semantic_smell_audit") or {}
    semantic_smell_required = "oxford_5000" in release_id or bool(semantic_smell_artifact)
    semantic_smell_path_value = semantic_smell_artifact.get("json_path") or semantic_smell_artifact.get("path") or ""
    semantic_smell_path = Path(semantic_smell_path_value) if semantic_smell_path_value else None
    semantic_smell_report = (
        read_json(semantic_smell_path)
        if semantic_smell_path_value and semantic_smell_path.exists()
        else {}
    )
    if semantic_smell_required:
        checked_artifacts.append({"gate": "english_example_semantic_smell_check", "path": semantic_smell_path_value})
        semantic_smell_findings = [
            finding
            for finding in semantic_smell_report.get("findings", [])
            if finding.get("release_id") == release_id
        ]
        semantic_smell_blockers = [
            finding
            for finding in semantic_smell_findings
            if finding.get("severity") == "blocker"
        ]
        semantic_smell_warnings = [
            finding
            for finding in semantic_smell_findings
            if finding.get("severity") == "warning"
        ]
        semantic_smell_ok = (
            bool(semantic_smell_path_value)
            and semantic_smell_path.exists()
            and semantic_smell_report.get("status") in {"pass", "ok", "passed"}
            and not semantic_smell_blockers
        )
        add_gate(
            gate_results,
            "english_learning_core",
            "english_example_semantic_smell_check",
            "pass" if semantic_smell_ok else "blocker",
            "English base examples passed the semantic-smell audit for this release."
            if semantic_smell_ok
            else "English base examples have semantic-smell blockers or the required audit is missing; repair the English example map before trusting support-language examples or final delivery.",
            path=semantic_smell_path_value,
            artifact_status=semantic_smell_artifact.get("status"),
            report_status=semantic_smell_report.get("status"),
            release_findings=len(semantic_smell_findings),
            release_blockers=len(semantic_smell_blockers),
            release_warnings=len(semantic_smell_warnings),
            sample=row_sample(semantic_smell_blockers, field="source_headword", limit=8),
        )

    edition_exports = contract.get("latest_edition_exports", {})
    delivery_manifests = edition_exports.get("delivery_manifests", [])
    readback_verified = bool(delivery_manifests) and all(
        item.get("google_sheet_readback_status") == "verified" for item in delivery_manifests
    )
    approval_ready = (
        contract.get("approved_for_generation") is True
        and contract.get("final_delivery_decision", {}).get("status") == "approved"
        and edition_exports.get("final_delivery_ready") is True
        and edition_exports.get("google_sheet_created") is True
        and readback_verified
    )
    add_gate(
        gate_results,
        "workbook_and_delivery",
        "final_delivery_approval_check",
        "pass" if approval_ready else "blocker",
        "Final delivery approval flags and Google Sheet readback are enabled."
        if approval_ready
        else "Final delivery approval/readback is incomplete; no ordinary deck-table Postgres import or Google Sheet delivery may be treated as current final.",
        approved_for_generation=contract.get("approved_for_generation"),
        final_delivery_decision=contract.get("final_delivery_decision", {}).get("status"),
        final_delivery_ready=edition_exports.get("final_delivery_ready"),
        google_sheet_created=edition_exports.get("google_sheet_created"),
        readback_verified=readback_verified,
    )

    post_final_artifact = contract.get("latest_post_final_language_spot_check") or {}
    post_final_path_value = post_final_artifact.get("path") or ""
    post_final_path = Path(post_final_path_value) if post_final_path_value else None
    post_final_report = (
        read_json(post_final_path)
        if post_final_path_value and post_final_path.exists()
        else {}
    )
    checked_artifacts.append({"gate": "post_final_language_spot_check", "path": post_final_path_value})
    post_final_ok = (
        post_final_artifact.get("report_id") == POST_FINAL_SPOT_CHECK_REPORT_ID
        and post_final_artifact.get("status") == "pass"
        and bool(post_final_path_value)
        and post_final_path.exists()
        and post_final_report.get("release_id") == release_id
        and post_final_report.get("status") == "pass"
        and post_final_report.get("sample_size_per_language", 0) >= POST_FINAL_SPOT_CHECK_MIN_SAMPLE_SIZE
        and post_final_report.get("languages_reviewed") == len(language_codes)
        and post_final_report.get("support_languages_reviewed") == len(target_language_codes)
        and post_final_report.get("blocker_count") == 0
        and post_final_report.get("final_workbook_check", {}).get("display_cells_checked", 0) > 0
        and post_final_report.get("final_workbook_check", {}).get("example_cells_checked", 0) > 0
        and post_final_report.get("transcription_rows_checked", 0)
        >= POST_FINAL_SPOT_CHECK_MIN_SAMPLE_SIZE * 2
    )
    if approval_ready:
        add_gate(
            gate_results,
            "workbook_and_delivery",
            "post_final_language_spot_check",
            "pass" if post_final_ok else "blocker",
            "Post-final deterministic spot check covers all 54 language variants and final US/UK workbook cells with no blockers."
            if post_final_ok
            else "Post-final deterministic spot check is missing, stale, incomplete or blocked after final Google Sheet readback.",
            path=post_final_path_value,
            report_id=post_final_artifact.get("report_id"),
            artifact_status=post_final_artifact.get("status"),
            report_status=post_final_report.get("status"),
            sample_size_per_language=post_final_report.get("sample_size_per_language"),
            minimum_sample_size_per_language=POST_FINAL_SPOT_CHECK_MIN_SAMPLE_SIZE,
            languages_reviewed=post_final_report.get("languages_reviewed"),
            expected_languages_reviewed=len(language_codes),
            support_languages_reviewed=post_final_report.get("support_languages_reviewed"),
            expected_support_languages=len(target_language_codes),
            blocker_count=post_final_report.get("blocker_count"),
        )
    else:
        add_gate(
            gate_results,
            "workbook_and_delivery",
            "post_final_language_spot_check",
            "warning",
            "Post-final deterministic spot check is required after final US/UK Google Sheet readback and is not evaluated as a release blocker before that stage.",
            path=post_final_path_value,
            final_delivery_ready=edition_exports.get("final_delivery_ready"),
            google_sheet_created=edition_exports.get("google_sheet_created"),
            readback_verified=readback_verified,
        )

    advisory_counts = summarize_advisory_warning_codes(post_final_report) if post_final_ok else {}
    total_advisory_codes = sum(advisory_counts.values())
    weak_language_codes = sum(
        count for code, count in advisory_counts.items() if code.startswith("source_audit:weak_language_")
    ) + advisory_counts.get("example_quality:weak_language_example_review_signal", 0)
    budget_breaches = {}
    if total_advisory_codes > POST_FINAL_WARNING_BUDGET["max_total_warning_codes"]:
        budget_breaches["total_warning_codes"] = {
            "actual": total_advisory_codes,
            "max": POST_FINAL_WARNING_BUDGET["max_total_warning_codes"],
        }
    if advisory_counts.get("source_audit:no_source_candidate", 0) > POST_FINAL_WARNING_BUDGET["max_source_no_candidate_codes"]:
        budget_breaches["source_audit:no_source_candidate"] = {
            "actual": advisory_counts.get("source_audit:no_source_candidate", 0),
            "max": POST_FINAL_WARNING_BUDGET["max_source_no_candidate_codes"],
        }
    if weak_language_codes > POST_FINAL_WARNING_BUDGET["max_weak_language_codes"]:
        budget_breaches["weak_language_warning_codes"] = {
            "actual": weak_language_codes,
            "max": POST_FINAL_WARNING_BUDGET["max_weak_language_codes"],
        }
    if (
        advisory_counts.get("example_quality:target_example_lexical_anchor", 0)
        > POST_FINAL_WARNING_BUDGET["max_target_example_lexical_anchor_codes"]
    ):
        budget_breaches["example_quality:target_example_lexical_anchor"] = {
            "actual": advisory_counts.get("example_quality:target_example_lexical_anchor", 0),
            "max": POST_FINAL_WARNING_BUDGET["max_target_example_lexical_anchor_codes"],
        }
    if (
        advisory_counts.get("source_audit:exact_english_surface_needs_loan_review", 0)
        > POST_FINAL_WARNING_BUDGET["max_exact_english_surface_loan_review_codes"]
    ):
        budget_breaches["source_audit:exact_english_surface_needs_loan_review"] = {
            "actual": advisory_counts.get("source_audit:exact_english_surface_needs_loan_review", 0),
            "max": POST_FINAL_WARNING_BUDGET["max_exact_english_surface_loan_review_codes"],
        }
    if post_final_ok:
        top_advisory_codes = sorted(advisory_counts.items(), key=lambda item: item[1], reverse=True)[:10]
        add_gate(
            gate_results,
            "support_language",
            "support_language_warning_budget",
            "warning" if budget_breaches else "pass",
            "Post-final advisory warning budget is exceeded; release remains non-blocked, but high-risk support-language signals require review planning."
            if budget_breaches
            else "Post-final advisory warning budget is within configured limits.",
            total_advisory_warning_codes=total_advisory_codes,
            warning_budget=POST_FINAL_WARNING_BUDGET,
            budget_breaches=budget_breaches,
            top_advisory_codes=top_advisory_codes,
            does_not_replace="native-speaker certification or full row-by-row linguistic audit",
        )
    elif approval_ready:
        add_gate(
            gate_results,
            "support_language",
            "support_language_warning_budget",
            "blocker",
            "Post-final advisory warning budget cannot be evaluated because the required post-final spot-check report is missing or blocked.",
            path=post_final_path_value,
        )
    else:
        add_gate(
            gate_results,
            "support_language",
            "support_language_warning_budget",
            "warning",
            "Post-final advisory warning budget is evaluated after final US/UK Google Sheet readback.",
            path=post_final_path_value,
        )

    blockers = [item for item in gate_results if item["status"] == "blocker"]
    warnings = [item for item in gate_results if item["status"] == "warning"]
    passes = [item for item in gate_results if item["status"] == "pass"]
    report = {
        "release_id": release_id,
        "report_id": REPORT_ID,
        "script_version": SCRIPT_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": "passed" if not blockers else "blocked",
        "summary": {
            "normalized_rows": len(normalized_rows),
            "selected_rows": len(candidate_rows),
            "workbook_rows": len(workbook_rows),
            "language_columns": len(language_codes),
            "target_support_languages": len(target_language_codes),
            "filled_support_display_cells": filled_target_cells,
            "filled_support_example_cells": filled_target_examples,
            "passed_gates": len(passes),
            "warning_gates": len(warnings),
            "blocked_gates": len(blockers),
        },
        "gate_scope": contract.get("qa_gate_scope"),
        "checked_artifacts": checked_artifacts,
        "gate_results": gate_results,
        "blockers": blockers,
        "warnings": warnings,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    report_json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        "# Oxford English-Learning Gates v1",
        "",
        f"Release: `{release_id}`",
        f"Status: `{report['status']}`",
        "",
        "This executable gate is release-specific to Oxford / English-learning courses. It does not change ordinary LunaCards polyglot deck gates.",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
        f"| Normalized source rows | {report['summary']['normalized_rows']} |",
        f"| Selected rows | {report['summary']['selected_rows']} |",
        f"| Workbook rows | {report['summary']['workbook_rows']} |",
        f"| Language columns | {report['summary']['language_columns']} |",
        f"| Support-language display cells | {filled_target_cells}/{expected_target_cells} |",
        f"| Support-language example cells | {filled_target_examples}/{expected_target_cells} |",
        f"| Passed gates | {len(passes)} |",
        f"| Warning gates | {len(warnings)} |",
        f"| Blocked gates | {len(blockers)} |",
        "",
        "## Blockers",
        "",
    ]
    lines.extend(
        [f"- `{item['gate']}`: {item['message']}" for item in blockers]
        if blockers
        else ["- None."]
    )
    lines.extend(["", "## Warnings", ""])
    lines.extend(
        [f"- `{item['gate']}`: {item['message']}" for item in warnings]
        if warnings
        else ["- None."]
    )
    lines.extend(["", "## Gate Results", ""])
    for item in gate_results:
        lines.append(f"- `{item['status']}` `{item['layer']}.{item['gate']}`: {item['message']}")
    lines.extend(
        [
            "",
            "## Report Files",
            "",
            f"- JSON: `{rel(report_json_path)}`",
            f"- Markdown: `{rel(report_md_path)}`",
            "",
        ]
    )
    report_md_path.write_text("\n".join(lines), encoding="utf-8")

    if blockers:
        print(
            f"Oxford English-learning gates blocked: blockers={len(blockers)} warnings={len(warnings)} report={rel(report_md_path)}",
            file=sys.stderr,
        )
        return 1

    print(
        f"Oxford English-learning gates passed: release={release_id} warnings={len(warnings)} report={rel(report_md_path)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
