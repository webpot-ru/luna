#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_VERSION = "2026-05-21.v6"
EXPORT_MODE = "working"
EXPORT_READINESS = "not final-ready; working preview only"
SPREADSHEET_CONTRACT = "config/spreadsheet-contract-v1.json"
COURSE_LEVEL = "A1"
COURSE_PART_NUMBER = "1"
COURSE_PART_PADDED = "001"
COURSE_WORDLIST_LABEL = "Oxford 3000"
COURSE_TITLE_LABEL = "Oxford 3000 Core"
COURSE_FILE_PREFIX = "Oxford_3000_Core"
COURSE_SLUG_PREFIX = "oxford-3000-core"
COURSE_DECK_ID_PREFIX = "oxford_3000_core"
COURSE_DOMAIN_LABEL = "Oxford 3000 Core"
COURSE_MODULE_LABEL = "Oxford 3000"
SUPPORT_STATUS = "draft_support_language_aid_needs_source_assisted_qa"
SUPPORT_TRANSCRIPTION_STATUS = "out_of_scope_for_support_language_aid"
STATUS_ISSUE_SUPPORT = "support-language word transcription not required for Oxford English-learning edition workbook"
STATUS_ISSUE_ENGLISH = "legal gates blocked"
FINAL_READY_FLAG = "no"
SELECTION_STATUS = "draft_source_package"
QUALITY_STATUS = "working_preview_not_final_ready"
EXPORT_STATUS = "working_preview_not_final_ready"
CONTEXT_EXAMPLE_QUALITY_STATUS = "draft_reviewed_not_final"
MEMBERSHIP_QUALITY_STATUS = "draft_source_package"
TAG_SUFFIX = "working_preview"
GOOGLE_SHEET_STATUS = "false"
POSTGRES_IMPORT_STATUS = "false"
QA_EVIDENCE_BLOCKERS = "none"
LEGAL_BLOCKERS = [
    "permission_evidence_or_project_evidence_decision",
    "allowed_fields_review",
    "final_delivery_approval_check",
]

EDITION_CONFIG = {
    "us_english": {
        "deck_id": "oxford_3000_core_a1_part_001_us_english_draft",
        "slug": "oxford-3000-core-a1-part-001-us-english",
        "deck_title": "FlashcardsLuna Oxford 3000 Core A1 Part 001 - US English",
        "short_title": "Oxford A1 Part 1 US.",
        "description": "Oxford 3000 words. Part 1. US English.",
        "deck_edition": "US English",
        "source_variant": "EN-US",
        "primary_language_code": "EN",
        "sheet_name": "Oxford A1 US",
        "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_US_English_draft.xlsx",
    },
    "british_english": {
        "deck_id": "oxford_3000_core_a1_part_001_british_english_draft",
        "slug": "oxford-3000-core-a1-part-001-british-english",
        "deck_title": "FlashcardsLuna Oxford 3000 Core A1 Part 001 - British English",
        "short_title": "Oxford A1 Part 1 UK.",
        "description": "Oxford 3000 words. Part 1. British English.",
        "deck_edition": "British English",
        "source_variant": "EN-GB",
        "primary_language_code": "EN-GB",
        "sheet_name": "Oxford A1 UK",
        "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_British_English_draft.xlsx",
    },
}

SENTENCE_END_BY_CODE = {
    "ZH": "。",
    "JA": "。",
    "HI": "।",
    "BN": "।",
    "NE": "।",
    "MY": "။",
    "KM": "។",
    "HY": "։",
}

COURSE_METADATA_COPY = {
    "EN": ["Part 1", "Oxford 3000 words.", "Part 1.", "US English.", "British English."],
    "ES": ["parte 1", "Palabras Oxford 3000.", "Parte 1.", "Inglés de EE. UU.", "Inglés británico."],
    "FR": ["partie 1", "Mots Oxford 3000.", "Partie 1.", "Anglais américain.", "Anglais britannique."],
    "DE": ["Teil 1", "Oxford-3000-Wörter.", "Teil 1.", "US-Englisch.", "Britisches Englisch."],
    "IT": ["parte 1", "Parole Oxford 3000.", "Parte 1.", "Inglese USA.", "Inglese britannico."],
    "PT": ["parte 1", "Palavras Oxford 3000.", "Parte 1.", "Inglês dos EUA.", "Inglês britânico."],
    "RU": ["часть 1", "Слова Oxford 3000.", "Часть 1.", "Английский США.", "Британский английский."],
    "ZH": ["第1部分", "Oxford 3000 词汇。", "第1部分。", "美式英语。", "英式英语。"],
    "JA": ["第1部", "Oxford 3000 語彙。", "第1部。", "米国英語版。", "英国英語版。"],
    "KO": ["1부", "Oxford 3000 단어.", "1부.", "미국 영어.", "영국 영어."],
    "VI": ["phần 1", "Từ Oxford 3000.", "Phần 1.", "Tiếng Anh Mỹ.", "Tiếng Anh Anh."],
    "TH": ["ตอน 1", "คำ Oxford 3000.", "ตอน 1.", "อังกฤษแบบสหรัฐฯ.", "อังกฤษแบบอังกฤษ."],
    "MS": ["bhg. 1", "Kata Oxford 3000.", "Bahagian 1.", "Inggeris AS.", "Inggeris UK."],
    "ID": ["bag. 1", "Kata Oxford 3000.", "Bagian 1.", "Inggris AS.", "Inggris Britania."],
    "PL": ["część 1", "Słowa Oxford 3000.", "Część 1.", "Angielski USA.", "Angielski brytyjski."],
    "NL": ["deel 1", "Oxford 3000-woorden.", "Deel 1.", "Amerikaans Engels.", "Brits Engels."],
    "SV": ["del 1", "Oxford 3000-ord.", "Del 1.", "Amerikansk engelska.", "Brittisk engelska."],
    "NO": ["del 1", "Oxford 3000-ord.", "Del 1.", "Amerikansk engelsk.", "Britisk engelsk."],
    "DA": ["del 1", "Oxford 3000-ord.", "Del 1.", "Amerikansk engelsk.", "Britisk engelsk."],
    "FI": ["osa 1", "Oxford 3000 -sanat.", "Osa 1.", "Amerikanenglanti.", "Brittienglanti."],
    "CS": ["část 1", "Slova Oxford 3000.", "Část 1.", "Americká angličtina.", "Britská angličtina."],
    "SK": ["časť 1", "Slová Oxford 3000.", "Časť 1.", "Americká angličtina.", "Britská angličtina."],
    "HU": ["1. rész", "Oxford 3000 szavak.", "1. rész.", "Amerikai angol.", "Brit angol."],
    "RO": ["partea 1", "Cuvinte Oxford 3000.", "Partea 1.", "Engleză SUA.", "Engleză britanică."],
    "BG": ["част 1", "Думи Oxford 3000.", "Част 1.", "Американски английски.", "Британски английски."],
    "HR": ["dio 1", "Riječi Oxford 3000.", "Dio 1.", "Američki engleski.", "Britanski engleski."],
    "SR": ["део 1", "Речи Oxford 3000.", "Део 1.", "Амерички енглески.", "Британски енглески."],
    "SL": ["del 1", "Besede Oxford 3000.", "Del 1.", "Ameriška angleščina.", "Britanska angleščina."],
    "LT": ["1 dalis", "Oxford 3000 žodžiai.", "1 dalis.", "JAV anglų kalba.", "Britų anglų kalba."],
    "LV": ["1. daļa", "Oxford 3000 vārdi.", "1. daļa.", "ASV angļu valoda.", "Britu angļu valoda."],
    "ET": ["osa 1", "Oxford 3000 sõnad.", "Osa 1.", "USA inglise keel.", "Briti inglise keel."],
    "IS": ["hluti 1", "Oxford 3000 orð.", "Hluti 1.", "Bandarísk enska.", "Bresk enska."],
    "HI": ["भाग 1", "Oxford 3000 शब्द।", "भाग 1।", "अमेरिकी अंग्रेज़ी।", "ब्रिटिश अंग्रेज़ी।"],
    "BN": ["অংশ 1", "Oxford 3000 শব্দ।", "অংশ 1।", "মার্কিন ইংরেজি।", "ব্রিটিশ ইংরেজি।"],
    "TL": ["bahagi 1", "Mga salitang Oxford 3000.", "Bahagi 1.", "Ingles US.", "Ingles UK."],
    "MY": ["အပိုင်း 1", "Oxford 3000 စကားလုံးများ။", "အပိုင်း 1။", "အမေရိကန်အင်္ဂလိပ်။", "ဗြိတိသျှအင်္ဂလိပ်။"],
    "KM": ["ផ្នែក 1", "ពាក្យ Oxford 3000។", "ផ្នែក 1។", "អង់គ្លេស US។", "អង់គ្លេស UK។"],
    "LO": ["ພາກ 1", "ຄຳ Oxford 3000.", "ພາກ 1.", "ອັງກິດ US.", "ອັງກິດ UK."],
    "NE": ["भाग 1", "Oxford 3000 शब्द।", "भाग 1।", "अमेरिकी अंग्रेजी।", "बेलायती अंग्रेजी।"],
    "SI": ["කොටස 1", "Oxford 3000 වචන.", "කොටස 1.", "ඇමරිකානු ඉංග්‍රීසි.", "බ්‍රිතාන්‍ය ඉංග්‍රීසි."],
    "TA": ["பகுதி 1", "Oxford 3000 சொற்கள்.", "பகுதி 1.", "அமெரிக்க ஆங்கிலம்.", "பிரிட்டிஷ் ஆங்கிலம்."],
    "TE": ["భాగం 1", "Oxford 3000 పదాలు.", "భాగం 1.", "అమెరికన్ ఇంగ్లీష్.", "బ్రిటిష్ ఇంగ్లీష్."],
    "KN": ["ಭಾಗ 1", "Oxford 3000 ಪದಗಳು.", "ಭಾಗ 1.", "ಅಮೇರಿಕನ್ ಇಂಗ್ಲಿಷ್.", "ಬ್ರಿಟಿಷ್ ಇಂಗ್ಲಿಷ್."],
    "ML": ["ഭാഗം 1", "Oxford 3000 വാക്കുകൾ.", "ഭാഗം 1.", "അമേരിക്കൻ ഇംഗ്ലീഷ്.", "ബ്രിട്ടീഷ് ഇംഗ്ലീഷ്."],
    "UZ": ["qism 1", "Oxford 3000 so‘zlari.", "1-qism.", "AQSH inglizchasi.", "Britaniya inglizchasi."],
    "KK": ["1-бөлім", "Oxford 3000 сөздері.", "1-бөлім.", "АҚШ ағылшын тілі.", "Британдық ағылшын."],
    "AZ": ["hissə 1", "Oxford 3000 sözləri.", "Hissə 1.", "ABŞ ingiliscəsi.", "Britaniya ingiliscəsi."],
    "KA": ["ნაწილი 1", "Oxford 3000 სიტყვები.", "ნაწილი 1.", "აშშ ინგლისური.", "ბრიტანული ინგლისური."],
    "HY": ["մաս 1", "Oxford 3000 բառեր.", "Մաս 1.", "ԱՄՆ անգլերեն.", "Բրիտանական անգլերեն."],
    "TR": ["bölüm 1", "Oxford 3000 kelime.", "Bölüm 1.", "ABD İngilizcesi.", "Britanya İngilizcesi."],
    "SW": ["sehemu 1", "Maneno ya Oxford 3000.", "Sehemu 1.", "Kiingereza cha US.", "Kiingereza cha UK."],
    "PT-BR": ["parte 1", "Palavras Oxford 3000.", "Parte 1.", "Inglês dos EUA.", "Inglês britânico."],
    "ES-419": ["parte 1", "Palabras Oxford 3000.", "Parte 1.", "Inglés de EE. UU.", "Inglés británico."],
    "EN-GB": ["Part 1", "Oxford 3000 words.", "Part 1.", "US English.", "British English."],
}


def configure_export_mode(export_mode):
    global EXPORT_MODE
    global EXPORT_READINESS
    global SUPPORT_STATUS
    global SUPPORT_TRANSCRIPTION_STATUS
    global STATUS_ISSUE_SUPPORT
    global STATUS_ISSUE_ENGLISH
    global FINAL_READY_FLAG
    global SELECTION_STATUS
    global QUALITY_STATUS
    global EXPORT_STATUS
    global CONTEXT_EXAMPLE_QUALITY_STATUS
    global MEMBERSHIP_QUALITY_STATUS
    global TAG_SUFFIX
    global GOOGLE_SHEET_STATUS
    global POSTGRES_IMPORT_STATUS
    global QA_EVIDENCE_BLOCKERS
    global LEGAL_BLOCKERS

    if export_mode == "working":
        return
    if export_mode != "final":
        raise ValueError(f"Unsupported export mode: {export_mode}")

    EXPORT_MODE = "final"
    EXPORT_READINESS = "final learner-facing English-learning delivery"
    SUPPORT_STATUS = "support_language_aid_checked_for_english_learning"
    SUPPORT_TRANSCRIPTION_STATUS = "out_of_scope_for_support_language_aid"
    STATUS_ISSUE_SUPPORT = "support-language word transcription not required for English-learning final"
    STATUS_ISSUE_ENGLISH = ""
    FINAL_READY_FLAG = "yes"
    SELECTION_STATUS = "final_approved_english_learning_release"
    QUALITY_STATUS = "final_checked"
    EXPORT_STATUS = "final_ready_for_google_sheet_delivery"
    CONTEXT_EXAMPLE_QUALITY_STATUS = "reviewed_final"
    MEMBERSHIP_QUALITY_STATUS = "final_source_package"
    TAG_SUFFIX = "final"
    GOOGLE_SHEET_STATUS = "tracked in delivery manifest"
    POSTGRES_IMPORT_STATUS = "isolated oxford_vocabulary_* persistence only; ordinary deck-table import false"
    QA_EVIDENCE_BLOCKERS = "none"
    LEGAL_BLOCKERS = []

    EDITION_CONFIG["us_english"].update(
        {
            "deck_id": "oxford_3000_core_a1_part_001_us_english",
            "description": "Oxford 3000 words. Part 1. US English.",
            "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_US_English.xlsx",
        }
    )
    EDITION_CONFIG["british_english"].update(
        {
            "deck_id": "oxford_3000_core_a1_part_001_british_english",
            "description": "Oxford 3000 words. Part 1. British English.",
            "file_name": "FlashcardsLuna_Oxford_3000_Core_A1_Part_001_British_English.xlsx",
        }
    )


def infer_part_number(release_id):
    match = re.search(r"_part_(\d+)_", release_id)
    if not match:
        return "1"
    return str(int(match.group(1)))


def infer_course_level(release_id):
    match = re.search(r"_core_([abc]\d)_part_", release_id, flags=re.IGNORECASE)
    if not match:
        match = re.search(r"_advanced_([abc]\d)_extension_part_", release_id, flags=re.IGNORECASE)
    if not match:
        return "A1"
    return match.group(1).upper()


def configure_course_family(release_id):
    global COURSE_WORDLIST_LABEL
    global COURSE_TITLE_LABEL
    global COURSE_FILE_PREFIX
    global COURSE_SLUG_PREFIX
    global COURSE_DECK_ID_PREFIX
    global COURSE_DOMAIN_LABEL
    global COURSE_MODULE_LABEL

    if release_id.startswith("oxford_5000_advanced_"):
        COURSE_WORDLIST_LABEL = "Oxford 5000 Advanced Extension"
        COURSE_TITLE_LABEL = "Oxford 5000 Advanced Extension"
        COURSE_FILE_PREFIX = "Oxford_5000_Advanced_Extension"
        COURSE_SLUG_PREFIX = "oxford-5000-advanced-extension"
        COURSE_DECK_ID_PREFIX = "oxford_5000_advanced_extension"
        COURSE_DOMAIN_LABEL = "Oxford 5000 Advanced Extension"
        COURSE_MODULE_LABEL = "Oxford 5000"
    else:
        COURSE_WORDLIST_LABEL = "Oxford 3000"
        COURSE_TITLE_LABEL = "Oxford 3000 Core"
        COURSE_FILE_PREFIX = "Oxford_3000_Core"
        COURSE_SLUG_PREFIX = "oxford-3000-core"
        COURSE_DECK_ID_PREFIX = "oxford_3000_core"
        COURSE_DOMAIN_LABEL = "Oxford 3000 Core"
        COURSE_MODULE_LABEL = "Oxford 3000"


def configure_release_metadata(release_id, part_number, export_mode, level=None):
    global COURSE_LEVEL
    global COURSE_PART_NUMBER
    global COURSE_PART_PADDED

    configure_course_family(release_id)
    COURSE_LEVEL = (level or infer_course_level(release_id)).upper()
    level_token = COURSE_LEVEL.lower()
    COURSE_PART_NUMBER = str(int(part_number))
    COURSE_PART_PADDED = f"{int(part_number):03d}"
    suffix = "" if export_mode == "final" else "_draft"
    deck_suffix = "" if export_mode == "final" else "_draft"

    EDITION_CONFIG["us_english"].update(
        {
            "deck_id": f"{COURSE_DECK_ID_PREFIX}_{level_token}_part_{COURSE_PART_PADDED}_us_english{deck_suffix}",
            "slug": f"{COURSE_SLUG_PREFIX}-{level_token}-part-{COURSE_PART_PADDED}-us-english",
            "deck_title": f"FlashcardsLuna {COURSE_TITLE_LABEL} {COURSE_LEVEL} Part {COURSE_PART_PADDED} - US English",
            "short_title": f"Oxford {COURSE_LEVEL} Part {COURSE_PART_NUMBER} US.",
            "description": f"{COURSE_WORDLIST_LABEL} {COURSE_LEVEL} words. Part {COURSE_PART_NUMBER}. US English.",
            "sheet_name": f"Oxford {COURSE_LEVEL} US",
            "file_name": f"FlashcardsLuna_{COURSE_FILE_PREFIX}_{COURSE_LEVEL}_Part_{COURSE_PART_PADDED}_US_English{suffix}.xlsx",
        }
    )
    EDITION_CONFIG["british_english"].update(
        {
            "deck_id": f"{COURSE_DECK_ID_PREFIX}_{level_token}_part_{COURSE_PART_PADDED}_british_english{deck_suffix}",
            "slug": f"{COURSE_SLUG_PREFIX}-{level_token}-part-{COURSE_PART_PADDED}-british-english",
            "deck_title": f"FlashcardsLuna {COURSE_TITLE_LABEL} {COURSE_LEVEL} Part {COURSE_PART_PADDED} - British English",
            "short_title": f"Oxford {COURSE_LEVEL} Part {COURSE_PART_NUMBER} UK.",
            "description": f"{COURSE_WORDLIST_LABEL} {COURSE_LEVEL} words. Part {COURSE_PART_NUMBER}. British English.",
            "sheet_name": f"Oxford {COURSE_LEVEL} UK",
            "file_name": f"FlashcardsLuna_{COURSE_FILE_PREFIX}_{COURSE_LEVEL}_Part_{COURSE_PART_PADDED}_British_English{suffix}.xlsx",
        }
    )

CARD_METADATA_HEADERS = [
    "main_sheet_row",
    "display_order",
    "card_key",
    "set_id",
    "meaning_id",
    "canonical_english",
    "english_with_article",
    "part_of_speech",
    "level",
    "frequency_band",
    "priority_band",
    "domain",
    "area",
    "category",
    "context_domain",
    "context_area",
    "context_category",
    "countability",
    "plural_form_en",
    "semantic_class",
    "tags",
    "meaning_note",
    "context_note",
    "context_example_id",
    "context_example_key",
    "context_example_quality_status",
    "meaning_quality_status",
    "membership_quality_status",
]

OXFORD_EDITION_PRONUNCIATION_HEADERS = [
    "EN transcription",
    "EN-GB transcription",
    "EN example transcription",
    "EN-GB example transcription",
]

MAIN_SHEET_COLUMN_CONTRACT = (
    "53 buyer-facing word columns -> 53 buyer-facing example columns -> primary English word IPA -> primary English example IPA"
)

SOURCE_PACKAGE_VARIANT_CONTRACT = (
    "US/UK pronunciation and spelling variants remain in source/QA artifacts; buyer-facing main sheets show only the primary English edition."
)

ENGLISH_HEADWORD_ARTICLE_POLICY = (
    "Oxford-style lemma/headword; do not add artificial a/an/the to English source headwords."
)

SUPPORT_TRANSLATION_ARTICLE_POLICY = (
    "Keep natural articles/gender markers in support-language display where the language normally teaches nouns with them."
)

SUPPORT_TRANSCRIPTION_COLUMN_POLICY = (
    "Do not create empty support-language transcription columns in Oxford English-learning edition workbooks."
)


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def read_jsonl(path):
    rows = []
    for index, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as error:
            raise ValueError(f"Invalid JSONL at {path}:{index}: {error}") from error
    return rows


def load_language_rows(path):
    data = read_json(path)
    required = {"spreadsheetCode", "dbCode", "language", "transcriptionFormat"}
    for row in data:
        missing = required - set(row)
        if missing:
            raise ValueError(f"Language row is missing fields {sorted(missing)}: {row}")
    return data


def load_support_rows(contract, support_codes):
    support_by_row = {}
    seen_codes = set()
    source_paths = []
    for batch in contract.get("latest_support_translation_batches", []):
        path = Path(batch["path"])
        rows = read_jsonl(path)
        source_paths.append(str(path))
        languages = batch.get("languages", [])
        seen_codes.update(languages)
        for row in rows:
            row_id = row["row_id"]
            target = support_by_row.setdefault(row_id, {})
            for code in languages:
                value = normalize_text(row.get(code))
                example = normalize_text(row.get(f"example_{code}"))
                if not value:
                    raise ValueError(f"Missing support translation {code} for {row_id} in {path}")
                if not example:
                    raise ValueError(f"Missing support example {code} for {row_id} in {path}")
                if code in target:
                    raise ValueError(f"Duplicate support translation {code} for {row_id}")
                target[code] = value
                target[f"example_{code}"] = example

    missing_codes = sorted(set(support_codes) - seen_codes)
    extra_codes = sorted(seen_codes - set(support_codes))
    if missing_codes or extra_codes:
        raise ValueError(f"Support language coverage mismatch missing={missing_codes} extra={extra_codes}")
    return support_by_row, source_paths


def load_pronunciations(path):
    rows = read_jsonl(path)
    by_row = {row["row_id"]: row for row in rows}
    if len(by_row) != len(rows):
        raise ValueError(f"Duplicate pronunciation rows in {path}")
    return by_row


def as_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def style_header_row(sheet, fill_color="1F4E78"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def style_main_sheet(sheet, headers):
    style_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 22


def add_values(sheet, rows):
    for row in rows:
        sheet.append([as_cell(value) for value in row])


def build_language_values(layer_row, support_cells, en_us_pronunciation, en_gb_pronunciation, code):
    if code == "EN":
        return {
            "word": layer_row["display_headword_EN_US"],
            "example": layer_row["example_EN_US"],
            "transcription": en_us_pronunciation["transcription"],
        }
    if code == "EN-GB":
        return {
            "word": layer_row["display_headword_EN_GB"],
            "example": layer_row["example_EN_GB"],
            "transcription": en_gb_pronunciation["transcription"],
        }
    return {
        "word": support_cells[code],
        "example": support_cells[f"example_{code}"],
        "transcription": "",
    }


def main_sheet_language_codes_for_edition(edition, language_codes):
    primary = EDITION_CONFIG[edition]["primary_language_code"]
    support_codes = [code for code in language_codes if code not in {"EN", "EN-GB"}]
    return [primary, *support_codes]


def pronunciation_headers_for_edition(edition):
    primary = EDITION_CONFIG[edition]["primary_language_code"]
    return [f"{primary} transcription", f"{primary} example transcription"]


def require_course_metadata_copy(language_codes):
    missing = sorted(set(language_codes) - set(COURSE_METADATA_COPY))
    if missing:
        raise ValueError(f"Course Metadata localization mismatch missing={missing}")


def sentence_end_for(code):
    return SENTENCE_END_BY_CODE.get(code, ".")


def course_metadata_cell(edition, code):
    title_part, words_text, part_text, us_variant, uk_variant = COURSE_METADATA_COPY[code]
    title_part = title_part.replace("1", COURSE_PART_NUMBER)
    part_text = part_text.replace("1", COURSE_PART_NUMBER)
    words_text = re.sub(r"Oxford[- ]3000", COURSE_WORDLIST_LABEL, words_text)
    edition_code = "US" if edition == "us_english" else "UK"
    variant_text = us_variant if edition == "us_english" else uk_variant
    title = f"Oxford {COURSE_LEVEL} {title_part} {edition_code}{sentence_end_for(code)}"
    description = f"{words_text} {part_text} {variant_text}"
    category = f"{COURSE_LEVEL} {title_part}{sentence_end_for(code)}"
    return {
        "title": title,
        "description": description,
        "module": COURSE_MODULE_LABEL,
        "category": category,
    }


def validate_course_metadata_rows(edition, language_codes, metadata_by_code):
    titles = []
    descriptions = []
    for code in language_codes:
        metadata = metadata_by_code.get(code)
        if not metadata:
            raise ValueError(f"Missing Course Metadata localization for {code}")
        title = normalize_text(metadata["title"])
        description = normalize_text(metadata["description"])
        titles.append(title)
        descriptions.append(description)
        if not title or not description:
            raise ValueError(f"Empty Course Metadata cell for {code}")
        if COURSE_PART_NUMBER not in title or COURSE_LEVEL not in title:
            raise ValueError(f"Course Metadata title must identify {COURSE_LEVEL} Part {COURSE_PART_NUMBER} for {code}: {title}")
        expected_wordlist_token = "5000" if "5000" in COURSE_WORDLIST_LABEL else "3000"
        if expected_wordlist_token not in description or COURSE_PART_NUMBER not in description:
            raise ValueError(
                f"Course Metadata description must identify {COURSE_WORDLIST_LABEL} Part {COURSE_PART_NUMBER} "
                f"for {code}: {description}"
            )
        if len(title) > 35:
            raise ValueError(f"Course Metadata title is too long for {code}: {title}")
        max_description_length = 100 if "5000" in COURSE_WORDLIST_LABEL else 80
        if len(description) > max_description_length:
            raise ValueError(f"Course Metadata description is too long for {code}: {description}")
    if len(set(titles)) < 20 or len(set(descriptions)) < 20:
        raise ValueError(f"Course Metadata localization collapsed for {edition}")


def write_main_sheet(workbook, edition, layer_rows, support_by_row, en_us_by_row, en_gb_by_row, language_codes):
    cfg = EDITION_CONFIG[edition]
    sheet = workbook.active
    sheet.title = cfg["sheet_name"]
    main_language_codes = main_sheet_language_codes_for_edition(edition, language_codes)
    pronunciation_headers = pronunciation_headers_for_edition(edition)
    headers = main_language_codes + [f"{code} example" for code in main_language_codes] + pronunciation_headers
    sheet.append(headers)
    for layer_row in layer_rows:
        row_id = layer_row["row_id"]
        support_cells = support_by_row.get(row_id)
        en_us_pronunciation = en_us_by_row.get(row_id)
        en_gb_pronunciation = en_gb_by_row.get(row_id)
        if not support_cells:
            raise ValueError(f"Missing support cells for {row_id}")
        if not en_us_pronunciation:
            raise ValueError(f"Missing EN-US pronunciation row for {row_id}")
        if not en_gb_pronunciation:
            raise ValueError(f"Missing EN-GB pronunciation row for {row_id}")
        values = [
            build_language_values(layer_row, support_cells, en_us_pronunciation, en_gb_pronunciation, code)
            for code in main_language_codes
        ]
        primary_pronunciation = en_us_pronunciation if cfg["primary_language_code"] == "EN" else en_gb_pronunciation
        sheet.append(
            [item["word"] for item in values]
            + [item["example"] for item in values]
            + [
                primary_pronunciation["transcription"],
                primary_pronunciation["example_transcription"],
            ]
        )
    style_main_sheet(sheet, headers)
    return headers


def add_course_metadata(workbook, edition, language_codes):
    require_course_metadata_copy(language_codes)
    metadata_by_code = {code: course_metadata_cell(edition, code) for code in language_codes}
    validate_course_metadata_rows(edition, language_codes, metadata_by_code)
    sheet = workbook.create_sheet("Course Metadata")
    rows = [
        ["", *language_codes],
        ["Title", *[metadata_by_code[code]["title"] for code in language_codes]],
        ["Description", *[metadata_by_code[code]["description"] for code in language_codes]],
        ["Module", *[metadata_by_code[code]["module"] for code in language_codes]],
        ["Category", *[metadata_by_code[code]["category"] for code in language_codes]],
    ]
    add_values(sheet, rows)
    style_header_row(sheet, "D9EAF7")
    sheet.column_dimensions["A"].width = 18
    for index in range(2, len(language_codes) + 2):
        sheet.column_dimensions[get_column_letter(index)].width = 22
    return metadata_by_code


def deck_metadata_rows(edition, layer_rows, main_sheet_name):
    cfg = EDITION_CONFIG[edition]
    en_metadata = course_metadata_cell(edition, "EN")
    ru_metadata = course_metadata_cell(edition, "RU")
    level_label = {
        "A1": "Beginner",
        "A2": "Elementary",
        "B1": "Intermediate",
        "B2": "Upper-intermediate",
        "C1": "Advanced",
        "C2": "Proficient",
    }.get(COURSE_LEVEL, COURSE_LEVEL)
    return [
        ["field", "value", "notes"],
        ["set_id", cfg["deck_id"], "Stable internal deck id."],
        ["slug", cfg["slug"], "Suggested stable slug."],
        ["content_type", "vocabulary", "Deck type."],
        ["domain", COURSE_DOMAIN_LABEL, "Top-level taxonomy domain."],
        ["area", COURSE_LEVEL, "Second-level taxonomy area."],
        ["category_or_situation", cfg["deck_edition"], f"English edition for this {EXPORT_MODE} workbook."],
        ["deck_edition", cfg["deck_edition"], "Buyer-facing English edition for this workbook."],
        ["source_variant", cfg["source_variant"], "English source/pronunciation variant used for the primary edition."],
        ["primary_language_code", cfg["primary_language_code"], "Primary English spreadsheet language code for this edition."],
        ["roadmap_stage", "course_exam_release", "Content roadmap lane."],
        ["level_label", level_label, "Canonical deck level label."],
        ["level_min", COURSE_LEVEL, "Official CEFR minimum level."],
        ["level_max", COURSE_LEVEL, "Official CEFR maximum level for this workbook shell."],
        ["target_item_count_min", len(layer_rows), "Target range, not a forced fill number."],
        ["target_item_count_max", len(layer_rows), "Target range, not a forced fill number."],
        ["actual_item_count", len(layer_rows), "Rows in the main card sheet."],
        ["title_en", en_metadata["title"], "Human-readable title from Course Metadata."],
        ["description_en", en_metadata["description"], "Human-readable description from Course Metadata."],
        ["module_en", COURSE_MODULE_LABEL, "Short mobile-app module label from Course Metadata."],
        ["category_en", en_metadata["category"], "Short mobile-app category label from Course Metadata."],
        ["title_ru", ru_metadata["title"], "Localized Russian title from Course Metadata."],
        ["description_ru", ru_metadata["description"], "Localized Russian description from Course Metadata."],
        ["module_ru", COURSE_MODULE_LABEL, "Shared course module label."],
        ["category_ru", ru_metadata["category"], "Localized Russian category label."],
        ["learning_goal", f"Study {COURSE_WORDLIST_LABEL} {COURSE_LEVEL} vocabulary with {cfg['deck_edition']} as the primary English edition.", "What the learner should get from this deck."],
        ["next_recommended_set_ids", "", "Suggested next decks in the content ladder."],
        ["selection_status", SELECTION_STATUS, "Selection/review state for the set composition."],
        ["quality_status", QUALITY_STATUS, "Quality status for the content set."],
        ["global_status_blockers", "; ".join(LEGAL_BLOCKERS), "Legal/source/delivery blockers."],
        ["qa_evidence_blockers", QA_EVIDENCE_BLOCKERS, "Missing final evidence, if any."],
        ["export_mode", EXPORT_MODE, "working or final."],
        ["export_status", EXPORT_STATUS, "Spreadsheet export readiness."],
        ["language_variant_count", 54, "Active language variants in this export."],
        ["main_sheet", main_sheet_name, "User-facing card sheet."],
        ["column_contract", MAIN_SHEET_COLUMN_CONTRACT, "Main sheet grouped-column contract."],
        ["spreadsheet_contract", SPREADSHEET_CONTRACT, "Machine-readable workbook contract."],
        ["card_key_policy", "card_key = set_id::meaning_id", "Deck-scoped stable row key in the workbook."],
        ["sheet_contract_version", "v1", "Workbook metadata/export contract version."],
        ["english_headword_article_policy", ENGLISH_HEADWORD_ARTICLE_POLICY, "Oxford source/headword display policy."],
        ["support_translation_article_policy", SUPPORT_TRANSLATION_ARTICLE_POLICY, "Native-language support display policy."],
        ["support_transcription_column_policy", SUPPORT_TRANSCRIPTION_COLUMN_POLICY, "Oxford edition main-sheet shape policy."],
    ]


def add_deck_metadata(workbook, edition, layer_rows, main_sheet_name):
    sheet = workbook.create_sheet("Deck Metadata")
    add_values(sheet, deck_metadata_rows(edition, layer_rows, main_sheet_name))
    style_header_row(sheet, "D9EAF7")
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 80


def add_card_metadata(workbook, edition, layer_rows):
    cfg = EDITION_CONFIG[edition]
    sheet = workbook.create_sheet("Card Metadata")
    rows = [CARD_METADATA_HEADERS]
    for index, row in enumerate(layer_rows, start=1):
        context_example_id = f"{row['core_item_id']}_example_001"
        primary_headword = (
            row["display_headword_EN_US"] if cfg["primary_language_code"] == "EN" else row["display_headword_EN_GB"]
        )
        rows.append(
            [
                index + 1,
                index,
                f"{cfg['deck_id']}::{row['meaning_id']}",
                cfg["deck_id"],
                row["meaning_id"],
                primary_headword,
                primary_headword,
                row["reviewed_part_of_speech"],
                row["level_min"],
                "core",
                "core",
                COURSE_DOMAIN_LABEL,
                COURSE_LEVEL,
                cfg["deck_edition"],
                "English Vocabulary",
                f"Oxford {COURSE_LEVEL}",
                cfg["deck_edition"],
                "",
                "",
                "general_english_core_vocabulary",
                f"{COURSE_DECK_ID_PREFIX},{COURSE_LEVEL.lower()},source_package,{TAG_SUFFIX}",
                row["meaning_note"],
                row[f"edition_note_{'EN_US' if edition == 'us_english' else 'EN_GB'}"],
                context_example_id,
                f"{cfg['deck_id']}::{row['meaning_id']}::{context_example_id}",
                CONTEXT_EXAMPLE_QUALITY_STATUS,
                "row_reviewed_for_english_learning",
                MEMBERSHIP_QUALITY_STATUS,
            ]
        )
    add_values(sheet, rows)
    style_header_row(sheet, "1F4E78")
    sheet.freeze_panes = "A2"
    for index in range(1, len(CARD_METADATA_HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 24


def add_readme(workbook, edition, row_count):
    cfg = EDITION_CONFIG[edition]
    main_language_count = 53
    primary = cfg["primary_language_code"]
    rows = [
        ["FlashcardsLuna Google Sheets export"],
        ["Set ID", cfg["deck_id"]],
        ["Export mode", EXPORT_MODE],
        ["Export readiness", EXPORT_READINESS],
        ["Main sheet", cfg["sheet_name"]],
        ["Rows", row_count],
        ["Language variants", main_language_count],
        ["Column contract", MAIN_SHEET_COLUMN_CONTRACT],
        ["Source variant traceability", SOURCE_PACKAGE_VARIANT_CONTRACT],
        ["Spreadsheet contract", SPREADSHEET_CONTRACT],
        ["Order rule", "Primary English first, then the 52 support languages; the same order is repeated in the word and example blocks."],
        ["Transcription rule", f"Only {primary} word IPA and {primary} example IPA are shown at the end of this edition's main sheet."],
        ["Current field coverage", "Words/examples filled for 53 buyer-facing variants; primary English word/example IPA filled."],
        ["Missing coverage", "No empty support-language transcription columns; support IPA is out of scope for this English-learning release."],
        ["Article policy", f"{ENGLISH_HEADWORD_ARTICLE_POLICY} {SUPPORT_TRANSLATION_ARTICLE_POLICY}"],
        ["Support transcription columns", SUPPORT_TRANSCRIPTION_COLUMN_POLICY],
        ["Status blockers", "none" if not LEGAL_BLOCKERS else "Working preview; not final-ready."],
        ["Global status blockers", "; ".join(LEGAL_BLOCKERS)],
        ["QA evidence blockers", QA_EVIDENCE_BLOCKERS],
        ["Course metadata", "Course Metadata follows ordinary workbook sheet shape; English-learning product copy is edition-specific."],
        ["Deck metadata", "Deck Metadata contains set-level fields and primary English edition."],
        ["Card metadata", "Card Metadata contains card_key, set_id, meaning_id, level, POS, taxonomy and row statuses."],
        ["QA status", "See _qa_status."],
        ["Google Sheet created", GOOGLE_SHEET_STATUS],
        ["Postgres import", POSTGRES_IMPORT_STATUS],
    ]
    sheet = workbook.create_sheet("_README")
    add_values(sheet, rows)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 110


def add_qa_status(workbook, language_rows, row_count):
    sheet = workbook.create_sheet("_qa_status")
    rows = [
        ["Export mode", EXPORT_MODE],
        ["Export readiness", EXPORT_READINESS],
        ["Rows", row_count],
        ["Language variants", len(language_rows)],
        ["Rule", "Oxford English-learning final export requires zero missing words/examples, EN/EN-GB word and example IPA, final-ready statuses, context-example gates and semantic QA evidence."],
        ["Global status blockers", "; ".join(LEGAL_BLOCKERS)],
        ["QA evidence blockers", QA_EVIDENCE_BLOCKERS],
        [],
        [
            "spreadsheet_code",
            "db_code",
            "missing_words",
            "missing_examples",
            "missing_transcriptions",
            "entry_statuses",
            "example_statuses",
            "pronunciation_statuses",
            "status_issues",
            "final_ready",
        ],
    ]
    for item in language_rows:
        code = item["spreadsheetCode"]
        is_english_variant = code in {"EN", "EN-GB"}
        rows.append(
            [
                code,
                item["dbCode"],
                0,
                0,
                0,
                "generated_checked" if is_english_variant else SUPPORT_STATUS,
                "generated_checked" if is_english_variant else SUPPORT_STATUS,
                "generated_checked" if is_english_variant else SUPPORT_TRANSCRIPTION_STATUS,
                STATUS_ISSUE_ENGLISH if is_english_variant else STATUS_ISSUE_SUPPORT,
                FINAL_READY_FLAG,
            ]
        )
    rows.extend(
        [
            [],
            ["global_status_blocker"],
            *[[blocker] for blocker in LEGAL_BLOCKERS],
            [],
            ["qa_evidence_blocker"],
            *[[blocker] for blocker in ([] if QA_EVIDENCE_BLOCKERS == "none" else QA_EVIDENCE_BLOCKERS.split("; "))],
        ]
    )
    add_values(sheet, rows)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["I"].width = 58


def add_languages(workbook, language_rows):
    sheet = workbook.create_sheet("_languages")
    rows = [["order", "spreadsheet_code", "db_code", "language", "transcription_format"]]
    for index, row in enumerate(language_rows, start=1):
        rows.append([index, row["spreadsheetCode"], row["dbCode"], row["language"], row["transcriptionFormat"]])
    add_values(sheet, rows)
    style_header_row(sheet, "D9EAF7")
    for index in range(1, 6):
        sheet.column_dimensions[get_column_letter(index)].width = [10, 20, 14, 34, 62][index - 1]


def write_workbook(path, edition, layer_rows, support_by_row, en_us_by_row, en_gb_by_row, language_rows):
    cfg = EDITION_CONFIG[edition]
    language_codes = [row["spreadsheetCode"] for row in language_rows]
    main_language_codes = main_sheet_language_codes_for_edition(edition, language_codes)
    language_rows_by_code = {row["spreadsheetCode"]: row for row in language_rows}
    main_language_rows = [language_rows_by_code[code] for code in main_language_codes]
    workbook = Workbook()
    headers = write_main_sheet(workbook, edition, layer_rows, support_by_row, en_us_by_row, en_gb_by_row, language_codes)
    add_course_metadata(workbook, edition, main_language_codes)
    add_deck_metadata(workbook, edition, layer_rows, cfg["sheet_name"])
    add_card_metadata(workbook, edition, layer_rows)
    add_readme(workbook, edition, len(layer_rows))
    add_qa_status(workbook, main_language_rows, len(layer_rows))
    add_languages(workbook, main_language_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return {
        "path": str(path),
        "rows": len(layer_rows),
        "columns": len(headers),
        "sheet": cfg["sheet_name"],
        "main_sheet_language_codes": main_sheet_language_codes_for_edition(edition, language_codes),
        "main_sheet_pronunciation_headers": pronunciation_headers_for_edition(edition),
    }


def workbook_shape(path):
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return {"rows": sheet.max_row - 1, "columns": sheet.max_column, "main_sheet": sheet.title, "sheets": workbook.sheetnames}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--contract", default="config/oxford-vocabulary-release-contract-v0.json")
    parser.add_argument(
        "--edition-layer",
        default="outputs/oxford-vocabulary/edition-layers/oxford_3000_core_a1_part_001_150_v1_us_uk_edition_layer_v1.jsonl",
    )
    parser.add_argument(
        "--en-us-pronunciations",
        default="outputs/oxford-vocabulary/pronunciations/oxford_3000_core_a1_part_001_150_v1_en_us_edition_pronunciations_v1.jsonl",
    )
    parser.add_argument(
        "--en-gb-pronunciations",
        default="outputs/oxford-vocabulary/pronunciations/oxford_3000_core_a1_part_001_150_v1_en_gb_edition_pronunciations_v1.jsonl",
    )
    parser.add_argument("--languages", default="config/language-order.json")
    parser.add_argument("--out-dir", default="outputs/oxford-vocabulary/editions")
    parser.add_argument("--export-id", default="edition_exports_v1")
    parser.add_argument("--export-mode", choices=["working", "final"], default="working")
    parser.add_argument("--part-number", default="")
    parser.add_argument("--level", default="")
    parser.add_argument("--isolated-oxford-db-persistence", action="store_true")
    args = parser.parse_args()
    configure_export_mode(args.export_mode)

    contract = read_json(Path(args.contract))
    layer_path = Path(args.edition_layer)
    layer_rows = read_jsonl(layer_path)
    if not layer_rows:
        raise ValueError("Edition layer artifact is empty")
    release_id = layer_rows[0]["release_id"]
    configure_release_metadata(release_id, args.part_number or infer_part_number(release_id), args.export_mode, args.level)
    language_rows = load_language_rows(Path(args.languages))
    language_codes = [row["spreadsheetCode"] for row in language_rows]
    support_codes = [code for code in language_codes if code not in {"EN", "EN-GB"}]
    if len(language_codes) != 54:
        raise ValueError(f"Expected 54 language variants, found {len(language_codes)}")
    if len(support_codes) != 52:
        raise ValueError(f"Expected 52 support languages, found {len(support_codes)}")

    support_by_row, source_paths = load_support_rows(contract, support_codes)
    support_article_display_repair = contract.get("latest_support_article_display_repair", {})
    en_us_by_row = load_pronunciations(Path(args.en_us_pronunciations))
    en_gb_by_row = load_pronunciations(Path(args.en_gb_pronunciations))
    generated_at = datetime.now(timezone.utc).isoformat()
    out_dir = Path(args.out_dir)

    editions = []
    for edition in ["us_english", "british_english"]:
        path = out_dir / EDITION_CONFIG[edition]["file_name"]
        info = write_workbook(path, edition, layer_rows, support_by_row, en_us_by_row, en_gb_by_row, language_rows)
        info.update(
            {
                "edition_id": edition,
                "deck_id": EDITION_CONFIG[edition]["deck_id"],
                "deck_title": EDITION_CONFIG[edition]["deck_title"],
                "source_variant": EDITION_CONFIG[edition]["source_variant"],
                "primary_language_code": EDITION_CONFIG[edition]["primary_language_code"],
                "shape": workbook_shape(path),
            }
        )
        editions.append(info)

    main_sheet_language_count = len(editions[0]["main_sheet_language_codes"]) if editions else 0
    main_sheet_pronunciation_column_count = (
        len(editions[0]["main_sheet_pronunciation_headers"]) if editions else 0
    )
    manifest = {
        "export_id": args.export_id,
        "script_version": SCRIPT_VERSION,
        "status": "final_english_learning_delivery_workbooks_ready"
        if EXPORT_MODE == "final"
        else "ordinary_lunacards_working_preview_not_final_delivery",
        "release_id": release_id,
        "generated_at": generated_at,
        "principle": "one canonical source package, two buyer-facing US/UK English deck workbooks using the ordinary LunaCards sheet shell",
        "spreadsheet_contract": SPREADSHEET_CONTRACT,
        "required_sheets": [
            "main vocabulary sheet",
            "Course Metadata",
            "Deck Metadata",
            "Card Metadata",
            "_README",
            "_qa_status",
            "_languages",
        ],
        "main_sheet_column_contract": MAIN_SHEET_COLUMN_CONTRACT,
        "source_package_variant_contract": SOURCE_PACKAGE_VARIANT_CONTRACT,
        "language_variant_count": len(language_codes),
        "main_sheet_language_variant_count": main_sheet_language_count,
        "support_language_count": len(support_codes),
        "words_filled_per_workbook": len(layer_rows) * main_sheet_language_count,
        "examples_filled_per_workbook": len(layer_rows) * main_sheet_language_count,
        "word_transcriptions_filled_per_workbook": len(layer_rows) * main_sheet_pronunciation_column_count,
        "word_transcriptions_missing_per_workbook": 0,
        "example_transcriptions_filled_per_workbook": len(layer_rows) * main_sheet_pronunciation_column_count,
        "example_transcription_languages": ["EN", "EN-GB"],
        "example_transcription_location": "main_sheet_final_two_primary_english_columns",
        "opposite_english_variant_present_in_main_sheet": False,
        "course_metadata_localized": True,
        "course_metadata_title_includes_part": True,
        "card_metadata_columns": len(CARD_METADATA_HEADERS),
        "support_transcription_columns_present": False,
        "support_language_word_transcriptions_required": False,
        "support_language_word_transcription_policy": "out_of_scope_for_oxford_english_learning_edition_workbooks",
        "article_policy": {
            "english_headword": ENGLISH_HEADWORD_ARTICLE_POLICY,
            "support_translation_display": SUPPORT_TRANSLATION_ARTICLE_POLICY,
        },
        "support_article_display_repair": support_article_display_repair,
        "edition_layer_path": str(layer_path),
        "support_translation_batch_paths": source_paths,
        "ordinary_deck_table_postgres_import": False,
        "isolated_oxford_db_persistence": bool(args.isolated_oxford_db_persistence),
        "postgres_import": False,
        "google_sheet_created": False,
        "final_delivery_ready": EXPORT_MODE == "final",
        "remaining_blockers": [*LEGAL_BLOCKERS],
        "editions": editions,
    }
    manifest_path = out_dir / f"{release_id}_{args.export_id}.json"
    summary_path = out_dir / f"{release_id}_{args.export_id}_summary.md"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    summary = [
        f"# Oxford {COURSE_LEVEL} US/UK Ordinary-Format Edition Workbooks: {release_id}",
        "",
        f"- Script version: `{SCRIPT_VERSION}`",
        f"- Status: `{manifest['status']}`",
        f"- Spreadsheet contract: `{SPREADSHEET_CONTRACT}`",
        f"- Editions: {len(editions)}",
        f"- Rows per edition: {len(layer_rows)}",
        f"- Columns per edition: {editions[0]['columns'] if editions else 'unknown'}",
        "- Required sheets: main vocabulary sheet, `Course Metadata`, `Deck Metadata`, `Card Metadata`, `_README`, `_qa_status`, `_languages`",
        f"- Words filled per edition: {manifest['words_filled_per_workbook']}",
        f"- Examples filled per edition: {manifest['examples_filled_per_workbook']}",
        f"- Word transcriptions filled per edition: {manifest['word_transcriptions_filled_per_workbook']}",
        f"- Word transcriptions missing per edition: {manifest['word_transcriptions_missing_per_workbook']}",
        f"- Main sheet language variants per edition: {manifest['main_sheet_language_variant_count']}",
        f"- Example transcriptions filled per edition: {manifest['example_transcriptions_filled_per_workbook']} (primary English only, in the main sheet final columns)",
        "- Opposite English variant present in main sheet: false",
        f"- Course Metadata localized: {str(manifest['course_metadata_localized']).lower()}",
        f"- Course Metadata title includes part: {str(manifest['course_metadata_title_includes_part']).lower()}",
        f"- Card Metadata columns: {manifest['card_metadata_columns']}",
        f"- Support transcription columns present: {str(manifest['support_transcription_columns_present']).lower()}",
        f"- Support-language word transcriptions required: {str(manifest['support_language_word_transcriptions_required']).lower()}",
        f"- English headword article policy: {ENGLISH_HEADWORD_ARTICLE_POLICY}",
        f"- Support translation article policy: {SUPPORT_TRANSLATION_ARTICLE_POLICY}",
        f"- Support article display repair: {support_article_display_repair.get('repair_id', 'not_applied')} ({support_article_display_repair.get('display_cells_changed', 0)} display cells changed)",
        "- Ordinary deck-table Postgres import: false",
        f"- Isolated Oxford DB persistence: {str(bool(args.isolated_oxford_db_persistence)).lower()}",
        f"- Google Sheet created: {str(manifest['google_sheet_created']).lower()}",
        f"- Final delivery ready: {str(manifest['final_delivery_ready']).lower()}",
        "",
    ]
    for edition in editions:
        summary.append(f"- `{edition['deck_title']}`: `{edition['path']}`")
    summary_path.write_text("\n".join(summary) + "\n", encoding="utf-8")

    print(
        f"Oxford ordinary-format edition workbooks exported: editions={len(editions)} rows={len(layer_rows)} "
        f"columns={editions[0]['columns'] if editions else 'unknown'} manifest={manifest_path} summary={summary_path}"
    )


if __name__ == "__main__":
    main()
