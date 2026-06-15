#!/usr/bin/env python3
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_VERSION = "2026-05-21.v1"
RELEASE_ID = "oxford_3000_core_a2_part_002_300_v1"
OLD_SOURCE_CANDIDATE_ID = "oxford_3000::1526::adj_a1"
NEW_SOURCE_CANDIDATE_ID = "oxford_3000::1526::light"
ROW_ID = f"{RELEASE_ID}::111"

SOURCE_JSONL = Path(f"outputs/oxford-vocabulary/source/{RELEASE_ID}.oxford_3000.normalized.jsonl")
SOURCE_CSV = Path(f"outputs/oxford-vocabulary/source/{RELEASE_ID}.oxford_3000.normalized.csv")
CANDIDATE_POOL = Path(f"outputs/oxford-vocabulary/candidate-pools/{RELEASE_ID}_candidate_pool.jsonl")
CANDIDATE_SUMMARY = Path(f"outputs/oxford-vocabulary/candidate-pools/{RELEASE_ID}_candidate_pool_summary.md")
ROW_REVIEW = Path(f"outputs/oxford-vocabulary/row-reviews/{RELEASE_ID}_row_review_v1.jsonl")
ROW_REVIEW_SUMMARY = Path(f"outputs/oxford-vocabulary/row-reviews/{RELEASE_ID}_row_review_v1_summary.md")
SOURCE_MANIFEST = Path(f"outputs/oxford-vocabulary/source/{RELEASE_ID}_source_snapshot_manifest.json")
SOURCE_DRAFT_WORKBOOK = Path("outputs/oxford-vocabulary/final/FlashcardsLuna_Oxford_3000_Core_A2_Part_002_source_draft.xlsx")
CONTRACT = Path(f"config/{RELEASE_ID}_contract_v0.json")

REPAIR = {
    "repair_id": "a2_part002_light_source_row_repair_v1",
    "script_path": "scripts/oxford/apply-oxford-a2-part002-light-source-repair.py",
    "script_version": SCRIPT_VERSION,
    "release_id": RELEASE_ID,
    "row_id": ROW_ID,
    "old_source_candidate_id": OLD_SOURCE_CANDIDATE_ID,
    "new_source_candidate_id": NEW_SOURCE_CANDIDATE_ID,
    "old_source_headword": "adj. A1,",
    "new_source_headword": "light",
    "old_raw_line": "adj. A1, v. A2",
    "new_raw_line": "light adj. A1, v. A2",
    "reason": "The PDF-derived source parser split the light row and promoted the level/POS fragment 'adj. A1,' to a fake headword.",
    "postgres_changes": False,
    "google_sheet_created": False,
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def replace_jsonl_row(path: Path, predicate, mutator) -> dict:
    lines = path.read_text(encoding="utf-8").splitlines()
    replaced = None
    next_lines: list[str] = []
    for line in lines:
        row = json.loads(line)
        if predicate(row):
            if replaced is not None:
                raise RuntimeError(f"Multiple matching rows in {path}")
            replaced = mutator(row)
            next_lines.append(json.dumps(replaced, ensure_ascii=False, separators=(",", ":")))
        else:
            next_lines.append(line)
    if replaced is None:
        raise RuntimeError(f"No matching row in {path}")
    path.write_text("\n".join(next_lines) + "\n", encoding="utf-8")
    return replaced


def replace_csv_row(path: Path) -> None:
    lines = path.read_text(encoding="utf-8").splitlines()
    if not lines:
        raise RuntimeError(f"Empty CSV: {path}")
    replaced = 0
    out_lines = [lines[0]]
    for line in lines[1:]:
        row = next(csv.reader([line]))
        if row and row[0] in {OLD_SOURCE_CANDIDATE_ID, NEW_SOURCE_CANDIDATE_ID}:
            row[0] = NEW_SOURCE_CANDIDATE_ID
            row[2] = "light"
            row[3] = "v."
            row[8] = "user_reported_oup_permission_pending_written_evidence"
            row[9] = "light adj. A1, v. A2"
            buffer = io.StringIO()
            csv.writer(buffer, lineterminator="").writerow(row)
            out_lines.append(buffer.getvalue())
            replaced += 1
        else:
            out_lines.append(line)
    if replaced != 1:
        raise RuntimeError(f"Expected one CSV replacement in {path}, got {replaced}")
    path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")


def repair_source_row(row: dict) -> dict:
    if row.get("source_headword") not in {"adj. A1,", "light"}:
        raise RuntimeError(f"Unexpected source row before repair: {row}")
    if row.get("raw_line") not in {"adj. A1, v. A2", "light adj. A1, v. A2"}:
        raise RuntimeError(f"Unexpected source row before repair: {row}")
    row.update(
        {
            "source_candidate_id": NEW_SOURCE_CANDIDATE_ID,
            "source_headword": "light",
            "normalized_headword": "light",
            "part_of_speech": "v.",
            "raw_line": "light adj. A1, v. A2",
            "source_repair": REPAIR,
        }
    )
    return row


def repair_candidate_row(row: dict) -> dict:
    if row.get("row_id") != ROW_ID or row.get("source_headword") not in {"adj. A1,", "light"}:
        raise RuntimeError(f"Unexpected candidate row before repair: {row}")
    row.update(
        {
            "meaning_id": "oxford3000_light_needs_review_01",
            "source_candidate_id": NEW_SOURCE_CANDIDATE_ID,
            "source_headword": "light",
            "normalized_headword": "light",
            "part_of_speech": "v.",
            "meaning_note": "Needs row review: Oxford 3000 source word 'light' in the A2 verb sense.",
            "source_repair": REPAIR,
        }
    )
    row["semantic_scene"] = {
        **(row.get("semantic_scene") or {}),
        "target_display": "light",
        "source_repair_status": "repaired_pdf_fragment_headword",
    }
    row["EN"] = "light"
    row["example_EN"] = ""
    return row


def repair_review_row(row: dict) -> dict:
    if row.get("row_id") != ROW_ID or row.get("source_headword") not in {"adj. A1,", "light"}:
        raise RuntimeError(f"Unexpected row-review row before repair: {row}")
    row.update(
        {
            "source_candidate_id": NEW_SOURCE_CANDIDATE_ID,
            "source_headword": "light",
            "reviewed_display_headword": "light",
            "original_part_of_speech": "v.",
            "reviewed_part_of_speech": "verb",
            "meaning_id": "oxford3000_light_common_a2_action_01",
            "meaning_note": 'The A2 verb sense of "light" used for starting a candle, lamp or fire.',
            "learner_value_note": 'Useful A2 English verb for learner communication: "light".',
            "source_repair": REPAIR,
        }
    )
    row["semantic_scene"] = {
        **(row.get("semantic_scene") or {}),
        "target_display": "light",
        "learner_sense": 'The A2 verb sense of "light" used for starting a candle, lamp or fire.',
        "source_headword": "light",
        "reviewed_part_of_speech": "verb",
        "source_repair_status": "repaired_pdf_fragment_headword",
    }
    return row


def repair_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook["Oxford 3000 Core"]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    row_number = None
    for index in range(2, sheet.max_row + 1):
        if sheet.cell(row=index, column=headers["row_id"]).value == ROW_ID:
            row_number = index
            break
    if row_number is None:
        raise RuntimeError(f"Workbook row not found: {ROW_ID}")
    if sheet.cell(row=row_number, column=headers["source_headword"]).value not in {"adj. A1,", "light"}:
        raise RuntimeError("Workbook source row was already changed or is unexpected")
    sheet.cell(row=row_number, column=headers["meaning_id"]).value = "oxford3000_light_needs_review_01"
    sheet.cell(row=row_number, column=headers["source_headword"]).value = "light"
    sheet.cell(row=row_number, column=headers["part_of_speech"]).value = "v."
    sheet.cell(row=row_number, column=headers["meaning_note"]).value = (
        "Needs row review: Oxford 3000 source word 'light' in the A2 verb sense."
    )
    sheet.cell(row=row_number, column=headers["semantic_scene"]).value = json.dumps(
        {
            "rule_version": "oxford-source-package-v0",
            "status": "needs_review",
            "target_display": "light",
            "topic_context": "general_english_core_vocabulary",
            "source_repair_status": "repaired_pdf_fragment_headword",
        },
        ensure_ascii=False,
    )
    sheet.cell(row=row_number, column=headers["EN"]).value = "light"
    sheet.cell(row=row_number, column=headers["example_EN"]).value = None
    sheet.cell(row=row_number, column=headers["qa_notes"]).value = (
        "Oxford source row selected; source fragment repaired from 'adj. A1,' to light verb sense; translations/examples/QA not generated yet."
    )
    workbook.save(path)


def append_summary_note(path: Path, note: str) -> None:
    text = path.read_text(encoding="utf-8")
    marker = f"- Source repair: {note}"
    if marker not in text:
        text = text.rstrip() + "\n" + marker + "\n"
        path.write_text(text, encoding="utf-8")


def update_manifest_and_contract(generated_at: str) -> None:
    repair = {**REPAIR, "generated_at": generated_at}
    manifest = read_json(SOURCE_MANIFEST)
    manifest_repairs = manifest.get("source_repairs") or []
    manifest["source_repairs"] = [item for item in manifest_repairs if item.get("repair_id") != repair["repair_id"]] + [repair]
    manifest["updated_at"] = generated_at
    write_json(SOURCE_MANIFEST, manifest)

    contract = read_json(CONTRACT)
    source_snapshot = contract.get("latest_source_snapshot") or {}
    source_repairs = source_snapshot.get("source_repairs") or []
    source_snapshot["source_repairs"] = [
        item for item in source_repairs if item.get("repair_id") != repair["repair_id"]
    ] + [repair]
    source_snapshot["status"] = "source_snapshot_ready_with_light_source_row_repair_not_delivery_ready"
    contract["latest_source_snapshot"] = source_snapshot
    contract["source_repairs"] = [
        item for item in contract.get("source_repairs", []) if item.get("repair_id") != repair["repair_id"]
    ] + [repair]
    contract["updated_at"] = generated_at
    write_json(CONTRACT, contract)


def main() -> None:
    generated_at = now_iso()
    replace_jsonl_row(
        SOURCE_JSONL,
        lambda row: row.get("source_candidate_id") in {OLD_SOURCE_CANDIDATE_ID, NEW_SOURCE_CANDIDATE_ID}
        and row.get("oxford_order") == 1526,
        repair_source_row,
    )
    replace_csv_row(SOURCE_CSV)
    replace_jsonl_row(CANDIDATE_POOL, lambda row: row.get("row_id") == ROW_ID, repair_candidate_row)
    replace_jsonl_row(ROW_REVIEW, lambda row: row.get("row_id") == ROW_ID, repair_review_row)
    repair_workbook(SOURCE_DRAFT_WORKBOOK)
    append_summary_note(
        CANDIDATE_SUMMARY,
        "`a2_part002_light_source_row_repair_v1` corrected the parsed source fragment `adj. A1,` to `light` for the A2 verb sense; no Google Sheet or DB import.",
    )
    append_summary_note(
        ROW_REVIEW_SUMMARY,
        "`a2_part002_light_source_row_repair_v1` corrected row 111 from parsed fragment `adj. A1,` to reviewed A2 verb `light`; no Google Sheet or DB import.",
    )
    update_manifest_and_contract(generated_at)
    print(
        json.dumps(
            {
                "release_id": RELEASE_ID,
                "repair_id": REPAIR["repair_id"],
                "row_id": ROW_ID,
                "old_source_headword": "adj. A1,",
                "new_source_headword": "light",
                "artifacts_updated": [
                    str(SOURCE_JSONL),
                    str(SOURCE_CSV),
                    str(CANDIDATE_POOL),
                    str(ROW_REVIEW),
                    str(SOURCE_DRAFT_WORKBOOK),
                    str(SOURCE_MANIFEST),
                    str(CONTRACT),
                ],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
